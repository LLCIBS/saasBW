#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Веб-интерфейс для настройки Call Analyzer
Позволяет управлять всеми настройками системы без редактирования кода
"""

import os
import sys
import json
import shutil
import yaml
import logging
import re
import importlib
import importlib.util
import requests
from datetime import datetime, timedelta
from pathlib import Path
from collections import OrderedDict
from flask import Flask, render_template, request, jsonify, redirect, url_for, flash
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
from flask_login import current_user, login_required
from copy import deepcopy
from contextlib import contextmanager
import threading
import subprocess
import time
from sqlalchemy import text

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'call_analyzer'))
load_dotenv(PROJECT_ROOT / '.env')

from common.user_settings import (
    default_config_template,
    default_prompts_template,
    default_vocabulary_template,
    default_logs_template,
    default_script_prompt_template,
    build_runtime_config
)

from config.settings import get_config
from database.models import (
    db,
    User,
    UserSettings,
    UserProfileData,
    FtpConnection,
    RostelecomAtsConnection,
    UserVocabulary,
    UserPrompt,
    UserScriptPrompt,
    UserConfig,
    UserStation,
    UserStationMapping,
    UserStationChatId,
    UserEmployeeExtension,
    ReportSchedule,
    BUSINESS_PROFILES,
    VOCAB_PRESETS,
)
from auth import login_manager
from auth.routes import auth_bp
try:
    from call_analyzer.service_manager import request_reload
except ImportError:
    def request_reload(user_id: int):
        pass

try:
    from call_analyzer.call_handler import (
        thebai_analyze,
        get_station_prompts,
        load_additional_vocab,
        save_transcript_analysis,
        parse_filename,
    )
    from call_analyzer.internal_transcription import transcribe_audio_with_internal_service
    from call_analyzer.exental_alert import run_exental_alert
except ImportError:
    thebai_analyze = None
    get_station_prompts = None
    load_additional_vocab = None
    save_transcript_analysis = None
    parse_filename = None
    transcribe_audio_with_internal_service = None
    run_exental_alert = None


def _dedupe_checklist_qa_text(text_value: str) -> str:
    """
    В некоторых ответах LLM чек-лист (нумерованный список 1..N) дублируется дважды.
    Удаляем второй блок, обрезая всё, что начинается со второго вхождения строки '1. ...'.
    """
    if not text_value:
        return text_value
    try:
        matches = list(re.finditer(r'(?m)^\s*1\.\s', text_value))
        if len(matches) >= 2:
            return text_value[:matches[1].start()].rstrip()
    except Exception:
        pass
    return text_value

# --- Нормализованный доступ к конфигурации пользователя ---
def _get_or_create_user_config_record(actual_user):
    """Возвращает запись user_config, создаёт при необходимости."""
    cfg = UserConfig.query.filter_by(user_id=actual_user.id).first()
    if not cfg:
        cfg = UserConfig(user_id=actual_user.id)
        db.session.add(cfg)
        db.session.commit()
    return cfg


def get_user_config_data(user=None):
    """Загружает конфигурацию пользователя из нормализованных таблиц."""
    actual_user = user if user else (current_user if hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None)
    if not actual_user:
        return default_config_template()

    cfg = UserConfig.query.filter_by(user_id=actual_user.id).first()
    config_data = default_config_template()

    if cfg:
        paths = config_data.get('paths') or {}
        paths.update({
            'source_type': cfg.source_type,
            'prompts_file': cfg.prompts_file,
            'base_records_path': cfg.base_records_path,
            'ftp_connection_id': cfg.ftp_connection_id,
            'rostelecom_ats_connection_id': getattr(cfg, 'rostelecom_ats_connection_id', None),
            'script_prompt_file': cfg.script_prompt_file,
            'additional_vocab_file': cfg.additional_vocab_file,
        })
        config_data['paths'] = paths

        config_data['api_keys'] = {
            'speechmatics_api_key': cfg.speechmatics_api_key or '',
            'thebai_api_key': cfg.thebai_api_key or '',
            'thebai_url': config_data['api_keys'].get('thebai_url', 'https://api.deepseek.com/v1/chat/completions'),
            'thebai_model': config_data['api_keys'].get('thebai_model', 'deepseek-reasoner'),
            'telegram_bot_token': cfg.telegram_bot_token or '',
        }

        config_data['telegram'] = {
            'alert_chat_id': cfg.alert_chat_id or '',
            'tg_channel_nizh': cfg.tg_channel_nizh or '',
            'tg_channel_other': cfg.tg_channel_other or '',
            'reports_chat_id': getattr(cfg, 'reports_chat_id', None) or '',
        }

        config_data['transcription'] = {
            'tbank_stereo_enabled': bool(cfg.tbank_stereo_enabled),
            'use_additional_vocab': bool(cfg.use_additional_vocab),
            'auto_detect_operator_name': bool(cfg.auto_detect_operator_name),
        }

        config_data['filename'] = {
            'enabled': bool(cfg.use_custom_filename_patterns),
            'patterns': cfg.filename_patterns or [],
            'extensions': cfg.filename_extensions or ['.mp3', '.wav']
        }

        config_data['allowed_stations'] = cfg.allowed_stations or []
        config_data['nizh_station_codes'] = cfg.nizh_station_codes or []
        config_data['legal_entity_keywords'] = cfg.legal_entity_keywords or []

    stations = {s.code: s.name for s in UserStation.query.filter_by(user_id=actual_user.id).all()}
    station_chat_ids = {}
    for row in UserStationChatId.query.filter_by(user_id=actual_user.id).all():
        station_chat_ids.setdefault(row.station_code, []).append(row.chat_id)
    station_mapping = {}
    for row in UserStationMapping.query.filter_by(user_id=actual_user.id).all():
        station_mapping.setdefault(row.main_station_code, []).append(row.sub_station_code)
    employee_by_extension = {
        row.extension: row.employee
        for row in UserEmployeeExtension.query.filter_by(user_id=actual_user.id).all()
    }

    config_data['stations'] = stations
    config_data['station_chat_ids'] = station_chat_ids
    config_data['station_mapping'] = station_mapping
    config_data['employee_by_extension'] = employee_by_extension

    return config_data


# Legacy config loader (call_analyzer/config.py)
LEGACY_CONFIG_MODULE_NAME = 'call_analyzer_legacy_config'
LEGACY_CONFIG_PATH = PROJECT_ROOT / 'call_analyzer' / 'config.py'


class MockConfig:
    SPEECHMATICS_API_KEY = ''
    THEBAI_API_KEY = ''
    TELEGRAM_BOT_TOKEN = ''
    ALERT_CHAT_ID = ''
    TG_CHANNEL_NIZH = ''
    TG_CHANNEL_OTHER = ''
    BASE_RECORDS_PATH = ''
    PROMPTS_FILE = ''
    ADDITIONAL_VOCAB_FILE = ''
    STATION_NAMES = {}
    STATION_CHAT_IDS = {}
    STATION_MAPPING = {}
    NIZH_STATION_CODES = []


def load_legacy_project_config():
    if not LEGACY_CONFIG_PATH.exists():
        print('Ошибка: файл config.py в call_analyzer не найден')
        return MockConfig()
    try:
        spec = importlib.util.spec_from_file_location(
            LEGACY_CONFIG_MODULE_NAME,
            str(LEGACY_CONFIG_PATH)
        )
        module = importlib.util.module_from_spec(spec)
        sys.modules[LEGACY_CONFIG_MODULE_NAME] = module
        spec.loader.exec_module(module)
        return module
    except Exception as exc:
        print(f'Ошибка: не удалось загрузить config.py из call_analyzer ({exc})')
        return MockConfig()


project_config = load_legacy_project_config()


def get_user_settings_record(user=None, auto_create=True):
    """Возвращает или создаёт запись с пользовательскими настройками."""
    actual_user = user
    if actual_user is None and hasattr(current_user, 'is_authenticated'):
        if current_user.is_authenticated:
            actual_user = current_user
    if actual_user is None or not getattr(actual_user, 'id', None):
        return None

    settings = getattr(actual_user, 'settings', None)
    if settings is None:
        settings = UserSettings.query.filter_by(user_id=actual_user.id).first()

    if not settings and auto_create:
        settings = UserSettings(user_id=actual_user.id, data={})
        db.session.add(settings)
        db.session.commit()
        actual_user.settings = settings

    return settings


def get_user_settings_section(section, default_factory, user=None):
    """Возвращает копию секции настроек пользователя."""
    settings = get_user_settings_record(user=user, auto_create=True)
    if not settings:
        return default_factory()

    data = settings.data or {}
    if section not in data or data[section] is None:
        data[section] = default_factory()
        settings.data = data
        db.session.add(settings)
        db.session.commit()

    return deepcopy(data[section])


def save_user_settings_section(section, value, user=None):
    """��������� ������ �������� ������������."""
    settings = get_user_settings_record(user=user, auto_create=True)
    if not settings:
        raise RuntimeError('��������� ������������ ����������.')

    payload = json.dumps(value, ensure_ascii=False)
    db.session.execute(
        text(
            "UPDATE user_settings "
            "SET data = jsonb_set(COALESCE(data, '{}'::jsonb), CAST(:path AS text[]), CAST(:payload AS jsonb), true) "
            "WHERE id = :sid"
        ),
        {
            'path': '{%s}' % section,
            'payload': payload,
            'sid': settings.id
        }
    )
    db.session.commit()
    db.session.refresh(settings)

    return deepcopy(value)


def _get_or_create_user_config_record(actual_user):
    """Возвращает запись user_config, создаёт при необходимости."""
    cfg = UserConfig.query.filter_by(user_id=actual_user.id).first()
    if not cfg:
        cfg = UserConfig(user_id=actual_user.id)
        db.session.add(cfg)
        db.session.commit()
    return cfg


def get_user_config_data(user=None):
    """Загружает конфигурацию пользователя из нормализованных таблиц."""
    actual_user = _resolve_current_user(user)
    if not actual_user:
        return default_config_template()

    cfg = UserConfig.query.filter_by(user_id=actual_user.id).first()
    config_data = default_config_template()

    # Paths
    if cfg:
        paths = config_data.get('paths') or {}
        paths.update({
            'source_type': cfg.source_type,
            'prompts_file': cfg.prompts_file,
            'base_records_path': cfg.base_records_path,
            'ftp_connection_id': cfg.ftp_connection_id,
            'rostelecom_ats_connection_id': getattr(cfg, 'rostelecom_ats_connection_id', None),
            'script_prompt_file': cfg.script_prompt_file,
            'additional_vocab_file': cfg.additional_vocab_file,
        })
        config_data['paths'] = paths

        # API Keys
        config_data['api_keys'] = {
            'speechmatics_api_key': cfg.speechmatics_api_key or '',
            'thebai_api_key': cfg.thebai_api_key or '',
            'thebai_url': config_data['api_keys'].get('thebai_url', 'https://api.deepseek.com/v1/chat/completions'),
            'thebai_model': config_data['api_keys'].get('thebai_model', 'deepseek-reasoner'),
            'telegram_bot_token': cfg.telegram_bot_token or '',
        }

        # Telegram
        config_data['telegram'] = {
            'alert_chat_id': cfg.alert_chat_id or '',
            'tg_channel_nizh': cfg.tg_channel_nizh or '',
            'tg_channel_other': cfg.tg_channel_other or '',
            'reports_chat_id': getattr(cfg, 'reports_chat_id', None) or '',
        }

        # Transcription
        config_data['transcription'] = {
            'tbank_stereo_enabled': bool(cfg.tbank_stereo_enabled),
            'use_additional_vocab': bool(cfg.use_additional_vocab),
            'auto_detect_operator_name': bool(cfg.auto_detect_operator_name),
        }

        # Filename formats
        config_data['filename'] = {
            'enabled': bool(getattr(cfg, 'use_custom_filename_patterns', False)),
            'patterns': getattr(cfg, 'filename_patterns', None) or [],
            'extensions': getattr(cfg, 'filename_extensions', None) or ['.mp3', '.wav'],
        }

        # Arrays/JSONB
        config_data['allowed_stations'] = cfg.allowed_stations or []
        config_data['nizh_station_codes'] = cfg.nizh_station_codes or []
        config_data['legal_entity_keywords'] = cfg.legal_entity_keywords or []

        # Отраслевой профиль (многоотраслевая платформа)
        config_data['business_profile'] = getattr(cfg, 'business_profile', None) or 'autoservice'

    if 'business_profile' not in config_data:
        config_data['business_profile'] = 'autoservice'

    # Stations
    stations = {s.code: s.name for s in UserStation.query.filter_by(user_id=actual_user.id).all()}
    station_chat_ids = {}
    for row in UserStationChatId.query.filter_by(user_id=actual_user.id).all():
        station_chat_ids.setdefault(row.station_code, []).append(row.chat_id)
    station_mapping = {}
    for row in UserStationMapping.query.filter_by(user_id=actual_user.id).all():
        station_mapping.setdefault(row.main_station_code, []).append(row.sub_station_code)
    employee_by_extension = {
        row.extension: row.employee
        for row in UserEmployeeExtension.query.filter_by(user_id=actual_user.id).all()
    }

    config_data['stations'] = stations
    config_data['station_chat_ids'] = station_chat_ids
    config_data['station_mapping'] = station_mapping
    config_data['employee_by_extension'] = employee_by_extension

    return config_data


def save_user_config_data(config_data, user=None):
    """Сохраняет конфигурацию пользователя в нормализованные таблицы."""
    actual_user = _resolve_current_user(user)
    if not actual_user:
        raise RuntimeError('Пользователь не определен.')

    cfg = _get_or_create_user_config_record(actual_user)

    paths = config_data.get('paths') or {}
    api_keys = config_data.get('api_keys') or {}
    telegram_cfg = config_data.get('telegram') or {}
    transcription_cfg = config_data.get('transcription') or {}

    cfg.source_type = paths.get('source_type')
    cfg.prompts_file = paths.get('prompts_file')
    cfg.base_records_path = paths.get('base_records_path')
    cfg.ftp_connection_id = paths.get('ftp_connection_id')
    cfg.rostelecom_ats_connection_id = paths.get('rostelecom_ats_connection_id')
    cfg.script_prompt_file = paths.get('script_prompt_file')
    cfg.additional_vocab_file = paths.get('additional_vocab_file')

    cfg.speechmatics_api_key = api_keys.get('speechmatics_api_key')
    cfg.thebai_api_key = api_keys.get('thebai_api_key')
    cfg.telegram_bot_token = api_keys.get('telegram_bot_token')

    cfg.alert_chat_id = telegram_cfg.get('alert_chat_id')
    cfg.tg_channel_nizh = telegram_cfg.get('tg_channel_nizh')
    cfg.tg_channel_other = telegram_cfg.get('tg_channel_other')
    cfg.reports_chat_id = telegram_cfg.get('reports_chat_id')

    cfg.tbank_stereo_enabled = bool(transcription_cfg.get('tbank_stereo_enabled', False))
    cfg.use_additional_vocab = bool(transcription_cfg.get('use_additional_vocab', True))
    # По умолчанию должно совпадать с default_config_template (True)
    cfg.auto_detect_operator_name = bool(transcription_cfg.get('auto_detect_operator_name', True))

    filename_cfg = config_data.get('filename') or {}
    cfg.use_custom_filename_patterns = bool(filename_cfg.get('enabled', False))
    cfg.filename_patterns = filename_cfg.get('patterns') or []
    cfg.filename_extensions = filename_cfg.get('extensions') or ['.mp3', '.wav']

    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(cfg, 'filename_patterns')
    flag_modified(cfg, 'filename_extensions')
    flag_modified(cfg, 'allowed_stations')
    flag_modified(cfg, 'nizh_station_codes')
    flag_modified(cfg, 'legal_entity_keywords')

    cfg.allowed_stations = config_data.get('allowed_stations') or []
    cfg.nizh_station_codes = config_data.get('nizh_station_codes') or []
    cfg.legal_entity_keywords = config_data.get('legal_entity_keywords') or []

    profile = config_data.get('business_profile')
    if profile and profile in ('autoservice', 'restaurant', 'dental', 'retail', 'medical', 'universal'):
        cfg.business_profile = profile

    db.session.add(cfg)

    # Станции, chat_id, маппинги, сотрудники
    UserStation.query.filter_by(user_id=actual_user.id).delete()
    UserStationChatId.query.filter_by(user_id=actual_user.id).delete()
    UserStationMapping.query.filter_by(user_id=actual_user.id).delete()
    UserEmployeeExtension.query.filter_by(user_id=actual_user.id).delete()

    stations = config_data.get('stations') or {}
    for code, name in stations.items():
        if code:
            db.session.add(UserStation(user_id=actual_user.id, code=str(code), name=str(name or code)))

    station_chat_ids = config_data.get('station_chat_ids') or {}
    for code, chat_list in station_chat_ids.items():
        if not code:
            continue
        for chat_id in chat_list or []:
            if chat_id:
                db.session.add(UserStationChatId(user_id=actual_user.id, station_code=str(code), chat_id=str(chat_id)))

    station_mapping = config_data.get('station_mapping') or {}
    for main_code, sub_list in station_mapping.items():
        if not main_code:
            continue
        for sub_code in sub_list or []:
            if sub_code:
                db.session.add(UserStationMapping(user_id=actual_user.id, main_station_code=str(main_code), sub_station_code=str(sub_code)))

    employees = config_data.get('employee_by_extension') or {}
    for ext, emp in employees.items():
        if ext and emp:
            db.session.add(UserEmployeeExtension(user_id=actual_user.id, extension=str(ext), employee=str(emp)))

    db.session.commit()
    return config_data


def normalize_prompts_payload(data):
    normalized = {
        'default': '',
        'anchors': {},
        'stations': {}
    }
    if not isinstance(data, dict):
        return normalized

    normalized['default'] = data.get('default') or ''

    anchors = data.get('anchors')
    if isinstance(anchors, dict):
        normalized['anchors'] = {
            str(key): str(value) if value is not None else ''
            for key, value in anchors.items()
        }
    else:
        derived = {}
        for key, value in data.items():
            if key in ('default', 'stations', 'anchors'):
                continue
            if isinstance(value, str):
                derived[str(key)] = value
        normalized['anchors'] = derived

    stations = data.get('stations')
    if isinstance(stations, dict):
        normalized['stations'] = {
            str(key): str(value) if value is not None else ''
            for key, value in stations.items()
        }
    else:
        normalized['stations'] = {}

    return normalized


def get_user_prompts_data(user=None):
    """Загружает промпты из таблицы user_prompts с обратной совместимостью"""
    if user is None:
        user = current_user if hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None
    if not user:
        return normalize_prompts_payload(default_prompts_template())
    
    # Сначала пытаемся загрузить из новых таблиц
    prompts = UserPrompt.query.filter_by(user_id=user.id).all()
    
    result = {
        'default': '',
        'anchors': {},
        'stations': {}
    }
    
    has_data = False
    for prompt in prompts:
        has_data = True
        if prompt.prompt_type == 'default':
            result['default'] = prompt.prompt_text or ''
        elif prompt.prompt_type == 'anchor':
            result['anchors'][prompt.prompt_key] = prompt.prompt_text or ''
        elif prompt.prompt_type == 'station':
            result['stations'][prompt.prompt_key] = prompt.prompt_text or ''
    
    # Если в новых таблицах пусто, читаем из старых UserSettings.data
    if not has_data:
        old_data = get_user_settings_section('prompts', default_prompts_template, user=user)
        if old_data and (old_data.get('anchors') or old_data.get('stations') or old_data.get('default')):
            # Автоматически мигрируем данные
            save_user_prompts_data(old_data, user=user)
            return normalize_prompts_payload(old_data)
    
    return normalize_prompts_payload(result)


def save_user_prompts_data(prompts_data, user=None):
    """Сохраняет промпты в таблицу user_prompts"""
    if user is None:
        user = current_user if hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None
    if not user:
        raise RuntimeError('Пользователь не определен.')
    
    normalized = normalize_prompts_payload(prompts_data)
    
    # Удаляем старые промпты
    UserPrompt.query.filter_by(user_id=user.id).delete()
    
    # Сохраняем default
    if normalized.get('default'):
        prompt = UserPrompt(
            user_id=user.id,
            prompt_type='default',
            prompt_key='default',
            prompt_text=str(normalized['default'])
        )
        db.session.add(prompt)
    
    # Сохраняем anchors
    anchors = normalized.get('anchors', {})
    for key, text in anchors.items():
        if text:  # Сохраняем только непустые
            prompt = UserPrompt(
                user_id=user.id,
                prompt_type='anchor',
                prompt_key=str(key),
                prompt_text=str(text)
            )
            db.session.add(prompt)
    
    # Сохраняем stations
    stations = normalized.get('stations', {})
    for station_code, text in stations.items():
        if text:  # Сохраняем только непустые
            prompt = UserPrompt(
                user_id=user.id,
                prompt_type='station',
                prompt_key=str(station_code),
                prompt_text=str(text)
            )
            db.session.add(prompt)
    
    db.session.commit()
    return deepcopy(normalized)


def get_user_vocabulary_data(user=None):
    """Загружает словарь из таблицы user_vocabulary с обратной совместимостью"""
    if user is None:
        user = current_user if hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None
    if not user:
        return default_vocabulary_template()
    
    vocab = UserVocabulary.query.filter_by(user_id=user.id).first()
    
    # Если в новых таблицах есть данные, возвращаем их
    if vocab and (vocab.additional_vocab or vocab.enabled is not None):
        return {
            'enabled': vocab.enabled,
            'additional_vocab': vocab.additional_vocab or []
        }
    
    # Если в новых таблицах пусто, читаем из старых UserSettings.data
    old_data = get_user_settings_section('vocabulary', default_vocabulary_template, user=user)
    if old_data and (old_data.get('additional_vocab') or old_data.get('enabled') is not None):
        # Автоматически мигрируем данные
        save_user_vocabulary_data(old_data, user=user)
        return old_data
    
    # Создаем запись с дефолтными значениями
    if not vocab:
        vocab = UserVocabulary(
            user_id=user.id,
            enabled=True,
            additional_vocab=[]
        )
        db.session.add(vocab)
        db.session.commit()
    
    return default_vocabulary_template()


def save_user_vocabulary_data(vocab_data, user=None):
    """Сохраняет словарь в таблицу user_vocabulary"""
    if user is None:
        user = current_user if hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None
    if not user:
        raise RuntimeError('Пользователь не определен.')
    
    vocab = UserVocabulary.query.filter_by(user_id=user.id).first()
    if not vocab:
        vocab = UserVocabulary(user_id=user.id)
        db.session.add(vocab)
    
    vocab.enabled = bool(vocab_data.get('enabled', True))
    vocab.additional_vocab = vocab_data.get('additional_vocab', []) or []
    vocab.updated_at = datetime.utcnow()
    
    db.session.commit()
    return deepcopy(vocab_data)


def get_user_logs(user=None):
    return get_user_settings_section('logs', default_logs_template, user=user)


def save_user_logs(logs, user=None):
    return save_user_settings_section('logs', logs, user=user)


def get_user_script_prompt(user=None):
    """Загружает script_prompt из таблицы user_script_prompts с обратной совместимостью"""
    if user is None:
        user = current_user if hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None
    if not user:
        return default_script_prompt_template()
    
    script_prompt = UserScriptPrompt.query.filter_by(user_id=user.id).first()
    
    # Если в новых таблицах есть данные, возвращаем их
    if script_prompt and (script_prompt.prompt_text or script_prompt.checklist):
        return {
            'prompt': script_prompt.prompt_text or '',
            'checklist': script_prompt.checklist or []
        }
    
    # Если в новых таблицах пусто, читаем из старых UserSettings.data
    old_data = get_user_settings_section('script_prompt', default_script_prompt_template, user=user)
    if old_data and (old_data.get('prompt') or old_data.get('checklist')):
        # Автоматически мигрируем данные
        save_user_script_prompt(old_data, user=user)
        return old_data
    
    # Создаем запись с дефолтными значениями
    if not script_prompt:
        script_prompt = UserScriptPrompt(
            user_id=user.id,
            prompt_text='',
            checklist=[]
        )
        db.session.add(script_prompt)
        db.session.commit()
    
    return default_script_prompt_template()


def save_user_script_prompt(data, user=None):
    """Сохраняет script_prompt в таблицу user_script_prompts"""
    if user is None:
        user = current_user if hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None
    if not user:
        raise RuntimeError('Пользователь не определен.')
    
    script_prompt = UserScriptPrompt.query.filter_by(user_id=user.id).first()
    if not script_prompt:
        script_prompt = UserScriptPrompt(user_id=user.id)
        db.session.add(script_prompt)
    
    script_prompt.prompt_text = str(data.get('prompt', ''))
    script_prompt.checklist = data.get('checklist', []) or []
    script_prompt.updated_at = datetime.utcnow()
    
    db.session.commit()
    return deepcopy(data)


def get_user_prompts_file_path(user=None):
    """
    Возвращает путь к prompts.yaml для пользователя (создаёт каталоги при необходимости).
    """
    runtime_cfg = build_user_runtime_config(user=user)
    prompts_file = (runtime_cfg.get('paths') or {}).get('prompts_file')
    if not prompts_file:
        return None
    path = Path(prompts_file)
    # Создаем директорию с правильными правами
    path.parent.mkdir(parents=True, exist_ok=True)
    # Устанавливаем права доступа (если директория только что создана)
    try:
        import os
        import stat
        # Устанавливаем права 755 на директорию
        os.chmod(path.parent, stat.S_IRWXU | stat.S_IRGRP | stat.S_IXGRP | stat.S_IROTH | stat.S_IXOTH)
    except Exception:
        pass  # Игнорируем ошибки прав доступа (может не хватать прав)
    return path


def load_prompts_file(user=None):
    """
    Загружает anchors/default/stations из prompts.yaml текущего пользователя.
    """
    result = {
        'anchors': {},
        'stations': {},
        'default': ''
    }
    path = get_user_prompts_file_path(user=user)
    if not path or not path.exists():
        return result

    try:
        with path.open('r', encoding='utf-8') as handler:
            raw = yaml.safe_load(handler) or {}
    except Exception as exc:
        app.logger.warning("Не удалось прочитать prompts-файл %s: %s", path, exc)
        return result

    if isinstance(raw.get('default'), str):
        result['default'] = raw.get('default') or ''

    stations = raw.get('stations')
    if isinstance(stations, dict):
        result['stations'] = {
            str(code): str(text) if text is not None else ''
            for code, text in stations.items()
        }

    anchors_block = raw.get('anchors')
    anchors = {}
    if isinstance(anchors_block, dict):
        anchors = {
            str(name): str(text) if text is not None else ''
            for name, text in anchors_block.items()
        }
    else:
        for key, value in raw.items():
            if key in ('default', 'stations', 'anchors'):
                continue
            if isinstance(value, str):
                anchors[str(key)] = value
    result['anchors'] = anchors
    return result


def write_prompts_file(prompts_data, user=None):
    """
    Сохраняет anchors/default/stations в пользовательский prompts.yaml.
    """
    path = get_user_prompts_file_path(user=user)
    if not path:
        return

    normalized = normalize_prompts_payload(prompts_data)
    payload = OrderedDict()
    if normalized['default']:
        payload['default'] = normalized['default']

    for name in sorted((normalized['anchors'] or {}).keys()):
        text = normalized['anchors'][name]
        payload[name] = text or ''

    payload['stations'] = normalized['stations'] or {}

    # YAML иногда ругается на OrderedDict, преобразуем в обычный dict через JSON-сериализацию
    try:
        serializable_payload = json.loads(json.dumps(payload, ensure_ascii=False))
    except Exception:
        serializable_payload = dict(payload)

    try:
        with path.open('w', encoding='utf-8') as handler:
            yaml.safe_dump(serializable_payload, handler, allow_unicode=True, sort_keys=False)
    except Exception as exc:
        app.logger.error("Не удалось сохранить prompts-файл %s: %s", path, exc)


def get_user_vocabulary_file_path(user=None):
    """Возвращает путь к файлу словаря пользователя"""
    runtime_cfg = build_user_runtime_config(user=user)
    paths = runtime_cfg.get('paths', {})
    vocab_file = paths.get('additional_vocab_file')
    if vocab_file:
        return Path(vocab_file)
    return None


def write_vocabulary_file(vocab_data, user=None):
    """Сохраняет словарь в YAML файл"""
    path = get_user_vocabulary_file_path(user=user)
    if not path:
        return
    
    # Формируем структуру для YAML
    payload = {
        'additional_vocab': vocab_data.get('additional_vocab', [])
    }
    
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open('w', encoding='utf-8') as handler:
            yaml.safe_dump(payload, handler, allow_unicode=True, sort_keys=False, default_flow_style=False)
    except Exception as exc:
        app.logger.error("Не удалось сохранить vocabulary файл %s: %s", path, exc)


def get_user_script_prompt_path(user=None):
    runtime_cfg = build_user_runtime_config(user=user)
    script_file = (runtime_cfg.get('paths') or {}).get('script_prompt_file')
    if not script_file:
        return None
    path = Path(script_file)
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def load_script_prompt_file(user=None, custom_path=None):
    if custom_path:
        path = Path(custom_path)
    else:
        path = get_user_script_prompt_path(user=user)
        
    if not path or not path.exists():
        return default_script_prompt_template()
    try:
        with path.open('r', encoding='utf-8') as handler:
            raw = yaml.safe_load(handler) or {}
    except Exception as exc:
        app.logger.warning("Не удалось прочитать script prompt файл %s: %s", path, exc)
        return default_script_prompt_template()

    data = default_script_prompt_template()
    if isinstance(raw, dict):
        checklist = raw.get('checklist')
        if isinstance(checklist, list):
            normalized = []
            for item in checklist:
                if not isinstance(item, dict):
                    continue
                normalized.append({
                    'title': str(item.get('title') or ''),
                    'prompt': str(item.get('prompt') or '')
                })
            data['checklist'] = normalized
        if isinstance(raw.get('prompt'), str):
            data['prompt'] = raw.get('prompt')
    return data


def write_script_prompt_file(data, user=None, custom_path=None):
    if custom_path:
        path = Path(custom_path)
    else:
        path = get_user_script_prompt_path(user=user)
        
    if not path:
        return
    payload = default_script_prompt_template()
    checklist = []
    for item in data.get('checklist') or []:
        if not isinstance(item, dict):
            continue
        checklist.append({
            'title': str(item.get('title') or ''),
            'prompt': str(item.get('prompt') or '')
        })
    payload['checklist'] = checklist
    payload['prompt'] = str(data.get('prompt') or '')
    try:
        # Убеждаемся, что папка существует
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open('w', encoding='utf-8') as handler:
            yaml.safe_dump(payload, handler, allow_unicode=True, sort_keys=False)
    except Exception as exc:
        app.logger.error("Не удалось сохранить script prompt файл %s: %s", path, exc)


def _resolve_current_user(user=None):
    """Возвращает фактического пользователя (текущий, если не передан явно)."""
    if user:
        return user
    if hasattr(current_user, 'is_authenticated') and current_user.is_authenticated:
        return current_user
    return None


def build_user_runtime_config(user=None, persist_defaults=True):
    """
    Собирает конфигурацию профиля с учётом дефолтов и legacy-config.
    Возвращает словарь с ключевыми секциями (api_keys, paths и др.).
    """
    actual_user = _resolve_current_user(user)
    config_data = get_user_config_data(user=actual_user)
    runtime, updated_config, changed = build_runtime_config(
        project_config,
        config_data,
        user_id=getattr(actual_user, 'id', None)
    )

    if changed and persist_defaults and actual_user:
        save_user_config_data(updated_config, user=actual_user)

    runtime['config_data'] = updated_config
    return runtime


def get_runtime_context(user=None):
    """
    Возвращает пару (runtime_config, runtime_dir) для текущего пользователя.
    runtime_dir указывает на подкаталог runtime внутри пользовательского BASE_RECORDS_PATH.
    """
    runtime_cfg = build_user_runtime_config(user=user)
    base_records_path = Path(runtime_cfg['paths']['base_records_path'])
    runtime_dir = base_records_path / 'runtime'
    runtime_dir.mkdir(parents=True, exist_ok=True)
    return runtime_cfg, runtime_dir


@contextmanager
def legacy_config_override(runtime_cfg):
    """
    Временное применение пользовательских настроек к legacy-модулю call_analyzer.config.
    Позволяет запускать отчёты в контексте конкретного профиля.
    Патчит как call_analyzer.config, так и config (если он загружен отдельно).
    """
    import sys
    configs_to_patch = []

    # 1. call_analyzer.config
    try:
        from call_analyzer import config as legacy_config
        configs_to_patch.append(legacy_config)
    except ImportError:
        pass

    # 2. config (если загружен как top-level модуль, что делает week_full.py и exental_alert.py)
    if 'config' in sys.modules:
        configs_to_patch.append(sys.modules['config'])

    if not configs_to_patch:
        yield None
        return

    sentinel = object()
    # overrides хранит список изменений для каждого модуля: {module_obj: {attr: prev_value}}
    overrides = {cfg: {} for cfg in configs_to_patch}

    def _set_attr(attr, value):
        for cfg in configs_to_patch:
            if cfg not in overrides:
                overrides[cfg] = {}
            
            # Сохраняем старое значение только один раз
            if attr not in overrides[cfg]:
                overrides[cfg][attr] = getattr(cfg, attr, sentinel)
            
            setattr(cfg, attr, value)

    def _restore():
        for cfg, attrs in overrides.items():
            for attr, prev in attrs.items():
                if prev is sentinel:
                    delattr(cfg, attr)
                else:
                    setattr(cfg, attr, prev)

    try:
        paths_cfg = runtime_cfg.get('paths', {})
        base_path = paths_cfg.get('base_records_path')
        prompts_file = paths_cfg.get('prompts_file')
        vocab_file = paths_cfg.get('additional_vocab_file')
        script_prompt_file = paths_cfg.get('script_prompt_file')

        if base_path:
            _set_attr('BASE_RECORDS_PATH', Path(base_path))
        if prompts_file:
            _set_attr('PROMPTS_FILE', Path(prompts_file))
        if vocab_file:
            _set_attr('ADDITIONAL_VOCAB_FILE', Path(vocab_file))
        if script_prompt_file:
            _set_attr('SCRIPT_PROMPT_8_PATH', Path(script_prompt_file))

        api_keys = runtime_cfg.get('api_keys', {})
        _set_attr('SPEECHMATICS_API_KEY', api_keys.get('speechmatics_api_key', ''))
        _set_attr('THEBAI_API_KEY', api_keys.get('thebai_api_key', ''))
        _set_attr('THEBAI_URL', api_keys.get('thebai_url', 'https://api.deepseek.com/v1/chat/completions'))
        _set_attr('THEBAI_MODEL', api_keys.get('thebai_model', 'deepseek-reasoner'))
        _set_attr('TELEGRAM_BOT_TOKEN', api_keys.get('telegram_bot_token', ''))

        telegram_cfg = runtime_cfg.get('telegram') or {}
        _set_attr('ALERT_CHAT_ID', telegram_cfg.get('alert_chat_id', ''))
        _set_attr('TG_CHANNEL_NIZH', telegram_cfg.get('tg_channel_nizh', ''))
        _set_attr('TG_CHANNEL_OTHER', telegram_cfg.get('tg_channel_other', ''))
        _set_attr('REPORTS_CHAT_ID', telegram_cfg.get('reports_chat_id', ''))

        transcription_cfg = runtime_cfg.get('transcription') or {}
        _set_attr('TBANK_STEREO_ENABLED', bool(transcription_cfg.get('tbank_stereo_enabled', False)))
        _set_attr('USE_ADDITIONAL_VOCAB', bool(transcription_cfg.get('use_additional_vocab', True)))
        _set_attr('AUTO_DETECT_OPERATOR_NAME', bool(transcription_cfg.get('auto_detect_operator_name', True)))

        _set_attr('EMPLOYEE_BY_EXTENSION', deepcopy(runtime_cfg.get('employee_by_extension') or {}))
        _set_attr('STATION_NAMES', deepcopy(runtime_cfg.get('stations') or {}))
        _set_attr('STATION_CHAT_IDS', deepcopy(runtime_cfg.get('station_chat_ids') or {}))
        _set_attr('STATION_MAPPING', deepcopy(runtime_cfg.get('station_mapping') or {}))
        _set_attr('NIZH_STATION_CODES', list(runtime_cfg.get('nizh_station_codes') or []))

        # Возвращаем первый конфиг для совместимости (хотя yield может и не использоваться)
        yield configs_to_patch[0] if configs_to_patch else None
    finally:
        _restore()


def append_user_log(message, level='INFO', module='system', user=None):
    """Добавляет запись в персональные логи пользователя."""
    logs = get_user_logs(user=user)
    logs.append({
        'timestamp': datetime.utcnow().isoformat(),
        'level': level,
        'module': module,
        'message': message
    })
    # ограничиваем размер списка
    logs = logs[-500:]
    save_user_logs(logs, user=user)
def reload_project_config():
    """Перезагружает конфигурацию проекта"""
    global project_config
    try:
        if isinstance(project_config, MockConfig):
            project_config = load_legacy_project_config()
        else:
            project_config = importlib.reload(project_config)
        logging.info("Конфигурация проекта перезагружена")
    except Exception as e:
        project_config = load_legacy_project_config()
        logging.error(f"Ошибка перезагрузки конфигурации: {e}")
app = Flask(__name__)
app.config.from_object(get_config())

# ВАЖНО: Убеждаемся, что SECRET_KEY установлен
if not app.config.get('SECRET_KEY') or app.config.get('SECRET_KEY') == 'dev-secret-key-change-in-production':
    import secrets
    app.config['SECRET_KEY'] = secrets.token_hex(32)
    logging.warning("SECRET_KEY был сгенерирован автоматически. Установите его в .env файле!")

# Настройки сессии для production
try:
    app.config['SESSION_PERMANENT'] = True
    # PERMANENT_SESSION_LIFETIME должен быть числом (секунды), а не timedelta
    app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=1).total_seconds()
except Exception as e:
    logging.error(f"Ошибка настройки сессии: {e}")

db.init_app(app)
login_manager.init_app(app)
app.register_blueprint(auth_bp)

AUTH_EXEMPT_ENDPOINTS = {'static'}


@app.before_request
def require_login():
    """�?�?�?�?�?�? �������?�?�?�? �������?�� �?� Flask-Login."""
    if app.config.get('LOGIN_DISABLED'):
        return
    endpoint = request.endpoint
    if not endpoint or endpoint in AUTH_EXEMPT_ENDPOINTS:
        return
    if endpoint.startswith('auth.'):
        return
    
    # Логируем информацию о текущем пользователе для отладки
    try:
        if hasattr(current_user, 'is_authenticated'):
            if current_user.is_authenticated:
                logging.debug(f"Пользователь {current_user.username} (ID: {current_user.id}) авторизован для {endpoint}")
                return
            else:
                logging.debug(f"Пользователь не авторизован, редирект на логин для {endpoint}")
        else:
            logging.debug(f"current_user не имеет is_authenticated для {endpoint}")
    except Exception as e:
        logging.debug(f"Ошибка проверки авторизации для {endpoint}: {e}")
    
    return redirect(url_for('auth.login', next=request.url))


# Автоматическая синхронизация при запуске
def initialize_app():
    """Инициализация приложения"""
    try:
        # Во время debug Flask запускает перезагрузчик и выполняет код дважды (родитель и дочерний процесс).
        # Выполняем инициализацию и автозапуск только в основном процессе перезагрузчика.
        if (
            app.debug
            and os.environ.get('FLASK_RUN_FROM_CLI') == 'true'
            and os.environ.get('WERKZEUG_RUN_MAIN') != 'true'
        ):
            return
        with app.app_context():
            db.create_all()
            ensure_default_admin()
        app.logger.info("Автоматическая синхронизация промптов выполнена")
        # Автозапуск сервиса анализа при старте веб-интерфейса
        ensure_service_running()
        
        # Инициализация планировщика отчетов
        try:
            from web_interface.scheduler_service import init_scheduler
            init_scheduler(app)
            app.logger.info("Планировщик отчетов инициализирован")
        except Exception as e:
            app.logger.error(f"Ошибка инициализации планировщика отчетов: {e}", exc_info=True)
    except Exception as e:
        app.logger.error(f"Ошибка автоматической синхронизации: {e}")

# Глобальные переменные для статуса сервиса
service_status = {
    'running': False,
    'pid': None,
    'last_start': None,
    'last_stop': None
}

def ensure_default_admin():
    """??????????? ??????? ???? ?? ?????? ?????????????? ??? ???????????."""
    try:
        if User.query.count() > 0:
            return

        username = os.getenv('DEFAULT_ADMIN_USERNAME', 'admin').strip()
        password = os.getenv('DEFAULT_ADMIN_PASSWORD', 'admin123').strip()
        email = os.getenv('DEFAULT_ADMIN_EMAIL', 'admin@example.com').strip()

        if not username or not password:
            app.logger.warning('?? ??????? ??????? ?????????????? ?? ?????????: ?????? ??????? ??????')
            return

        admin = User(username=username, email=email, role='admin')
        admin.set_password(password)
        db.session.add(admin)
        db.session.flush()
        db.session.add(UserSettings(user_id=admin.id, data={}))
        db.session.commit()

        app.logger.warning('?????? ????????????? ?? ?????????. ??????? ?????? ????? ??????? ?????.')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'?????? ??? ???????? ?????????????? ?? ?????????: {e}')


def get_project_root():
    """Возвращает корневую папку проекта"""
    return PROJECT_ROOT

def run_systemd_command(action):
    """Выполняет команду systemctl для управления сервисом на Linux"""
    import subprocess
    import platform
    import os
    
    # Определяем имя сервиса
    service_name = 'call-analyzer'  # По умолчанию используем единый сервис
    
    try:
        # Проверяем, какой сервис существует
        result = subprocess.run(
            ['systemctl', 'list-units', '--type=service', '--all'],
            capture_output=True,
            text=True,
            timeout=5
        )
        
        # Проверяем, какой сервис доступен
        if 'call-analyzer.service' in result.stdout:
            service_name = 'call-analyzer'
        elif 'call-analyzer-service.service' in result.stdout:
            service_name = 'call-analyzer-service'
        
        # Выполняем команду через sudo
        # ВАЖНО: Требуется настройка sudo без пароля для callanalyzer
        # Используем полный путь к systemctl, так как он может быть не в PATH
        systemctl_path = '/usr/bin/systemctl'
        if not os.path.exists(systemctl_path):
            # Пробуем найти systemctl
            result = subprocess.run(['which', 'systemctl'], capture_output=True, text=True, timeout=2)
            if result.returncode == 0:
                systemctl_path = result.stdout.strip()
            else:
                return False, "systemctl не найден в системе"
        
        if action == 'start':
            cmd = ['sudo', systemctl_path, 'start', service_name]
        elif action == 'stop':
            cmd = ['sudo', systemctl_path, 'stop', service_name]
        elif action == 'restart':
            cmd = ['sudo', systemctl_path, 'restart', service_name]
        elif action == 'status':
            cmd = [systemctl_path, 'is-active', service_name]
        else:
            return False, f"Неизвестная команда: {action}"
        
        # Выполняем команду
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=10,
            env=os.environ.copy()
        )
        
        if result.returncode == 0:
            if action == 'status':
                return result.stdout.strip() == 'active', result.stdout.strip()
            return True, f"Команда {action} выполнена успешно"
        else:
            error_msg = result.stderr or result.stdout or f"Ошибка выполнения команды {action}"
            # Если ошибка связана с правами, даем подсказку
            if 'permission denied' in error_msg.lower() or 'sudo' in error_msg.lower():
                error_msg += " (требуются права sudo для управления systemd сервисами)"
            return False, error_msg
            
    except subprocess.TimeoutExpired:
        return False, "Таймаут выполнения команды"
    except FileNotFoundError:
        return False, "systemctl не найден (не Linux система?)"
    except Exception as e:
        return False, f"Ошибка: {str(e)}"

def run_script(script_filename, wait=False):
    """Запускает bat-скрипт из корня проекта.

    Возвращает (success: bool, message: str).
    """
    try:
        project_root = get_project_root()
        script_path = project_root / script_filename
        if not script_path.exists():
            return False, f"Файл {script_filename} не найден"

        if wait:
            subprocess.run([str(script_path)], cwd=str(project_root))
        else:
            subprocess.Popen([str(script_path)], cwd=str(project_root))
        return True, f"Скрипт {script_filename} выполнен"
    except Exception as e:
        return False, str(e)

def ensure_service_running():
    """Проверяет и запускает сервис, если он не запущен."""
    try:
        import platform
        if platform.system() == 'Linux':
            # На Linux проверяем systemd
            status = get_service_status()
            if status.get('running'):
                app.logger.info('Сервис Call Analyzer запущен через systemd')
            else:
                app.logger.info('Сервис Call Analyzer не запущен. Используйте systemctl для управления')
        else:
            # На Windows сервис запускается автоматически через app.py (multiprocessing)
            # Просто отмечаем, что он должен быть запущен
            service_status['running'] = True
            service_status['last_start'] = datetime.now()
            app.logger.info('Сервис запущен при старте веб-интерфейса')
    except Exception as e:
        app.logger.error(f'Ошибка автозапуска сервиса: {e}')

def load_yaml_config(file_path):
    """Загружает YAML конфигурацию"""
    try:
        if not file_path.exists():
            return {}
        with open(file_path, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f)
    except Exception as e:
        logging.error(f"Ошибка загрузки {file_path}: {e}")
        return {}

def save_yaml_config(file_path, data):
    """Сохраняет YAML конфигурацию"""
    try:
        file_path.parent.mkdir(parents=True, exist_ok=True)
        with open(file_path, 'w', encoding='utf-8') as f:
            yaml.dump(data, f, default_flow_style=False, allow_unicode=True)
        return True
    except Exception as e:
        logging.error(f"Ошибка сохранения {file_path}: {e}")
        return False

def sync_prompts_from_config(user=None):
    """?????????????? ???????????????? ??????? ? ??? ?????????."""
    try:
        config_data = get_user_config_data(user=user)
        prompts_data = get_user_prompts_data(user=user)
        stations = config_data.get('stations') or {}
        if not isinstance(stations, dict):
            stations = {}
        prompts_data.setdefault('stations', {})
        default_prompt = prompts_data.get('default') or '?????? ?? ?????????'
        changed = False

        for station_code in stations.keys():
            if station_code not in prompts_data['stations']:
                prompts_data['stations'][station_code] = default_prompt
                changed = True

        for station_code in list(prompts_data['stations'].keys()):
            if station_code not in stations:
                prompts_data['stations'].pop(station_code)
                changed = True

        if changed:
            save_user_prompts_data(prompts_data, user=user)
            try:
                write_prompts_file(prompts_data, user=user)
            except Exception as file_error:
                app.logger.error("Не удалось обновить prompts-файл: %s", file_error)
        return changed
    except Exception as e:
        app.logger.error(f"?????? ????????????? ????????: {e}")
        return False


def get_service_status():
    """Проверяет статус сервиса Call Analyzer."""
    try:
        import platform
        running = False
        pid = None

        if platform.system() == 'Linux':
            # Проверка через systemd на Linux
            try:
                # Используем полный путь к systemctl
                systemctl_path = '/usr/bin/systemctl'
                if not os.path.exists(systemctl_path):
                    result = subprocess.run(['which', 'systemctl'], capture_output=True, text=True, timeout=2)
                    if result.returncode == 0:
                        systemctl_path = result.stdout.strip()
                    else:
                        systemctl_path = 'systemctl'  # Фолбэк
                
                # Определяем имя сервиса
                service_name = 'call-analyzer-service'
                result = subprocess.run(
                    [systemctl_path, 'list-units', '--type=service', '--state=active'],
                    capture_output=True,
                    text=True,
                    timeout=5
                )
                
                if 'call-analyzer.service' in result.stdout:
                    service_name = 'call-analyzer'
                elif 'call-analyzer-service.service' in result.stdout:
                    service_name = 'call-analyzer-service'
                
                # Проверяем статус
                result = subprocess.run(
                    [systemctl_path, 'is-active', service_name],
                    capture_output=True,
                    text=True,
                    timeout=5
                )
                
                if result.returncode == 0 and result.stdout.strip() == 'active':
                    running = True
                    # Получаем PID главного процесса
                    try:
                        result = subprocess.run(
                            [systemctl_path, 'show', '--property=MainPID', '--value', service_name],
                            capture_output=True,
                            text=True,
                            timeout=5
                        )
                        if result.returncode == 0:
                            pid_str = result.stdout.strip()
                            if pid_str and pid_str.isdigit():
                                pid = int(pid_str)
                    except Exception:
                        pass
            except Exception as e:
                logging.debug(f"Ошибка проверки статуса через systemd: {e}")
        else:
            # Windows: сервис запускается автоматически через app.py (multiprocessing)
            # Если веб-интерфейс Flask работает (раз эта функция вызвана), 
            # значит app.py запущен, а следовательно и service_manager тоже
            running = True
            pid = os.getpid()  # PID текущего процесса Flask
            app.logger.debug(f"Windows: Сервис автоматически считается запущенным (Flask PID={pid})")

        service_status['running'] = running
        service_status['pid'] = pid
    except Exception as e:
        logging.error(f"Ошибка проверки статуса сервиса: {e}")
        service_status['running'] = False
        service_status['pid'] = None

    return service_status

@app.route('/')
@login_required
def index():
    """Главная страница"""
    status = get_service_status()
    return render_template('index.html', status=status, active_page='dashboard')

@app.route('/api/status')
@login_required
def api_status():
    """API для получения статуса системы"""
    status = get_service_status()
    
    return jsonify({
        'service': status,
        'timestamp': datetime.now().isoformat()
    })

@app.route('/api/rescan-current-day', methods=['POST'])
@login_required
def api_rescan_current_day():
    """API для ручного переобхода файлов в папке текущего дня"""
    try:
        # Импортируем функцию переобхода из call_analyzer.main
        try:
            from call_analyzer.main import scan_current_day_folder
            from call_analyzer.call_handler import CallHandler, get_current_folder
        except ImportError as e:
            app.logger.error(f"Ошибка импорта модулей обработки файлов: {e}")
            return jsonify({
                'success': False,
                'message': 'Ошибка импорта модулей обработки файлов'
            }), 500
        
        # Получаем путь к папке текущего дня для текущего пользователя
        try:
            # Получаем базовый путь для текущего пользователя
            user_config = UserConfig.query.filter_by(user_id=current_user.id).first()
            if user_config and user_config.base_records_path:
                base_path = Path(user_config.base_records_path)
            else:
                # Используем дефолтный путь
                config = get_config()
                base_path = Path(config.BASE_RECORDS_PATH) / 'users' / str(current_user.id)
            
            # Формируем путь к папке текущего дня
            today = datetime.now()
            current_day_path = base_path / str(today.year) / f"{today.month:02d}" / f"{today.day:02d}"
            
            if not current_day_path.exists():
                return jsonify({
                    'success': False,
                    'message': f'Папка текущего дня не найдена: {current_day_path}'
                }), 404
            
            # Создаем временный экземпляр CallHandler для обработки
            event_handler = CallHandler()
            
            # Получаем пользовательскую конфигурацию для корректной обработки станций
            runtime_cfg = build_user_runtime_config()
            
            # Запускаем переобход в отдельном потоке, чтобы не блокировать запрос
            def run_scan():
                try:
                    # Применяем пользовательский конфиг, чтобы CallHandler видел
                    # правильные STATION_NAMES, STATION_MAPPING, EMPLOYEE_BY_EXTENSION
                    with legacy_config_override(runtime_cfg):
                        processed_count = scan_current_day_folder(event_handler, str(current_day_path))
                    app.logger.info(f"[WEB] Ручной переобход завершен. Обработано файлов: {processed_count}")
                except Exception as e:
                    app.logger.error(f"[WEB] Ошибка в потоке переобхода: {e}", exc_info=True)
            
            scan_thread = threading.Thread(target=run_scan, daemon=True)
            scan_thread.start()
            
            app.logger.info(f"[WEB] Запущен ручной переобход папки: {current_day_path}")
            
            return jsonify({
                'success': True,
                'message': f'Запущен переобход папки текущего дня',
                'path': str(current_day_path)
            })
            
        except Exception as e:
            app.logger.error(f"[WEB] Ошибка при переобходе папки: {e}", exc_info=True)
            return jsonify({
                'success': False,
                'message': f'Ошибка при переобходе папки: {str(e)}'
            }), 500
            
    except Exception as e:
        app.logger.error(f"[WEB] Критическая ошибка в api_rescan_current_day: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Критическая ошибка: {str(e)}'
        }), 500

@app.route('/api/summary/realtime')
@login_required
def api_summary_realtime():
    """Онлайн-сводка (как вкладка Сводный отчет) за указанный интервал или за сегодня.
    Параметры: start_date=YYYY-MM-DD, end_date=YYYY-MM-DD (необязательные).
    """
    try:
        # Ленивая загрузка, чтобы не тянуть зависимости раньше времени
        from reports.week_full import compute_realtime_summary

        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')

        if start_date_str:
            try:
                start_dt = datetime.strptime(start_date_str, '%Y-%m-%d')
            except ValueError:
                return jsonify({'success': False, 'message': 'Некорректный start_date'}), 400
        else:
            today = datetime.now()
            start_dt = datetime(today.year, today.month, today.day)

        if end_date_str:
            try:
                end_dt = datetime.strptime(end_date_str, '%Y-%m-%d')
            except ValueError:
                return jsonify({'success': False, 'message': 'Некорректный end_date'}), 400
        else:
            # До конца текущего дня
            end_dt = datetime.now()

        # Профильный конфиг: пути и разрешённые станции читаем из настроек пользователя
        runtime_cfg = build_user_runtime_config()
        allowed_stations = runtime_cfg.get('allowed_stations')
        if not allowed_stations:
            allowed_stations = None
        base_records_path = runtime_cfg['paths']['base_records_path']

        # Админам всегда показываем весь набор станций
        user = current_user if hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None
        if user and user.role == 'admin':
            allowed_stations = None

        station_names = runtime_cfg.get('stations') or {}
        station_mapping = runtime_cfg.get('station_mapping') or {}
        employee_by_extension = runtime_cfg.get('employee_by_extension') or {}
        script_prompt_path = runtime_cfg['paths'].get('script_prompt_file')

        summary = compute_realtime_summary(
            start_dt,
            end_dt,
            allowed_stations=allowed_stations,
            base_folder=base_records_path,
            station_names=station_names,
            station_mapping=station_mapping,
            employee_by_extension=employee_by_extension,
            script_prompt_path=script_prompt_path
        )

        # Считаем агрегаты для быстрых метрик на карточках
        total_calls = 0
        if summary and summary.get('stations'):
            for st in summary['stations']:
                for c in st.get('consultants', []):
                    total_calls += int(c.get('calls', 0))

        # Карточка «Станций»: число настроенных станций (как на /stations), а не по обработанным звонкам
        stations_count = len(station_names)

        # Считаем активные переводы и перезвоны
        active_transfers = 0
        active_recalls = 0
        try:
            # Сначала пытаемся получить из БД
            user = current_user if hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None
            if user:
                from database.models import TransferCase, RecallCase
                # Считаем только активные (status = 'waiting')
                active_transfers = TransferCase.query.filter_by(
                    user_id=user.id,
                    status='waiting'
                ).count()
                active_recalls = RecallCase.query.filter_by(
                    user_id=user.id,
                    status='waiting'
                ).count()
            
            # Если в БД нет данных, пытаемся из JSON (fallback для старых установок)
            if active_transfers == 0 and active_recalls == 0:
                _, runtime_dir = get_runtime_context()
                
                # Подсчет активных переводов из JSON
                transfers_file = runtime_dir / 'transfer_cases.json'
                if transfers_file.exists():
                    with open(transfers_file, 'r', encoding='utf-8') as f:
                        transfers = json.load(f)
                        # Считаем только активные (status = 'waiting')
                        active_transfers = sum(1 for t in transfers if t.get('status') == 'waiting')
                
                # Подсчет активных перезвонов из JSON
                recalls_file = runtime_dir / 'recall_cases.json'
                if recalls_file.exists():
                    with open(recalls_file, 'r', encoding='utf-8') as f:
                        recalls = json.load(f)
                        # Считаем только активные (status = 'waiting')
                        active_recalls = sum(1 for r in recalls if r.get('status') == 'waiting')
        except Exception as e:
            app.logger.error(f"Ошибка подсчета активных переводов/перезвонов: {e}")

        return jsonify({
            'success': True,
            'generated_at': summary.get('generated_at'),
            'total_calls': total_calls,
            'stations_count': stations_count,
            'active_transfers': active_transfers,
            'active_recalls': active_recalls,
            'ranking': summary.get('ranking', []),
            'stations': summary.get('stations', []),
            'total_questions': summary.get('total_questions', 0)
        })
    except Exception as e:
        app.logger.error(f"Ошибка realtime summary: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/config')
@login_required
def config_page():
    """Страница конфигурации"""
    return render_template('config.html', active_page='config', business_profiles=BUSINESS_PROFILES)

@app.route('/api/config/load')
@login_required
def api_config_load():
    """API для получения персональных настроек пользователя."""
    config_data = get_user_config_data()
    return jsonify(config_data)

@app.route('/settings')
@login_required
def settings_page():
    """Страница настроек пользователя"""
    return render_template('settings.html', active_page='settings')

@app.route('/api/settings/load')
@login_required
def api_settings_load():
    """API для получения данных текущего пользователя."""
    result = {
        'username': current_user.username,
        'email': current_user.email or '',
        'entity_type': 'legal',  # По умолчанию
        'legal_entity': {},
        'physical_entity': {}
    }
    
    # Загружаем данные о типе лица из UserProfileData
    profile_data = UserProfileData.query.filter_by(user_id=current_user.id).first()
    if profile_data:
        result['entity_type'] = profile_data.entity_type or 'legal'
        result['legal_entity'] = {
            'name': profile_data.legal_name or '',
            'inn': profile_data.legal_inn or '',
            'kpp': profile_data.legal_kpp or '',
            'ogrn': profile_data.legal_ogrn or '',
            'legal_address': profile_data.legal_address or '',
            'actual_address': profile_data.actual_address or ''
        }
        result['physical_entity'] = {
            'full_name': profile_data.physical_full_name or '',
            'inn': profile_data.physical_inn or '',
            'passport_series': profile_data.passport_series or '',
            'passport_number': profile_data.passport_number or '',
            'registration_address': profile_data.registration_address or ''
        }
    
    return jsonify(result)

@app.route('/api/settings/update', methods=['POST'])
@login_required
def api_settings_update():
    """API для обновления данных пользователя."""
    try:
        data = request.get_json() or {}
        
        # Обработка данных о типе лица (если присутствуют)
        # Если есть entity_type, обрабатываем только данные о лице и выходим
        if 'entity_type' in data and data.get('entity_type'):
            entity_type = data.get('entity_type')
            legal_entity = data.get('legal_entity', {})
            physical_entity = data.get('physical_entity', {})
            
            # Валидация данных о лице
            if entity_type == 'legal':
                if not legal_entity.get('name', '').strip():
                    return jsonify({'success': False, 'message': 'Название организации обязательно для заполнения'}), 400
                if not legal_entity.get('inn', '').strip():
                    return jsonify({'success': False, 'message': 'ИНН обязателен для заполнения'}), 400
                
                inn = legal_entity.get('inn', '').strip()
                if inn and not re.match(r'^\d{10}$|^\d{12}$', inn):
                    return jsonify({'success': False, 'message': 'ИНН должен содержать 10 или 12 цифр'}), 400
                
                kpp = legal_entity.get('kpp', '').strip()
                if kpp and not re.match(r'^\d{9}$', kpp):
                    return jsonify({'success': False, 'message': 'КПП должен содержать 9 цифр'}), 400
                
                ogrn = legal_entity.get('ogrn', '').strip()
                if ogrn and not re.match(r'^\d{13}$', ogrn):
                    return jsonify({'success': False, 'message': 'ОГРН должен содержать 13 цифр'}), 400
            elif entity_type == 'physical':
                if not physical_entity.get('full_name', '').strip():
                    return jsonify({'success': False, 'message': 'ФИО обязательно для заполнения'}), 400
                
                inn = physical_entity.get('inn', '').strip()
                if inn and not re.match(r'^\d{12}$', inn):
                    return jsonify({'success': False, 'message': 'ИНН физического лица должен содержать 12 цифр'}), 400
            
            # Сохранение данных о лице в UserProfileData
            profile_data = UserProfileData.query.filter_by(user_id=current_user.id).first()
            if not profile_data:
                profile_data = UserProfileData(user_id=current_user.id)
                db.session.add(profile_data)
            
            # Обновляем поля
            profile_data.entity_type = entity_type
            
            # Поля для юридического лица
            profile_data.legal_name = legal_entity.get('name', '').strip() or None
            profile_data.legal_inn = legal_entity.get('inn', '').strip() or None
            profile_data.legal_kpp = legal_entity.get('kpp', '').strip() or None
            profile_data.legal_ogrn = legal_entity.get('ogrn', '').strip() or None
            profile_data.legal_address = legal_entity.get('legal_address', '').strip() or None
            profile_data.actual_address = legal_entity.get('actual_address', '').strip() or None
            
            # Поля для физического лица
            profile_data.physical_full_name = physical_entity.get('full_name', '').strip() or None
            profile_data.physical_inn = physical_entity.get('inn', '').strip() or None
            profile_data.passport_series = physical_entity.get('passport_series', '').strip() or None
            profile_data.passport_number = physical_entity.get('passport_number', '').strip() or None
            profile_data.registration_address = physical_entity.get('registration_address', '').strip() or None
            
            db.session.commit()
            
            return jsonify({'success': True, 'message': 'Данные о лице успешно сохранены'})
        
        # Обработка личных данных (логин, email, пароль)
        # Проверяем, что это действительно запрос на обновление личных данных
        # (должен быть хотя бы username или password)
        if 'username' not in data and 'password' not in data and 'email' not in data:
            return jsonify({'success': False, 'message': 'Не указаны данные для обновления'}), 400
        
        username = data.get('username', '').strip()
        email = (data.get('email') or '').strip() or None
        password = (data.get('password') or '').strip() or None
        
        # Валидация логина (только если он передан в запросе)
        if 'username' in data:
            if not username:
                return jsonify({'success': False, 'message': 'Логин не может быть пустым'}), 400
        
        # Проверка уникальности username (если изменился)
        if username != current_user.username:
            existing_user = User.query.filter_by(username=username).filter(User.id != current_user.id).first()
            if existing_user:
                return jsonify({'success': False, 'message': 'Пользователь с таким логином уже существует'}), 400
        
        # Проверка уникальности email (если изменился и указан)
        if email and email != current_user.email:
            existing_user = User.query.filter_by(email=email).filter(User.id != current_user.id).first()
            if existing_user:
                return jsonify({'success': False, 'message': 'Пользователь с таким email уже существует'}), 400
        
        # Валидация формата email (если указан)
        if email:
            email_regex = r'^[^\s@]+@[^\s@]+\.[^\s@]+$'
            if not re.match(email_regex, email):
                return jsonify({'success': False, 'message': 'Некорректный формат email'}), 400
        
        # Валидация пароля (если указан)
        if password:
            if len(password) < 6:
                return jsonify({'success': False, 'message': 'Пароль должен содержать минимум 6 символов'}), 400
            current_user.set_password(password)
        
        # Обновление данных
        current_user.username = username
        if email is not None:
            current_user.email = email
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Настройки успешно сохранены'})
        
    except Exception as e:
        db.session.rollback()
        logging.error(f'Ошибка обновления настроек пользователя: {e}')
        return jsonify({'success': False, 'message': f'Ошибка сохранения: {str(e)}'}), 500

@app.route('/stations')
@login_required
def stations_page():
    """Страница управления станциями"""
    return render_template('stations.html', active_page='stations')

@app.route('/api/stations')
@login_required
def api_stations():
    """API ??? ????????? ?????? ??????? ????????????."""
    config_data = get_user_config_data()
    stations = config_data.get('stations', {})
    station_chat_ids = config_data.get('station_chat_ids', {})
    station_mapping = config_data.get('station_mapping', {})

    result = []
    for code, name in stations.items():
        result.append({
            'code': code,
            'name': name,
            'chat_ids': station_chat_ids.get(code, []),
            'sub_stations': station_mapping.get(code, [])
        })

    return jsonify(result)

@app.route('/api/stations/save', methods=['POST'])
@login_required
def api_stations_save():
    """API ??? ?????????? ??????? ????????????."""
    try:
        data = request.get_json() or {}
        config_data = get_user_config_data()

        stations = dict(config_data.get('stations') or {})
        station_chat_ids = dict(config_data.get('station_chat_ids') or {})
        station_mapping = dict(config_data.get('station_mapping') or {})

        stations_payload = data.get('stations')
        if stations_payload is not None:
            stations = {
                item['code']: item['name']
                for item in stations_payload
                if item.get('code')
            }

        chat_ids_payload = data.get('station_chat_ids')
        if chat_ids_payload is not None:
            station_chat_ids = {k: v or [] for k, v in chat_ids_payload.items()}

        mapping_payload = data.get('station_mapping')
        if mapping_payload is not None:
            for code, payload in mapping_payload.items():
                if payload is None:
                    stations.pop(code, None)
                    station_chat_ids.pop(code, None)
                    station_mapping.pop(code, None)
                    continue

                name = (payload.get('name') or '').strip()
                if name:
                    stations[code] = name
                elif code not in stations:
                    stations[code] = code

                if 'chat_ids' in payload:
                    station_chat_ids[code] = payload.get('chat_ids') or []

                if 'sub_stations' in payload:
                    station_mapping[code] = payload.get('sub_stations') or []

        if 'nizh_station_codes' in data:
            config_data['nizh_station_codes'] = data.get('nizh_station_codes') or []

        config_data['stations'] = stations
        config_data['station_chat_ids'] = station_chat_ids
        config_data['station_mapping'] = station_mapping

        save_user_config_data(config_data)
        sync_prompts_from_config()
        append_user_log('Станции обновлены', module='stations')
        return jsonify({'success': True, 'message': 'Станции сохранены'})
    except Exception as e:
        app.logger.error(f"Ошибка сохранения станций: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/prompts')
@login_required
def prompts_page():
    """Страница управления промптами"""
    return render_template('prompts.html', active_page='prompts')

@app.route('/api/prompts')
@login_required
def api_prompts():
    """Возвращает промпты пользователя (с учётом файла в профиле)."""
    try:
        sync_prompts_from_config()
        prompts_data = get_user_prompts_data()
        file_data = load_prompts_file()
        changed = False

        if file_data.get('stations'):
            if prompts_data.get('stations') != file_data.get('stations'):
                prompts_data['stations'] = file_data['stations']
                changed = True

        if file_data.get('default') and prompts_data.get('default') != file_data.get('default'):
            prompts_data['default'] = file_data['default']
            changed = True

        if changed:
            save_user_prompts_data(prompts_data)

        return jsonify(prompts_data)
    except Exception as e:
        app.logger.error(f"Ошибка в api_prompts: {e}", exc_info=True)
        # Возвращаем структуру, которую ожидает фронт, но пустую, или ошибку
        return jsonify({'error': str(e)}), 500

@app.route('/api/prompts/save', methods=['POST'])
@login_required
def api_prompts_save():
    """Сохраняет промпты (якоря + станции) и обновляет файл пользователя."""
    try:
        user = current_user if hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None
        data = request.get_json() or {}
        normalized = normalize_prompts_payload(data)
        save_user_prompts_data(normalized, user=user)
        write_prompts_file(normalized, user=user)
        
        # Синхронизация с конфигом (добавление/удаление станций)
        sync_prompts_from_config(user=user)
        
        # Запрос перезапуска воркера для применения изменений
        if user and hasattr(user, 'id'):
            try:
                request_reload(user.id)
                app.logger.info(f"Запрошен перезапуск воркера для пользователя {user.id} после изменения промптов")
            except Exception as exc:
                app.logger.warning(f"Не удалось запросить перезапуск воркера: {exc}")
        
        append_user_log('Промпты обновлены', module='prompts')
        return jsonify({'success': True, 'message': 'Промпты сохранены и применены. Воркер будет перезапущен автоматически.'})
    except Exception as e:
        app.logger.error(f"Ошибка сохранения промптов: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/prompts/sync', methods=['POST'])
@login_required
def api_prompts_sync():
    """Синхронизирует промпты со списком станций и обновляет файл."""
    try:
        user = current_user if hasattr(current_user, 'is_authenticated') and current_user.is_authenticated else None
        
        # Сохраняем данные, пришедшие с фронта, чтобы не потерять изменения до синхронизации
        data = request.get_json() or {}
        payload_before_sync = None
        if data:
            payload_before_sync = normalize_prompts_payload(data)
            save_user_prompts_data(payload_before_sync, user=user)
        else:
            payload_before_sync = get_user_prompts_data(user=user)

        # Независимо от того, были ли изменения в конфиге станций, записываем текущее состояние в файл
        try:
            write_prompts_file(payload_before_sync, user=user)
        except Exception as file_error:
            app.logger.error("Не удалось обновить файл промптов (до синхронизации): %s", file_error)

        if sync_prompts_from_config(user=user):
            # Если синхронизация добавила/удалила станции, повторно пишем файл уже с обновлёнными данными
            updated_data = get_user_prompts_data(user=user)
            try:
                write_prompts_file(updated_data, user=user)
            except Exception as file_error:
                app.logger.error("Не удалось обновить файл промптов (после синхронизации): %s", file_error)
            return jsonify({'success': True, 'message': 'Промпты синхронизированы'})
        
        return jsonify({'success': True, 'message': 'Промпты сохранены (конфиг без изменений)'})
    except Exception as e:
        app.logger.error(f"Ошибка синхронизации промптов: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/script-prompt', methods=['GET'])
@login_required
def api_script_prompt_get():
    """Возвращает script_prompt в формате JSON. Поддерживает параметр file_path."""
    try:
        file_path = request.args.get('file_path')
        
        if file_path:
            # Загружаем из указанного файла
            data = load_script_prompt_file(user=current_user, custom_path=file_path)
        else:
            # Стандартная логика с БД и файлом пользователя
            data = get_user_script_prompt()
            file_data = load_script_prompt_file()
            changed = False

            if file_data.get('checklist') and data.get('checklist') != file_data.get('checklist'):
                data['checklist'] = file_data['checklist']
                changed = True
            if isinstance(file_data.get('prompt'), str) and data.get('prompt') != file_data.get('prompt'):
                data['prompt'] = file_data['prompt']
                changed = True

            if changed:
                save_user_script_prompt(data)

        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/script-prompt/save', methods=['POST'])
@login_required
def api_script_prompt_save():
    """Сохраняет script_prompt. Поддерживает параметр file_path."""
    try:
        payload = request.get_json() or {}
        data = payload.get('data', {})
        file_path = payload.get('file_path')
        
        if file_path:
            # Сохраняем только в указанный файл
            write_script_prompt_file(data, user=current_user, custom_path=file_path)
        else:
            # Сохраняем в БД и в стандартный файл пользователя
            save_user_script_prompt(data)
            write_script_prompt_file(data)
            
        append_user_log('Чек-лист сохранен', module='prompts')
        return jsonify({'success': True, 'message': 'Чек-лист сохранен'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/generate-prompt', methods=['POST'])
@login_required
def api_generate_prompt():
    """API для автогенерации промпта через DeepSeek"""
    try:
        data = request.get_json()
        title = (data.get('title') or '').strip()
        existing_prompt = (data.get('existingPrompt') or '').strip()

        if not title:
            return jsonify({'success': False, 'message': 'Название пункта не может быть пустым'})

        # Генерируем короткий промпт в формате двух предложений
        context = f"""Сгенерируй КОРОТКИЙ текстовый промпт для пункта чек-листа: "{title}".
Формат ответа ДОЛЖЕН быть строго таким (ровно одно-два предложения в одну строку, без переносов и списков):
Считать [ОТВЕТ: ДА], если ... . Считать [ОТВЕТ: НЕТ], если ... .
Опиши критерии простым языком, без лишних деталей и без кавычек вокруг всего текста. Не добавляй ничего кроме этих двух предложений."""

        import requests
        runtime_cfg = build_user_runtime_config()
        api_keys = runtime_cfg.get('api_keys', {})
        thebai_api_key = api_keys.get('thebai_api_key')
        thebai_url = api_keys.get('thebai_url') or 'https://api.deepseek.com/v1/chat/completions'
        thebai_model = api_keys.get('thebai_model') or 'deepseek-reasoner'

        if not thebai_api_key:
            return jsonify({'success': False, 'message': 'Укажите TheB.ai API key в настройках профиля'}), 400

        config_data = get_user_config_data()
        profile = config_data.get('business_profile', 'autoservice')
        profile_expert = {
            'autoservice': 'в автосервисе',
            'restaurant': 'в ресторанах и общепите',
            'dental': 'в стоматологических клиниках',
            'retail': 'в розничной торговле',
            'medical': 'в медицинских центрах',
            'universal': 'в любой сфере услуг',
        }.get(profile, 'в любой сфере услуг')

        prompt_request = {
            "model": "deepseek-chat", # Используем deepseek-chat (V3) для скорости и стабильности
            "messages": [
                {
                    "role": "system",
                    "content": f"Ты эксперт по созданию инструкций для оценки качества телефонных звонков {profile_expert}. Твоя задача - создавать четкие, реалистичные и подробные критерии оценки. ВАЖНО: Верни ТОЛЬКО текст критерия. Не добавляй никаких вводных слов, пояснений или markdown-разметки.",
                },
                {
                    "role": "user",
                    "content": context
                },
            ],
            "temperature": 0.3,
            "max_tokens": 800,  # Увеличено для подробных промптов
        }

        # Retry логика для надежности
        max_retries = 3
        timeout = 90  # Увеличенный таймаут для генерации промпта
        response = None
        last_error = None
        
        for attempt in range(max_retries):
            try:
                response = requests.post(
                    thebai_url,
                    headers={
                        'Authorization': f'Bearer {thebai_api_key}',
                        'Content-Type': 'application/json',
                    },
                    json=prompt_request,
                    timeout=timeout,
                )
                
                if response.status_code == 200:
                    result = response.json()
                    app.logger.info(f"Ответ API: {result}")
                    
                    # Проверяем структуру ответа
                    if not result.get('choices') or len(result['choices']) == 0:
                        app.logger.error("Пустой ответ от API")
                        return jsonify({'success': False, 'message': 'Пустой ответ от API'})
                    
                    generated_prompt = (result.get('choices', [{}])[0]
                                            .get('message', {})
                                            .get('content', '')).strip()
                    
                    if not generated_prompt:
                        app.logger.error("Пустой промпт в ответе")
                        return jsonify({'success': False, 'message': 'Пустой промпт в ответе'})
                    
                    # Удаляем только внешние кавычки, если они есть
                    if generated_prompt.startswith('"') and generated_prompt.endswith('"'):
                        generated_prompt = generated_prompt[1:-1].strip()
                    elif generated_prompt.startswith("'") and generated_prompt.endswith("'"):
                        generated_prompt = generated_prompt[1:-1].strip()
                    
                    app.logger.info(f"Сгенерирован промпт для '{title}': {generated_prompt}")
                    return jsonify({'success': True, 'prompt': generated_prompt})
                
                # Если статус не 200, логируем и пробуем снова (если не последняя попытка)
                app.logger.warning(f"Попытка {attempt + 1}/{max_retries}: Ошибка DeepSeek API: {response.status_code} - {response.text}")
                if attempt < max_retries - 1:
                    time.sleep(2)  # Небольшая задержка перед повтором
                    continue
                else:
                    return jsonify({'success': False, 'message': f'Ошибка API: {response.status_code}'}), 502
                    
            except requests.exceptions.Timeout as e:
                last_error = f"Таймаут запроса (>{timeout}с)"
                app.logger.warning(f"Попытка {attempt + 1}/{max_retries}: {last_error}")
                if attempt < max_retries - 1:
                    time.sleep(2)
                    continue
            except requests.exceptions.RequestException as e:
                last_error = f"Ошибка сети: {str(e)}"
                app.logger.warning(f"Попытка {attempt + 1}/{max_retries}: {last_error}")
                if attempt < max_retries - 1:
                    time.sleep(2)
                    continue
        
        # Если все попытки исчерпаны
        error_message = last_error or "Неизвестная ошибка"
        app.logger.error(f"Все попытки исчерпаны. Последняя ошибка: {error_message}")
        return jsonify({'success': False, 'message': f'Ошибка генерации после {max_retries} попыток: {error_message}'}), 500

    except Exception as e:
        app.logger.error(f"Ошибка генерации промпта: {e}")
        return jsonify({'success': False, 'message': f'Ошибка генерации: {str(e)}'}), 500

@app.route('/api/generate-anchor-prompt', methods=['POST'])
@login_required
def api_generate_anchor_prompt():
    """API для генерации полного промпта якоря на основе описания задачи"""
    try:
        data = request.get_json()
        user_intent = (data.get('intent') or '').strip()

        if not user_intent:
            return jsonify({'success': False, 'message': 'Описание задачи не может быть пустым'}), 400

        config_data = get_user_config_data()
        profile = config_data.get('business_profile', 'autoservice')
        profile_expert = {
            'autoservice': 'в автосервисе',
            'restaurant': 'в ресторанах и общепите',
            'dental': 'в стоматологических клиниках',
            'retail': 'в розничной торговле',
            'medical': 'в медицинских центрах',
            'universal': 'в любой сфере услуг',
        }.get(profile, 'в любой сфере услуг')

        # Системный промпт для генерации промпта
        system_instruction = f"""Ты — эксперт по созданию системных промптов (инструкций) для AI-анализа телефонных разговоров {profile_expert}.
Твоя задача: На основе описания пользователя создать подробный, структурированный и эффективный промпт, который будет использоваться другой нейросетью для анализа диалогов.

Требования к генерируемому промпту:
1. Он должен быть написан от лица инструктора к исполнителю (AI-аналитику).
2. Он должен включать четкие критерии классификации звонков (Тип, Класс, Результат), основанные на пожеланиях пользователя.
3. Он должен требовать СТРОГИЙ формат вывода с тегами, так как это используется для автоматического парсинга.
4. Обязательная структура ответа (тегов), которую ты должен включить в генерируемый промпт:
   - [ТИПЗВОНКА:ЦЕЛЕВОЙ] или [ТИПЗВОНКА:НЕЦЕЛЕВОЙ]
   - [КЛАСС:A], [КЛАСС:B] и т.д. (если применимо)
   - [РЕЗУЛЬТАТ:...] (например, ЗАПИСЬ, ОТКАЗ, ПЕРЕЗВОНИТЬ, КОНСУЛЬТАЦИЯ и т.д.)
   - Краткое объяснение после тегов.
5. ВАЖНО: Верни ТОЛЬКО текст сгенерированного промпта. Не добавляй никаких вводных слов, пояснений или markdown-разметки (никаких ```yaml или ```). Только чистый текст промпта.

Контекст задачи пользователя:
"""

        import requests
        runtime_cfg = build_user_runtime_config()
        api_keys = runtime_cfg.get('api_keys', {})
        thebai_api_key = api_keys.get('thebai_api_key')
        thebai_url = api_keys.get('thebai_url') or 'https://api.deepseek.com/v1/chat/completions'
        # Используем deepseek-chat (V3) для более предсказуемого формата текста
        thebai_model = 'deepseek-chat'

        if not thebai_api_key:
            return jsonify({'success': False, 'message': 'Укажите DeepSeek/TheB.ai API key в настройках профиля'}), 400

        prompt_request = {
            "model": thebai_model,
            "messages": [
                {
                    "role": "system",
                    "content": system_instruction
                },
                {
                    "role": "user",
                    "content": f"Создай промпт для следующей задачи: {user_intent}"
                },
            ],
            "temperature": 0.6, # Чуть выше для креативности в инструкциях
        }

        # Retry логика для надежности
        max_retries = 3
        timeout = 120  # Увеличенный таймаут для генерации большого промпта
        response = None
        last_error = None
        
        for attempt in range(max_retries):
            try:
                response = requests.post(
                    thebai_url,
                    headers={
                        'Authorization': f'Bearer {thebai_api_key}',
                        'Content-Type': 'application/json',
                    },
                    json=prompt_request,
                    timeout=timeout,
                )
                break  # Успешный запрос, выходим из цикла
            except (requests.exceptions.Timeout, requests.exceptions.ConnectionError, requests.exceptions.RequestException) as e:
                last_error = e
                if attempt < max_retries - 1:
                    wait_time = (attempt + 1) * 2  # Экспоненциальная задержка: 2, 4, 6 секунд
                    app.logger.warning(f"Попытка {attempt + 1}/{max_retries} не удалась, повтор через {wait_time} сек: {e}")
                    time.sleep(wait_time)
                else:
                    app.logger.error(f"Все {max_retries} попытки не удались: {e}")
                    raise
        
        if not response:
            return jsonify({'success': False, 'message': 'Не удалось получить ответ от API после всех попыток'}), 503
        
        if response.status_code == 200:
            result = response.json()
            if not result.get('choices') or len(result['choices']) == 0:
                return jsonify({'success': False, 'message': 'Пустой ответ от API'})
            
            generated_prompt = (result.get('choices', [{}])[0]
                                    .get('message', {})
                                    .get('content', '')).strip()
            
            # Убираем markdown code blocks если они есть
            if generated_prompt.startswith("```") and generated_prompt.endswith("```"):
                lines = generated_prompt.split('\n')
                if lines[0].startswith("```"):
                    lines = lines[1:]
                if lines[-1].startswith("```"):
                    lines = lines[:-1]
                generated_prompt = '\n'.join(lines)
            elif generated_prompt.startswith("```"):
                 generated_prompt = generated_prompt.replace("```yaml", "").replace("```", "")
            
            return jsonify({'success': True, 'prompt': generated_prompt.strip()})

        app.logger.error(f"Ошибка API при генерации якоря: {response.status_code} - {response.text}")
        return jsonify({'success': False, 'message': f'Ошибка API: {response.status_code}'}), 502

    except requests.exceptions.Timeout:
        app.logger.error("Таймаут при генерации якоря (превышено время ожидания ответа от API)")
        return jsonify({'success': False, 'message': 'Превышено время ожидания ответа от API. Попробуйте еще раз.'}), 504
    except requests.exceptions.ConnectionError as e:
        app.logger.error(f"Ошибка соединения при генерации якоря: {e}")
        return jsonify({'success': False, 'message': 'Ошибка соединения с API. Проверьте интернет-соединение и попробуйте еще раз.'}), 503
    except Exception as e:
        app.logger.error(f"Ошибка генерации якоря: {e}", exc_info=True)
        error_msg = str(e)
        if 'timeout' in error_msg.lower() or 'timed out' in error_msg.lower():
            return jsonify({'success': False, 'message': 'Превышено время ожидания ответа от API. Попробуйте еще раз.'}), 504
        return jsonify({'success': False, 'message': f'Ошибка генерации: {error_msg}'}), 500

@app.route('/api/regenerate-anchor-prompt', methods=['POST'])
@login_required
def api_regenerate_anchor_prompt():
    """API для доработки существующего промпта якоря с добавлением нового условия"""
    try:
        data = request.get_json()
        current_prompt = (data.get('current_prompt') or '').strip()
        additional_condition = (data.get('additional_condition') or '').strip()

        if not current_prompt:
            return jsonify({'success': False, 'message': 'Текущий промпт не может быть пустым'}), 400
        
        if not additional_condition:
            return jsonify({'success': False, 'message': 'Дополнительное условие не может быть пустым'}), 400

        config_data = get_user_config_data()
        profile = config_data.get('business_profile', 'autoservice')
        profile_expert = {
            'autoservice': 'в автосервисе',
            'restaurant': 'в ресторанах и общепите',
            'dental': 'в стоматологических клиниках',
            'retail': 'в розничной торговле',
            'medical': 'в медицинских центрах',
            'universal': 'в любой сфере услуг',
        }.get(profile, 'в любой сфере услуг')

        # Системный промпт для доработки промпта
        system_instruction = f"""Ты — эксперт по улучшению системных промптов (инструкций) для AI-анализа телефонных разговоров {profile_expert}.
Твоя задача: Доработать существующий промпт, добавив в него новое условие или требование, указанное пользователем.

Важные требования:
1. Сохрани всю структуру и логику существующего промпта.
2. Интегрируй новое условие естественным образом, не нарушая существующую логику.
3. Если новое условие конфликтует со старым, приоритет отдай новому, но сохрани остальные части промпта.
4. Промпт должен оставаться структурированным и четким.
5. Обязательно сохрани формат вывода с тегами [ТИПЗВОНКА:...], [КЛАСС:...], [РЕЗУЛЬТАТ:...] если они были в оригинале.
6. ВАЖНО: Верни ТОЛЬКО текст доработанного промпта. Не добавляй никаких вводных слов, пояснений или markdown-разметки (никаких ```yaml или ```). Только чистый текст промпта."""

        import requests
        runtime_cfg = build_user_runtime_config()
        api_keys = runtime_cfg.get('api_keys', {})
        thebai_api_key = api_keys.get('thebai_api_key')
        thebai_url = api_keys.get('thebai_url') or 'https://api.deepseek.com/v1/chat/completions'
        # Используем deepseek-chat (V3) для более предсказуемого формата текста
        thebai_model = 'deepseek-chat'

        if not thebai_api_key:
            return jsonify({'success': False, 'message': 'Укажите DeepSeek/TheB.ai API key в настройках профиля'}), 400

        user_message = f"""Текущий промпт:
{current_prompt}

Новое условие или требование, которое нужно добавить:
{additional_condition}

Доработай промпт, интегрировав новое условие в существующую структуру."""

        prompt_request = {
            "model": thebai_model,
            "messages": [
                {
                    "role": "system",
                    "content": system_instruction
                },
                {
                    "role": "user",
                    "content": user_message
                },
            ],
            "temperature": 0.3,
        }

        # Retry логика для надежности
        max_retries = 3
        timeout = 120
        response = None
        last_error = None
        
        for attempt in range(max_retries):
            try:
                response = requests.post(
                    thebai_url,
                    headers={
                        'Authorization': f'Bearer {thebai_api_key}',
                        'Content-Type': 'application/json',
                    },
                    json=prompt_request,
                    timeout=timeout,
                )
                break
            except (requests.exceptions.Timeout, requests.exceptions.ConnectionError, requests.exceptions.RequestException) as e:
                last_error = e
                if attempt < max_retries - 1:
                    wait_time = (attempt + 1) * 2
                    app.logger.warning(f"Попытка {attempt + 1}/{max_retries} не удалась, повтор через {wait_time} сек: {e}")
                    time.sleep(wait_time)
                else:
                    app.logger.error(f"Все {max_retries} попытки не удались: {e}")
                    raise
        
        if not response:
            return jsonify({'success': False, 'message': 'Не удалось получить ответ от API после всех попыток'}), 503
        
        if response.status_code == 200:
            result = response.json()
            if not result.get('choices') or len(result['choices']) == 0:
                return jsonify({'success': False, 'message': 'Пустой ответ от API'})
            
            regenerated_prompt = (result.get('choices', [{}])[0]
                                    .get('message', {})
                                    .get('content', '')).strip()
            
            # Убираем markdown code blocks если они есть
            if regenerated_prompt.startswith("```") and regenerated_prompt.endswith("```"):
                lines = regenerated_prompt.split('\n')
                if lines[0].startswith("```"):
                    lines = lines[1:]
                if lines[-1].startswith("```"):
                    lines = lines[:-1]
                regenerated_prompt = '\n'.join(lines)
            elif regenerated_prompt.startswith("```"):
                 regenerated_prompt = regenerated_prompt.replace("```yaml", "").replace("```", "")
            
            return jsonify({'success': True, 'prompt': regenerated_prompt.strip()})

        app.logger.error(f"Ошибка API при перегенерации якоря: {response.status_code} - {response.text}")
        return jsonify({'success': False, 'message': f'Ошибка API: {response.status_code}'}), 502

    except requests.exceptions.Timeout:
        app.logger.error("Таймаут при перегенерации якоря (превышено время ожидания ответа от API)")
        return jsonify({'success': False, 'message': 'Превышено время ожидания ответа от API. Попробуйте еще раз.'}), 504
    except requests.exceptions.ConnectionError as e:
        app.logger.error(f"Ошибка соединения при перегенерации якоря: {e}")
        return jsonify({'success': False, 'message': 'Ошибка соединения с API. Проверьте интернет-соединение и попробуйте еще раз.'}), 503
    except Exception as e:
        app.logger.error(f"Ошибка перегенерации якоря: {e}", exc_info=True)
        error_msg = str(e)
        if 'timeout' in error_msg.lower() or 'timed out' in error_msg.lower():
            return jsonify({'success': False, 'message': 'Превышено время ожидания ответа от API. Попробуйте еще раз.'}), 504
        return jsonify({'success': False, 'message': f'Ошибка перегенерации: {error_msg}'}), 500

@app.route('/api/regenerate-checklist-item-prompt', methods=['POST'])
@login_required
def api_regenerate_checklist_item_prompt():
    """API для доработки промпта пункта чек-листа с добавлением нового условия"""
    try:
        data = request.get_json()
        current_prompt = (data.get('current_prompt') or '').strip()
        additional_condition = (data.get('additional_condition') or '').strip()
        item_title = (data.get('item_title') or '').strip()

        if not current_prompt:
            return jsonify({'success': False, 'message': 'Текущий промпт не может быть пустым'}), 400
        
        if not additional_condition:
            return jsonify({'success': False, 'message': 'Дополнительное условие не может быть пустым'}), 400

        config_data = get_user_config_data()
        profile = config_data.get('business_profile', 'autoservice')
        profile_expert = {
            'autoservice': 'в автосервисе',
            'restaurant': 'в ресторанах и общепите',
            'dental': 'в стоматологических клиниках',
            'retail': 'в розничной торговле',
            'medical': 'в медицинских центрах',
            'universal': 'в любой сфере услуг',
        }.get(profile, 'в любой сфере услуг')

        # Системный промпт для доработки промпта пункта чек-листа
        system_instruction = f"""Ты — эксперт по улучшению промптов для оценки телефонных разговоров {profile_expert} по чек-листам.
Твоя задача: Доработать существующий промпт для конкретного пункта чек-листа, добавив в него новое условие или требование, указанное пользователем.

Важные требования:
1. Сохрани всю структуру и логику существующего промпта.
2. Интегрируй новое условие естественным образом, не нарушая существующую логику.
3. Если новое условие конфликтует со старым, приоритет отдай новому, но сохрани остальные части промпта.
4. Промпт должен оставаться четким и конкретным - это критерий оценки для одного пункта чек-листа.
5. Промпт должен быть сфокусирован на проверке выполнения конкретного критерия.
6. Сохрани формат и стиль оригинального промпта.
7. ВАЖНО: Верни ТОЛЬКО текст доработанного промпта. Не добавляй никаких вводных слов, пояснений или markdown-разметки (никаких ```yaml или ```). Только чистый текст промпта."""

        import requests
        runtime_cfg = build_user_runtime_config()
        api_keys = runtime_cfg.get('api_keys', {})
        thebai_api_key = api_keys.get('thebai_api_key')
        thebai_url = api_keys.get('thebai_url') or 'https://api.deepseek.com/v1/chat/completions'
        # Используем deepseek-chat (V3) вместо reasoner для более предсказуемого формата текста
        thebai_model = 'deepseek-chat' 

        if not thebai_api_key:
            return jsonify({'success': False, 'message': 'Укажите DeepSeek/TheB.ai API key в настройках профиля'}), 400
        
        user_message = f"""Текущий промпт для пункта чек-листа:
{current_prompt}

Новое условие или требование, которое нужно добавить:
{additional_condition}

Доработай промпт, интегрировав новое условие в существующую структуру."""

        prompt_request = {
            "model": thebai_model,
            "messages": [
                {
                    "role": "system",
                    "content": system_instruction
                },
                {
                    "role": "user",
                    "content": user_message
                },
            ],
            "temperature": 0.3, # Понижаем температуру для большей стабильности
        }

        # Retry логика для надежности
        max_retries = 3
        timeout = 120
        response = None
        last_error = None
        
        for attempt in range(max_retries):
            try:
                response = requests.post(
                    thebai_url,
                    headers={
                        'Authorization': f'Bearer {thebai_api_key}',
                        'Content-Type': 'application/json',
                    },
                    json=prompt_request,
                    timeout=timeout,
                )
                break
            except (requests.exceptions.Timeout, requests.exceptions.ConnectionError, requests.exceptions.RequestException) as e:
                last_error = e
                if attempt < max_retries - 1:
                    wait_time = (attempt + 1) * 2
                    app.logger.warning(f"Попытка {attempt + 1}/{max_retries} не удалась, повтор через {wait_time} сек: {e}")
                    time.sleep(wait_time)
                else:
                    app.logger.error(f"Все {max_retries} попытки не удались: {e}")
                    raise
        
        if not response:
            return jsonify({'success': False, 'message': 'Не удалось получить ответ от API после всех попыток'}), 503
        
        if response.status_code == 200:
            result = response.json()
            if not result.get('choices') or len(result['choices']) == 0:
                return jsonify({'success': False, 'message': 'Пустой ответ от API'})
            
            regenerated_prompt = (result.get('choices', [{}])[0]
                                    .get('message', {})
                                    .get('content', '')).strip()
            
            # Убираем markdown code blocks если они есть
            if regenerated_prompt.startswith("```") and regenerated_prompt.endswith("```"):
                lines = regenerated_prompt.split('\n')
                if lines[0].startswith("```"):
                    lines = lines[1:]
                if lines[-1].startswith("```"):
                    lines = lines[:-1]
                regenerated_prompt = '\n'.join(lines)
            elif regenerated_prompt.startswith("```"):
                 regenerated_prompt = regenerated_prompt.replace("```yaml", "").replace("```", "")
            
            return jsonify({'success': True, 'prompt': regenerated_prompt.strip()})

        app.logger.error(f"Ошибка API при доработке пункта чек-листа: {response.status_code} - {response.text}")
        return jsonify({'success': False, 'message': f'Ошибка API: {response.status_code}'}), 502

    except requests.exceptions.Timeout:
        app.logger.error("Таймаут при доработке пункта чек-листа")
        return jsonify({'success': False, 'message': 'Превышено время ожидания ответа от API. Попробуйте еще раз.'}), 504
    except requests.exceptions.ConnectionError as e:
        app.logger.error(f"Ошибка соединения при доработке пункта чек-листа: {e}")
        return jsonify({'success': False, 'message': 'Ошибка соединения с API. Проверьте интернет-соединение и попробуйте еще раз.'}), 503
    except Exception as e:
        app.logger.error(f"Ошибка доработки пункта чек-листа: {e}", exc_info=True)
        error_msg = str(e)
        if 'timeout' in error_msg.lower() or 'timed out' in error_msg.lower():
            return jsonify({'success': False, 'message': 'Превышено время ожидания ответа от API. Попробуйте еще раз.'}), 504
        return jsonify({'success': False, 'message': f'Ошибка доработки: {error_msg}'}), 500

@app.route('/vocabulary')
@login_required
def vocabulary_page():
    """Страница управления словарем"""
    return render_template('vocabulary.html', active_page='vocabulary')

@app.route('/checklists')
@login_required
def checklists_page():
    """Страница настройки чек-листов"""
    return render_template('checklists.html', active_page='checklists')

@app.route('/api/checklists/meta')
@login_required
def api_checklists_meta():
    """Возвращает данные для страницы тестирования чек-листов."""
    runtime_cfg = build_user_runtime_config()
    stations_dict = runtime_cfg.get('stations', {}) or {}
    stations = [{'code': code, 'name': name} for code, name in stations_dict.items()]
    script_prompt_path = (
        runtime_cfg.get('paths', {}).get('script_prompt_file')
        or str(getattr(project_config, 'SCRIPT_PROMPT_8_PATH', ''))
    )
    return jsonify({
        'script_prompt_file': script_prompt_path,
        'stations': stations
    })

@app.route('/api/checklists/test', methods=['POST'])
@login_required
def api_checklists_test():
    """Прогон загруженного файла через транскрибацию и чек-лист без Telegram уведомлений."""
    if not transcribe_audio_with_internal_service or not thebai_analyze or not save_transcript_analysis:
        return jsonify({'success': False, 'message': 'Модуль обработки недоступен на сервере.'}), 500

    upload = request.files.get('file')
    if not upload or not upload.filename:
        return jsonify({'success': False, 'message': 'Файл не выбран.'}), 400

    station_code = (request.form.get('station_code') or '').strip()
    script_prompt_override = (request.form.get('script_prompt_file') or '').strip()

    runtime_cfg = build_user_runtime_config()
    # Локально отключаем Telegram, чтобы тесты чек-листов не слали уведомления
    runtime_cfg = deepcopy(runtime_cfg)
    api_keys = runtime_cfg.get('api_keys') or {}
    api_keys['telegram_bot_token'] = ''
    runtime_cfg['api_keys'] = api_keys
    telegram_cfg = runtime_cfg.get('telegram') or {}
    telegram_cfg['alert_chat_id'] = ''
    telegram_cfg['tg_channel_nizh'] = ''
    telegram_cfg['tg_channel_other'] = ''
    telegram_cfg['reports_chat_id'] = ''
    runtime_cfg['telegram'] = telegram_cfg
    if script_prompt_override:
        runtime_cfg.setdefault('paths', {})
        runtime_cfg['paths']['script_prompt_file'] = script_prompt_override

    try:
        with legacy_config_override(runtime_cfg):
            # Куда складываем тестовые файлы и результаты
            paths_cfg = runtime_cfg.get('paths', {}) or {}
            base_records = paths_cfg.get('base_records_path') or getattr(project_config, 'BASE_RECORDS_PATH', Path.cwd())
            base_records = Path(base_records)

            # Отдельная папка testscalls в корне профиля пользователя, чтобы не мешать боевой обработке
            today_subdir = datetime.now().strftime("%Y/%m/%d")
            test_root = base_records / "testscalls" / today_subdir
            save_dir = test_root / datetime.now().strftime("%H%M%S_%f")
            save_dir.mkdir(parents=True, exist_ok=True)

            safe_name = secure_filename(upload.filename)
            if not safe_name:
                safe_name = f"test_{int(time.time())}.wav"
            saved_path = save_dir / safe_name
            upload.save(saved_path)

            # Метаданные из имени файла
            phone_number = ''
            call_time = datetime.now()
            parsed_station = None
            if parse_filename:
                try:
                    phone_number, parsed_station, parsed_dt = parse_filename(saved_path.name)
                    if parsed_dt:
                        call_time = parsed_dt
                except Exception as exc:
                    app.logger.warning("Не удалось распарсить имя файла для теста: %s", exc)

            if not station_code:
                station_code = parsed_station or 'test'

            # Транскрипция
            stereo_mode = bool((runtime_cfg.get('transcription') or {}).get('tbank_stereo_enabled', False))
            additional_vocab = load_additional_vocab() if load_additional_vocab else []
            transcript_text = transcribe_audio_with_internal_service(
                saved_path,
                stereo_mode=stereo_mode,
                additional_vocab=additional_vocab,
                timeout=int(os.getenv("CHECKLIST_TEST_TRANSCRIPTION_TIMEOUT", "600") or "600")
            )
            if not transcript_text:
                return jsonify({'success': False, 'message': 'Не удалось получить транскрипт.'}), 500

            # Анализ основным промптом
            prompt_text = ''
            anchor_name = None
            try:
                # Получаем промпты пользователя для определения имени якоря
                user_prompts = get_user_prompts_data(user=current_user)
                anchors = user_prompts.get('anchors') or {}
                
                prompts = get_station_prompts() if get_station_prompts else {}
                prompt_text = prompts.get(str(station_code)) or prompts.get('default') or ''
                
                # Пытаемся найти имя якоря по тексту промпта
                for name, text in anchors.items():
                    if text == prompt_text:
                        anchor_name = name
                        break
            except Exception as exc:
                app.logger.error("Не удалось получить промпт станции: %s", exc)

            if not prompt_text:
                prompt_text = "Проанализируй звонок и определи результат разговора."

            analysis_text = thebai_analyze(transcript_text, prompt_text)
            
            # Определяем целевой ли звонок (по аналогии с call_handler.py)
            analysis_clean = analysis_text.upper().replace(" ", "")
            is_target = "[ТИПЗВОНКА:ЦЕЛЕВОЙ]" in analysis_clean or "ПЕРВИЧНЫЙ" in analysis_clean

            # Сохраняем транскрипцию и анализ в отдельную тестовую папку,
            # не используя стандартный путь transcriptions, чтобы не мешать рабочим данным
            test_trans_dir = save_dir / "transcriptions"
            test_trans_dir.mkdir(parents=True, exist_ok=True)
            transcript_file = test_trans_dir / f"{saved_path.stem}.txt"
            try:
                with transcript_file.open("w", encoding="utf-8") as f:
                    f.write("Диалог:\n\n")
                    f.write(transcript_text)
                    f.write("\n\nАнализ:\n\n")
                    f.write(analysis_text)
            except Exception as exc:
                app.logger.error("Не удалось сохранить тестовый файл транскрипции: %s", exc)
                return jsonify({'success': False, 'message': 'Не удалось сохранить результаты теста.'}), 500

            checklist_analysis_path = None
            checklist_payload = None
            try:
                result = run_exental_alert(
                    str(transcript_file),
                    station_code=str(station_code),
                    phone_number=phone_number or '',
                    date_str=call_time.strftime("%Y-%m-%d %H:%M:%S"),
                    operator_station_code=station_code
                )
                if isinstance(result, (list, tuple)) and len(result) == 4:
                    caption, analysis_full, qa_text, overall = result
                    checklist_payload = {
                        'caption': caption,
                        'qa_text': _dedupe_checklist_qa_text(qa_text),
                        'overall': overall
                    }
                # Файл чек-листа создаётся рядом в script_8/
                checklist_analysis_path = Path(transcript_file).parent / "script_8" / f"{Path(transcript_file).stem}_analysis.txt"
            except Exception as exc:
                app.logger.error("Ошибка при разборе чек-листа: %s", exc, exc_info=True)

            return jsonify({
                'success': True,
                'message': 'Файл обработан. Результаты сохранены в тестовую папку.',
                'paths': {
                    'audio': str(saved_path),
                    'transcript': str(transcript_file),
                    'analysis': str(transcript_file),
                    'checklist': str(checklist_analysis_path) if checklist_analysis_path else None
                },
                'checklist': checklist_payload,
                'anchor': {
                    'analysis': analysis_text,
                    'prompt': prompt_text,
                    'anchor_name': anchor_name,
                    'station_code': station_code,
                    'is_target': is_target
                },
                'transcript_text': transcript_text  # Для последующего анализа
            })
    except Exception as exc:
        app.logger.error("Сбой при тестировании чек-листа: %s", exc, exc_info=True)
        return jsonify({'success': False, 'message': f'Ошибка тестирования: {exc}'}), 500

@app.route('/api/checklists/analyze', methods=['POST'])
@login_required
def api_checklists_analyze():
    """Детальный анализ разбора каждого пункта чек-листа с обоснованием."""
    data = request.get_json() or {}
    transcript_text = data.get('transcript_text', '').strip()
    qa_text = data.get('qa_text', '').strip()
    
    if not transcript_text or not qa_text:
        return jsonify({'success': False, 'message': 'Не переданы данные транскрипта или чек-листа.'}), 400
    
    # Формируем промпт для детального анализа каждого пункта
    analysis_prompt = f"""Ты — эксперт по анализу клиентских звонков. Перед тобой транскрипт звонка и результат разбора по чек-листу.

ТРАНСКРИПТ ЗВОНКА:
{transcript_text}

РЕЗУЛЬТАТ РАЗБОРА ПО ЧЕК-ЛИСТУ:
{qa_text}

Твоя задача: для КАЖДОГО пункта чек-листа написать детальное обоснование, ПОЧЕМУ этот пункт был оценен именно так (ДА или НЕТ).

В обосновании укажи:
- Цитаты или парафразы из транскрипта, которые подтверждают или опровергают выполнение пункта
- Краткое объяснение логики оценки
- Если пункт НЕ выполнен — что именно отсутствовало

Формат ответа — строго для каждого пункта по порядку:

1. [Детальное обоснование для пункта 1]

2. [Детальное обоснование для пункта 2]

...

Обоснования должны быть конкретными и ссылаться на реальные фрагменты диалога."""

    try:
        runtime_cfg = build_user_runtime_config()
        thebai_api_key = (runtime_cfg.get('api_keys') or {}).get('thebai_api_key', '')
        thebai_url = (runtime_cfg.get('api_keys') or {}).get('thebai_url', 'https://api.deepseek.com/v1/chat/completions')
        
        if not thebai_api_key:
            return jsonify({'success': False, 'message': 'API ключ для анализа не настроен.'}), 400
        
        # Отправляем запрос к TheB.ai / DeepSeek
        payload = {
            "model": (runtime_cfg.get('api_keys') or {}).get('thebai_model', 'deepseek-reasoner'),
            "messages": [
                {"role": "system", "content": "Ты эксперт по анализу клиентских звонков."},
                {"role": "user", "content": analysis_prompt}
            ],
            "temperature": 0.3,
        }
        
        response = requests.post(
            thebai_url,
            headers={
                'Authorization': f'Bearer {thebai_api_key}',
                'Content-Type': 'application/json',
            },
            json=payload,
            timeout=120
        )
        
        if response.status_code != 200:
            app.logger.error(f"Ошибка API при анализе чек-листа: {response.status_code} - {response.text}")
            return jsonify({'success': False, 'message': f'Ошибка API: {response.status_code}'}), 502
        
        result = response.json()
        analysis_result = (result.get('choices', [{}])[0]
                          .get('message', {})
                          .get('content', '')).strip()
        
        if not analysis_result:
            return jsonify({'success': False, 'message': 'Пустой ответ от API анализа.'}), 500
        
        # Парсим результат: разбиваем по номерам пунктов
        explanations = _parse_checklist_explanations(analysis_result)
        
        return jsonify({
            'success': True,
            'explanations': explanations,
            'raw_analysis': analysis_result
        })
        
    except requests.exceptions.Timeout:
        app.logger.error("Таймаут при анализе чек-листа")
        return jsonify({'success': False, 'message': 'Превышено время ожидания ответа от API.'}), 504
    except Exception as exc:
        app.logger.error(f"Ошибка анализа чек-листа: {exc}", exc_info=True)
        return jsonify({'success': False, 'message': f'Ошибка анализа: {exc}'}), 500

def _parse_checklist_explanations(analysis_text):
    """Парсит текст анализа и возвращает dict {номер_пункта: обоснование}"""
    explanations = {}
    lines = analysis_text.split('\n')
    
    current_num = None
    current_text = []
    
    for line in lines:
        line = line.strip()
        # Ищем строки вида "1. текст" или "1) текст"
        match = re.match(r'^(\d+)[\.\)]\s*(.*)$', line)
        if match:
            # Сохраняем предыдущий блок
            if current_num is not None and current_text:
                explanations[str(current_num)] = '\n'.join(current_text).strip()
            # Начинаем новый блок
            current_num = match.group(1)
            rest = match.group(2).strip()
            current_text = [rest] if rest else []
        elif current_num is not None and line:
            # Продолжение текущего пункта
            current_text.append(line)
    
    # Сохраняем последний блок
    if current_num is not None and current_text:
        explanations[str(current_num)] = '\n'.join(current_text).strip()
    
    return explanations

@app.route('/api/anchors/analyze', methods=['POST'])
@login_required
def api_anchors_analyze():
    """Детальный анализ результата якоря с обоснованием."""
    data = request.get_json() or {}
    transcript_text = data.get('transcript_text', '').strip()
    analysis_text = data.get('analysis_text', '').strip()
    prompt_text = data.get('prompt_text', '').strip()
    
    if not transcript_text or not analysis_text or not prompt_text:
        return jsonify({'success': False, 'message': 'Не переданы данные транскрипта, анализа или промпта.'}), 400
    
    analysis_prompt = f"""Ты — эксперт по анализу клиентских звонков. Перед тобой транскрипт звонка и результат его классификации.

ТРАНСКРИПТ ЗВОНКА:
{transcript_text}

ПРОМПТ ЯКОРЯ:
{prompt_text}

РЕЗУЛЬТАТ КЛАССИФИКАЦИИ:
{analysis_text}

Твоя задача: написать КРАТКОЕ обоснование (2-4 предложения), ПОЧЕМУ звонок получил такую оценку.

ФОРМАТ ОТВЕТА (строго соблюдай):
✓ [Статус: Целевой/Нецелевой]
✓ Объяснение: [1-2 ключевых факта из разговора, которые определили статус]

ВАЖНО: 
- Ответ должен быть КРАТКИМ (максимум 3-4 строки)
- Без подробного пересказа всего разговора
- Только главные факты, определившие статус"""

    try:
        runtime_cfg = build_user_runtime_config()
        thebai_api_key = (runtime_cfg.get('api_keys') or {}).get('thebai_api_key', '')
        thebai_url = (runtime_cfg.get('api_keys') or {}).get('thebai_url', 'https://api.deepseek.com/v1/chat/completions')
        
        if not thebai_api_key:
            return jsonify({'success': False, 'message': 'API ключ для анализа не настроен.'}), 400
        
        # Отправляем запрос к TheB.ai / DeepSeek
        payload = {
            "model": (runtime_cfg.get('api_keys') or {}).get('thebai_model', 'deepseek-reasoner'),
            "messages": [
                {"role": "system", "content": "Ты эксперт по анализу клиентских звонков. Давай КРАТКИЕ и четкие ответы."},
                {"role": "user", "content": analysis_prompt}
            ],
            "temperature": 0.2,  # Снижаем для более предсказуемого короткого ответа
            "max_tokens": 300,   # Ограничиваем длину ответа для краткости
        }
        
        response = requests.post(
            thebai_url,
            headers={
                'Authorization': f'Bearer {thebai_api_key}',
                'Content-Type': 'application/json',
            },
            json=payload,
            timeout=120
        )
        
        if response.status_code != 200:
            app.logger.error(f"Ошибка API при анализе якоря: {response.status_code} - {response.text}")
            return jsonify({'success': False, 'message': f'Ошибка API: {response.status_code}'}), 502
        
        result = response.json()
        explanation = (result.get('choices', [{}])[0]
                          .get('message', {})
                          .get('content', '')).strip()
        
        return jsonify({'success': True, 'explanation': explanation})
    except Exception as exc:
        app.logger.error("Сбой при анализе якоря: %s", exc, exc_info=True)
        return jsonify({'success': False, 'message': str(exc)}), 500

@app.route('/api/vocabulary')
@login_required
def api_vocabulary():
    """API для получения словаря и пресетов по отраслевому профилю."""
    vocab_data = get_user_vocabulary_data()
    config_data = get_user_config_data()
    profile = config_data.get('business_profile', 'autoservice')
    presets = VOCAB_PRESETS.get(profile, VOCAB_PRESETS.get('universal', {}))
    vocab_data['business_profile'] = profile
    vocab_data['vocab_presets'] = presets
    return jsonify(vocab_data)

@app.route('/api/vocabulary/save', methods=['POST'])
@login_required
def api_vocabulary_save():
    """API для сохранения словаря."""
    try:
        data = request.get_json() or {}
        save_user_vocabulary_data(data)
        
        # Синхронизируем vocabulary.enabled с transcription.use_additional_vocab
        if 'enabled' in data:
            vocab_enabled = bool(data.get('enabled', True))
            # Получаем текущие настройки транскрипции
            config_data = get_user_config_data()
            transcription_cfg = config_data.get('transcription') or {}
            transcription_cfg['use_additional_vocab'] = vocab_enabled
            config_data['transcription'] = transcription_cfg
            save_user_config_data(config_data)
            # Обновляем runtime конфигурацию и применяем к config модулю
            runtime_cfg = build_user_runtime_config()
            # Применяем конфигурацию к call_analyzer.config
            try:
                from call_analyzer import config as legacy_config
                if hasattr(legacy_config, '_apply_profile_dict'):
                    legacy_config._apply_profile_dict(runtime_cfg.get('config_data', {}))
            except ImportError:
                pass
        
        # Синхронизируем словарь из БД в файл
        try:
            write_vocabulary_file(data)
        except Exception as file_error:
            app.logger.error("Не удалось сохранить vocabulary файл: %s", file_error)
        append_user_log('Словарь сохранен', module='vocabulary')
        return jsonify({'success': True, 'message': 'Словарь сохранен'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/reports')
@login_required
def reports_page():
    """Страница отчетов"""
    return render_template('reports.html', active_page='reports')

# Глобальная переменная для отслеживания статуса генерации отчетов
report_generation_status = {}

# Глобальная переменная для отслеживания прогресса генерации отчетов
report_generation_progress = {}

@app.route('/api/reports/generate', methods=['POST'])
@login_required
def api_reports_generate():
    """API для генерации отчетов"""
    try:
        data = request.get_json()
        report_type = data.get('type', 'week_full')
        # Параметры интервала из UI
        start_date_str = data.get('start_date')
        end_date_str = data.get('end_date')
        
        # Логируем полученные даты для отладки
        app.logger.info(f"Генерация отчета {report_type}: start_date={start_date_str}, end_date={end_date_str}")
        
        # Проверяем доступность модулей отчетов
        try:
            if report_type == 'week_full':
                from reports.week_full import run_week_full
            elif report_type == 'rr_3':
                from reports.rr_3 import run_rr_3
            elif report_type == 'rr_bad':
                from reports.rr_bad import run_rr_bad
            else:
                return jsonify({'success': False, 'message': f'Неизвестный тип отчета: {report_type}'})
        except ImportError as e:
            return jsonify({'success': False, 'message': f'Модуль отчета не найден: {str(e)}'})
        
        # Инициализируем статус и прогресс генерации
        report_generation_status[report_type] = {
            'status': 'running',
            'started_at': datetime.now().isoformat(),
            'message': 'Генерация отчета запущена',
            'progress': 0,
            'current_step': 'Инициализация...'
        }
        
        report_generation_progress[report_type] = {
            'progress': 0,
            'current_step': 'Инициализация...',
            'total_steps': 5,
            'completed_steps': 0
        }
        
        runtime_cfg = build_user_runtime_config()

        # Запускаем генерацию отчета в отдельном потоке
        def generate_report():
            try:
                # Обновляем прогресс
                report_generation_progress[report_type]['current_step'] = 'Загрузка модуля отчета...'
                report_generation_progress[report_type]['progress'] = 10
                report_generation_status[report_type]['progress'] = 10
                report_generation_status[report_type]['current_step'] = 'Загрузка модуля отчета...'
                
                if report_type == 'week_full':
                    report_generation_progress[report_type]['current_step'] = 'Анализ данных за неделю...'
                    report_generation_progress[report_type]['progress'] = 30
                    report_generation_status[report_type]['progress'] = 30
                    report_generation_status[report_type]['current_step'] = 'Анализ данных за неделю...'
                    # Подготовим даты
                    start_dt = None
                    end_dt = None
                    if start_date_str:
                        try:
                            start_dt = datetime.strptime(start_date_str, '%Y-%m-%d')
                        except ValueError as e:
                            app.logger.error(f"Ошибка парсинга start_date '{start_date_str}': {e}")
                    if end_date_str:
                        try:
                            end_dt = datetime.strptime(end_date_str, '%Y-%m-%d')
                        except ValueError as e:
                            app.logger.error(f"Ошибка парсинга end_date '{end_date_str}': {e}")
                    
                    app.logger.info(f"Запуск week_full с датами: start_dt={start_dt}, end_dt={end_dt}")
                    
                    # Получаем base_folder из runtime_cfg
                    base_folder = runtime_cfg['paths']['base_records_path']
                    app.logger.info(f"Используем base_folder: {base_folder}")
                    
                    # Запуск с интервалом
                    try:
                        with legacy_config_override(runtime_cfg):
                            result_path = run_week_full(start_date=start_dt, end_date=end_dt, base_folder=base_folder)
                    except TypeError:
                        with legacy_config_override(runtime_cfg):
                            result_path = run_week_full(base_folder=base_folder)
                    
                elif report_type == 'rr_3':
                    report_generation_progress[report_type]['current_step'] = 'Анализ станций Ретрак...'
                    report_generation_progress[report_type]['progress'] = 30
                    report_generation_status[report_type]['progress'] = 30
                    report_generation_status[report_type]['current_step'] = 'Анализ станций Ретрак...'
                    # Подготовим даты
                    date_from = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
                    date_to = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None
                    try:
                        with legacy_config_override(runtime_cfg):
                            run_rr_3(date_from=date_from, date_to=date_to)
                    except TypeError:
                        with legacy_config_override(runtime_cfg):
                            run_rr_3()
                    
                elif report_type == 'rr_bad':
                    report_generation_progress[report_type]['current_step'] = 'Анализ плохих звонков...'
                    report_generation_progress[report_type]['progress'] = 30
                    report_generation_status[report_type]['progress'] = 30
                    report_generation_status[report_type]['current_step'] = 'Анализ плохих звонков...'
                    # Подготовим даты
                    date_from = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
                    date_to = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None
                    try:
                        with legacy_config_override(runtime_cfg):
                            run_rr_bad(date_from=date_from, date_to=date_to)
                    except TypeError:
                        with legacy_config_override(runtime_cfg):
                            run_rr_bad()
                
                # Обновляем прогресс на завершение
                report_generation_progress[report_type]['current_step'] = 'Создание Excel файла...'
                report_generation_progress[report_type]['progress'] = 80
                report_generation_status[report_type]['progress'] = 80
                report_generation_status[report_type]['current_step'] = 'Создание Excel файла...'
                
                # Проверяем результат для week_full
                if report_type == 'week_full':
                    if not result_path or not os.path.exists(result_path):
                        report_generation_status[report_type] = {
                            'status': 'error',
                            'started_at': report_generation_status[report_type]['started_at'],
                            'completed_at': datetime.now().isoformat(),
                            'message': 'Не удалось создать отчет week_full. Проверьте наличие аудио и прав записи.',
                            'progress': 0,
                            'current_step': 'Ошибка'
                        }
                        report_generation_progress[report_type] = {
                            'progress': 0,
                            'current_step': 'Ошибка',
                            'total_steps': 5,
                            'completed_steps': 0
                        }
                        return
                
                # Финальное обновление
                report_generation_progress[report_type]['current_step'] = 'Отправка в Telegram...'
                report_generation_progress[report_type]['progress'] = 95
                report_generation_status[report_type]['progress'] = 95
                report_generation_status[report_type]['current_step'] = 'Отправка в Telegram...'
                
                time.sleep(1)
                
                # Обновляем статус на успешное завершение
                report_generation_status[report_type] = {
                    'status': 'completed',
                    'started_at': report_generation_status[report_type]['started_at'],
                    'completed_at': datetime.now().isoformat(),
                    'message': f'Отчет {report_type} успешно сгенерирован',
                    'progress': 100,
                    'current_step': 'Завершено'
                }
                
                report_generation_progress[report_type] = {
                    'progress': 100,
                    'current_step': 'Завершено',
                    'total_steps': 5,
                    'completed_steps': 5
                }
                
                logging.info(f"Отчет {report_type} успешно сгенерирован")
                
            except Exception as e:
                # Обновляем статус на ошибку
                report_generation_status[report_type] = {
                    'status': 'error',
                    'started_at': report_generation_status[report_type]['started_at'],
                    'completed_at': datetime.now().isoformat(),
                    'message': f'Ошибка генерации отчета {report_type}: {str(e)}',
                    'progress': 0,
                    'current_step': 'Ошибка'
                }
                
                report_generation_progress[report_type] = {
                    'progress': 0,
                    'current_step': 'Ошибка',
                    'total_steps': 5,
                    'completed_steps': 0
                }
                
                logging.error(f"Ошибка генерации отчета {report_type}: {e}")
        
        thread = threading.Thread(target=generate_report)
        thread.daemon = True
        thread.start()
        
        return jsonify({'success': True, 'message': f'Генерация отчета {report_type} запущена'})
        
    except Exception as e:
        logging.error(f"Ошибка запуска генерации отчета: {e}")
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/reports/status')
@login_required
def api_reports_status():
    """API для получения статуса генерации отчетов"""
    return jsonify(report_generation_status)

@app.route('/api/reports/progress')
@login_required
def api_reports_progress():
    """API для получения прогресса генерации отчетов"""
    return jsonify(report_generation_progress)

# API для управления расписаниями отчетов
def _schedule_to_dict(s: ReportSchedule):
    """Преобразовать модель расписания в словарь для JSON"""
    return {
        'id': s.id,
        'report_type': s.report_type,
        'schedule_type': s.schedule_type,
        'enabled': s.enabled,
        'daily_time': s.daily_time,
        'interval_value': s.interval_value,
        'interval_unit': s.interval_unit,
        'weekly_day': s.weekly_day,
        'weekly_time': s.weekly_time,
        'cron_expression': s.cron_expression,
        'period_type': s.period_type or 'last_week',
        'period_n_days': s.period_n_days,
        'last_run_at': s.last_run_at.isoformat() if s.last_run_at else None,
        'next_run_at': s.next_run_at.isoformat() if s.next_run_at else None,
    }

@app.route('/api/reports/schedules', methods=['GET'])
@login_required
def api_get_report_schedules():
    """Получить все расписания пользователя"""
    try:
        schedules = ReportSchedule.query.filter_by(user_id=current_user.id).all()
        return jsonify({'success': True, 'items': [_schedule_to_dict(s) for s in schedules]})
    except Exception as e:
        app.logger.error(f"Ошибка получения расписаний: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/reports/schedules/debug', methods=['GET'])
@login_required
def api_debug_report_schedules():
    """Диагностика: показать задачи в планировщике и данные в БД"""
    try:
        from web_interface.scheduler_service import scheduler
        
        # Данные из БД
        db_schedules = ReportSchedule.query.filter_by(user_id=current_user.id).all()
        db_info = []
        for s in db_schedules:
            db_info.append({
                'report_type': s.report_type,
                'schedule_type': s.schedule_type,
                'enabled': s.enabled,
                'daily_time': s.daily_time,
                'weekly_day': s.weekly_day,
                'weekly_time': s.weekly_time,
                'interval_value': s.interval_value,
                'interval_unit': s.interval_unit,
                'cron_expression': s.cron_expression,
                'next_run_at': s.next_run_at.isoformat() if s.next_run_at else None,
            })
        
        # Задачи в планировщике
        scheduler_jobs = []
        if scheduler:
            for job in scheduler.get_jobs():
                job_info = {
                    'id': job.id,
                    'name': job.name,
                    'next_run_time': job.next_run_time.isoformat() if job.next_run_time else None,
                    'trigger': str(job.trigger),
                }
                scheduler_jobs.append(job_info)
        
        return jsonify({
            'success': True,
            'db_schedules': db_info,
            'scheduler_jobs': scheduler_jobs,
            'scheduler_running': scheduler is not None and scheduler.running if scheduler else False,
        })
    except Exception as e:
        app.logger.error(f"Ошибка диагностики расписаний: {e}", exc_info=True)
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/reports/schedules', methods=['POST'])
@login_required
def api_save_report_schedule():
    """Создать или обновить расписание"""
    try:
        data = request.get_json() or {}
        report_type = data.get('report_type')
        
        if report_type not in ('week_full', 'rr_3', 'rr_bad'):
            return jsonify({'success': False, 'message': 'Неверный тип отчета'}), 400

        schedule = ReportSchedule.query.filter_by(
            user_id=current_user.id, 
            report_type=report_type
        ).first()
        
        if not schedule:
            schedule = ReportSchedule(user_id=current_user.id, report_type=report_type)

        new_schedule_type = data.get('schedule_type') or 'daily'
        schedule.schedule_type = new_schedule_type
        schedule.enabled = bool(data.get('enabled', True))
        
        # Очищаем поля других типов расписания, чтобы избежать конфликтов
        # и устанавливаем только поля текущего типа
        if new_schedule_type == 'daily':
            schedule.daily_time = data.get('daily_time') or '12:00'
            schedule.interval_value = None
            schedule.interval_unit = None
            schedule.weekly_day = None
            schedule.weekly_time = None
            schedule.cron_expression = None
        elif new_schedule_type == 'interval':
            schedule.daily_time = None
            schedule.interval_value = data.get('interval_value') or 1
            schedule.interval_unit = data.get('interval_unit') or 'days'
            schedule.weekly_day = None
            schedule.weekly_time = None
            schedule.cron_expression = None
        elif new_schedule_type == 'weekly':
            schedule.daily_time = None
            schedule.interval_value = None
            schedule.interval_unit = None
            schedule.weekly_day = data.get('weekly_day') if data.get('weekly_day') is not None else 0
            schedule.weekly_time = data.get('weekly_time') or '08:00'
            schedule.cron_expression = None
        elif new_schedule_type == 'custom':
            schedule.daily_time = None
            schedule.interval_value = None
            schedule.interval_unit = None
            schedule.weekly_day = None
            schedule.weekly_time = None
            schedule.cron_expression = data.get('cron_expression')
        else:
            # Fallback на daily
            schedule.daily_time = data.get('daily_time') or '12:00'
            schedule.interval_value = None
            schedule.interval_unit = None
            schedule.weekly_day = None
            schedule.weekly_time = None
            schedule.cron_expression = None
        
        app.logger.info(f"Сохранение расписания: type={new_schedule_type}, weekly_day={schedule.weekly_day}, weekly_time={schedule.weekly_time}, daily_time={schedule.daily_time}")
        
        # Период генерации
        schedule.period_type = data.get('period_type') or 'last_week'
        period_n_days = data.get('period_n_days')
        if period_n_days:
            schedule.period_n_days = int(period_n_days)
        else:
            schedule.period_n_days = None

        db.session.add(schedule)
        db.session.commit()

        # Обновляем задачу в планировщике
        from web_interface.scheduler_service import refresh_schedule
        refresh_schedule(schedule.id, app)

        app.logger.info(f"Расписание сохранено: user={current_user.id}, report={report_type}, type={schedule.schedule_type}")
        return jsonify({'success': True, 'item': _schedule_to_dict(schedule)})
        
    except Exception as e:
        app.logger.error(f"Ошибка сохранения расписания: {e}", exc_info=True)
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/reports/schedules/<int:schedule_id>', methods=['DELETE'])
@login_required
def api_delete_report_schedule(schedule_id):
    """Удалить расписание"""
    try:
        schedule = ReportSchedule.query.filter_by(
            id=schedule_id, 
            user_id=current_user.id
        ).first()
        
        if not schedule:
            return jsonify({'success': False, 'message': 'Расписание не найдено'}), 404

        # Удаляем задачу из планировщика
        from web_interface.scheduler_service import scheduler
        if scheduler:
            job_id = f"user_{schedule.user_id}_{schedule.report_type}"
            try:
                scheduler.remove_job(job_id)
            except Exception:
                pass

        db.session.delete(schedule)
        db.session.commit()

        app.logger.info(f"Расписание удалено: id={schedule_id}, user={current_user.id}")
        return jsonify({'success': True})
        
    except Exception as e:
        app.logger.error(f"Ошибка удаления расписания: {e}", exc_info=True)
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/reports/schedules/<int:schedule_id>/toggle', methods=['POST'])
@login_required
def api_toggle_report_schedule(schedule_id):
    """Включить/выключить расписание"""
    try:
        schedule = ReportSchedule.query.filter_by(
            id=schedule_id, 
            user_id=current_user.id
        ).first()
        
        if not schedule:
            return jsonify({'success': False, 'message': 'Расписание не найдено'}), 404

        schedule.enabled = not schedule.enabled
        db.session.commit()

        # Обновляем задачу в планировщике
        from web_interface.scheduler_service import refresh_schedule
        refresh_schedule(schedule.id, app)

        app.logger.info(f"Расписание {'включено' if schedule.enabled else 'выключено'}: id={schedule_id}, user={current_user.id}")
        return jsonify({'success': True, 'item': _schedule_to_dict(schedule)})
        
    except Exception as e:
        app.logger.error(f"Ошибка переключения расписания: {e}", exc_info=True)
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/check-excel')
@login_required
def api_check_excel():
    """API для проверки содержимого Excel файла отчета"""
    try:
        import openpyxl
        from pathlib import Path
        
        # Ищем последний созданный отчет
        runtime_cfg = build_user_runtime_config()
        base_path_str = runtime_cfg['paths']['base_records_path']
        if not base_path_str:
            return jsonify({'error': 'Не задан путь к каталогу со звонками в настройках профиля'})
        base_path = Path(str(base_path_str))
        excel_files = list(base_path.rglob('Отчет_по_скрипту_*.xlsx'))
        
        if not excel_files:
            return jsonify({'error': 'Файлы отчетов не найдены'})
        
        # Берем самый новый файл
        latest_file = max(excel_files, key=lambda x: x.stat().st_mtime)
        
        wb = openpyxl.load_workbook(latest_file)
        
        result = {
            'file_path': str(latest_file),
            'sheets': wb.sheetnames,
            'sheet_count': len(wb.sheetnames),
            'sheet_details': {}
        }
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            result['sheet_details'][sheet_name] = {
                'rows': sheet.max_row,
                'columns': sheet.max_column,
                'headers': []
            }
            
            # Получаем заголовки первой строки
            if sheet.max_row > 0:
                for col in range(1, min(sheet.max_column + 1, 6)):  # первые 5 столбцов
                    cell_value = sheet.cell(row=1, column=col).value
                    result['sheet_details'][sheet_name]['headers'].append(str(cell_value) if cell_value else '')
        
        return jsonify(result)
        
    except Exception as e:
        return jsonify({'error': f'Ошибка при проверке Excel файла: {str(e)}'})

@app.route('/test-reports')
@login_required
def test_reports_page():
    """Страница тестирования отчетов"""
    return '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Тест отчетов - Call Analyzer</title>
        <meta charset="utf-8">
        <style>
            body { font-family: Arial, sans-serif; margin: 40px; }
            .header { background: #2c3e50; color: white; padding: 20px; border-radius: 5px; }
            .content { margin: 20px 0; }
            .test-section { background: #ecf0f1; padding: 15px; margin: 10px 0; border-radius: 5px; }
            .test-button { background: #3498db; color: white; padding: 10px 20px; border: none; border-radius: 5px; cursor: pointer; margin: 5px; }
            .test-button:hover { background: #2980b9; }
            .result { margin: 10px 0; padding: 10px; border-radius: 5px; }
            .success { background: #d5f4e6; border: 1px solid #27ae60; }
            .error { background: #fadbd8; border: 1px solid #e74c3c; }
            .info { background: #d6eaf8; border: 1px solid #3498db; }
            
            .progress-bar {
                width: 100%;
                height: 20px;
                background-color: #ecf0f1;
                border-radius: 10px;
                overflow: hidden;
                margin: 10px 0;
            }
            
            .progress-fill {
                height: 100%;
                background: linear-gradient(90deg, #3498db, #2ecc71);
                width: 0%;
                transition: width 0.3s ease;
                border-radius: 10px;
            }
            
            .progress-text {
                text-align: center;
                font-weight: bold;
                color: #2c3e50;
                margin: 5px 0;
            }
        </style>
    </head>
    <body>
        <div class="header">
            <h1>🧪 Тест отчетов Call Analyzer</h1>
            <p>Проверка функционала генерации отчетов</p>
        </div>
        
        <div class="content">
            <div class="test-section">
                <h2>📊 Тест системы отчетов</h2>
                <button class="test-button" onclick="testReportsSystem()">Проверить систему отчетов</button>
                <div id="reports-test-result"></div>
            </div>
            
            <div class="test-section">
                <h2>🔑 Тест API ключей</h2>
                <button class="test-button" onclick="testApiKeys()">Проверить API ключи</button>
                <div id="api-test-result"></div>
            </div>
            
            <div class="test-section">
                <h2>📈 Тест генерации отчетов</h2>
                <button class="test-button" onclick="testReportGeneration()">Запустить тестовый отчет</button>
                <div id="generation-test-result"></div>
            </div>
            
            <div class="test-section">
                <h2>📋 Статус генерации</h2>
                <button class="test-button" onclick="checkGenerationStatus()">Проверить статус</button>
                <div id="status-result"></div>
            </div>
            
            <div class="test-section">
                <h2>📊 Прогресс генерации</h2>
                <button class="test-button" onclick="checkGenerationProgress()">Проверить прогресс</button>
                <div id="progress-result"></div>
                <div id="progress-bar-container" style="display: none;">
                    <div class="progress-bar">
                        <div class="progress-fill" id="progress-fill"></div>
                    </div>
                    <div class="progress-text" id="progress-text"></div>
                </div>
            </div>
        </div>
        
        <script>
            function showResult(elementId, message, type) {
                const element = document.getElementById(elementId);
                element.innerHTML = `<div class="result ${type}">${message}</div>`;
            }
            
            async function testReportsSystem() {
                showResult('reports-test-result', '⏳ Проверка системы отчетов...', 'info');
                
                try {
                    const response = await fetch('/api/reports/test');
                    const data = await response.json();
                    
                    if (data.dependencies && data.dependencies.all_installed) {
                        showResult('reports-test-result', 
                            '✅ Система отчетов работает корректно<br>' +
                            '📦 Все зависимости установлены<br>' +
                            '📁 Тестовый файл: ' + (data.test_result.test_file || 'N/A'), 
                            'success');
                    } else {
                        showResult('reports-test-result', 
                            '❌ Проблемы с системой отчетов<br>' +
                            '📦 Недостающие зависимости: ' + (data.dependencies.missing.join(', ') || 'N/A'), 
                            'error');
                    }
                } catch (error) {
                    showResult('reports-test-result', '❌ Ошибка: ' + error.message, 'error');
                }
            }
            
            async function testApiKeys() {
                showResult('api-test-result', '⏳ Проверка API ключей...', 'info');
                
                try {
                    const response = await fetch('/api/config/test');
                    const data = await response.json();
                    
                    let result = '🔑 Результаты тестирования API:<br>';
                    
                    if (data.test_results) {
                        for (const [api, result_data] of Object.entries(data.test_results)) {
                            const status = result_data.status === 'success' ? '✅' : '❌';
                            result += `${status} ${api}: ${result_data.message}<br>`;
                        }
                    }
                    
                    const type = data.summary && data.summary.working_apis > 0 ? 'success' : 'error';
                    showResult('api-test-result', result, type);
                } catch (error) {
                    showResult('api-test-result', '❌ Ошибка: ' + error.message, 'error');
                }
            }
            
            async function testReportGeneration() {
                showResult('generation-test-result', '⏳ Запуск тестового отчета...', 'info');
                
                try {
                    const response = await fetch('/api/reports/generate', {
                        method: 'POST',
                        headers: {'Content-Type': 'application/json'},
                        body: JSON.stringify({type: 'week_full'})
                    });
                    
                    const data = await response.json();
                    
                    if (data.success) {
                        showResult('generation-test-result', 
                            '✅ Генерация отчета запущена<br>' +
                            '📝 ' + data.message, 
                            'success');
                    } else {
                        showResult('generation-test-result', 
                            '❌ Ошибка запуска генерации<br>' +
                            '📝 ' + data.message, 
                            'error');
                    }
                } catch (error) {
                    showResult('generation-test-result', '❌ Ошибка: ' + error.message, 'error');
                }
            }
            
            async function checkGenerationStatus() {
                showResult('status-result', '⏳ Проверка статуса генерации...', 'info');
                
                try {
                    const response = await fetch('/api/reports/status');
                    const data = await response.json();
                    
                    if (Object.keys(data).length === 0) {
                        showResult('status-result', '📊 Нет активных генераций отчетов', 'info');
                    } else {
                        let result = '📊 Статус генерации отчетов:<br>';
                        for (const [reportType, status] of Object.entries(data)) {
                            result += `📋 ${reportType}: ${status.status}<br>`;
                            result += `📅 Начато: ${status.started_at}<br>`;
                            if (status.completed_at) {
                                result += `✅ Завершено: ${status.completed_at}<br>`;
                            }
                            if (status.progress !== undefined) {
                                result += `📊 Прогресс: ${status.progress}%<br>`;
                            }
                            if (status.current_step) {
                                result += `🔄 Текущий шаг: ${status.current_step}<br>`;
                            }
                            result += `📝 ${status.message}<br><br>`;
                        }
                        showResult('status-result', result, 'success');
                    }
                } catch (error) {
                    showResult('status-result', '❌ Ошибка: ' + error.message, 'error');
                }
            }
            
            async function checkGenerationProgress() {
                showResult('progress-result', '⏳ Проверка прогресса генерации...', 'info');
                
                try {
                    const response = await fetch('/api/reports/progress');
                    const data = await response.json();
                    
                    if (Object.keys(data).length === 0) {
                        showResult('progress-result', '📊 Нет активных генераций отчетов', 'info');
                        document.getElementById('progress-bar-container').style.display = 'none';
                    } else {
                        let result = '📊 Прогресс генерации отчетов:<br>';
                        let hasActiveGeneration = false;
                        
                        for (const [reportType, progress] of Object.entries(data)) {
                            if (progress.progress > 0 && progress.progress < 100) {
                                hasActiveGeneration = true;
                                result += `📋 ${reportType}: ${progress.progress}%<br>`;
                                result += `🔄 ${progress.current_step}<br>`;
                                result += `📊 Шагов: ${progress.completed_steps}/${progress.total_steps}<br><br>`;
                                
                                // Обновляем прогресс-бар
                                updateProgressBar(progress.progress, progress.current_step);
                            } else if (progress.progress === 100) {
                                result += `✅ ${reportType}: Завершено<br>`;
                            }
                        }
                        
                        if (hasActiveGeneration) {
                            document.getElementById('progress-bar-container').style.display = 'block';
                            // Автоматически обновляем прогресс каждые 2 секунды
                            setTimeout(checkGenerationProgress, 2000);
                        } else {
                            document.getElementById('progress-bar-container').style.display = 'none';
                        }
                        
                        showResult('progress-result', result, 'success');
                    }
                } catch (error) {
                    showResult('progress-result', '❌ Ошибка: ' + error.message, 'error');
                }
            }
            
            function updateProgressBar(progress, currentStep) {
                const progressFill = document.getElementById('progress-fill');
                const progressText = document.getElementById('progress-text');
                
                if (progressFill && progressText) {
                    progressFill.style.width = progress + '%';
                    progressText.textContent = `${progress}% - ${currentStep}`;
                }
            }
        </script>
    </body>
    </html>
    '''

@app.route('/api/reports/test')
@login_required
def api_reports_test():
    """API для тестирования системы отчетов"""
    try:
        # Проверяем зависимости
        dependencies = ['pandas', 'openpyxl', 'requests', 'yaml', 'watchdog', 'APScheduler']
        missing_deps = []
        
        for dep in dependencies:
            try:
                __import__(dep)
            except ImportError:
                missing_deps.append(dep)
        
        # Тестируем создание простого отчета
        test_result = {'success': False, 'message': ''}
        
        if not missing_deps:
            try:
                import pandas as pd
                
                # Создаем тестовые данные
                data = {
                    'Дата': ['2024-01-01', '2024-01-02', '2024-01-03'],
                    'Станция': ['NN01', 'NN02', 'NN01'],
                    'Звонков': [10, 15, 12],
                    'Качество': [8.5, 7.8, 9.2]
                }
                
                df = pd.DataFrame(data)
                
                # Сохраняем тестовый файл
                test_file = get_project_root() / 'test_report.xlsx'
                df.to_excel(test_file, index=False)
                
                test_result = {
                    'success': True,
                    'message': 'Система отчетов работает корректно',
                    'test_file': str(test_file)
                }
                
            except Exception as e:
                test_result = {
                    'success': False,
                    'message': f'Ошибка создания тестового отчета: {str(e)}'
                }
        else:
            test_result = {
                'success': False,
                'message': f'Недостающие зависимости: {", ".join(missing_deps)}'
            }
        
        return jsonify({
            'dependencies': {
                'missing': missing_deps,
                'all_installed': len(missing_deps) == 0
            },
            'test_result': test_result
        })
        
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/transfers')
@login_required
def transfers_page():
    """Страница управления переводами"""
    return render_template('transfers.html', active_page='transfers')

@app.route('/api/transfers')
@login_required
def api_transfers():
    """API для получения списка переводов"""
    try:
        _, runtime_dir = get_runtime_context()
        transfers_file = runtime_dir / 'transfer_cases.json'
        if transfers_file.exists():
            with open(transfers_file, 'r', encoding='utf-8') as f:
                transfers = json.load(f)
        else:
            transfers = []
        
        return jsonify(transfers)
        
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/recalls')
@login_required
def recalls_page():
    """Страница управления перезвонами"""
    return render_template('recalls.html', active_page='recalls')

@app.route('/api/recalls')
@login_required
def api_recalls():
    """API для получения списка перезвонов"""
    try:
        _, runtime_dir = get_runtime_context()
        recalls_file = runtime_dir / 'recall_cases.json'
        if recalls_file.exists():
            with open(recalls_file, 'r', encoding='utf-8') as f:
                recalls = json.load(f)
        else:
            recalls = []
        
        return jsonify(recalls)
        
    except Exception as e:
        return jsonify({'error': str(e)})

def _parse_datetime_field(value):
    """Преобразует ISO-строку в datetime или возвращает None."""
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    try:
        return datetime.fromisoformat(value)
    except Exception:
        return None


@app.route('/ats')
@login_required
def ats_page():
    """Страница интеграций с АТС"""
    return render_template('ats.html', active_page='ats')


def _process_rostelecom_recording(conn_id, user_id, session_id, from_number, request_number, request_pin, call_type, timestamp_str):
    """Фоновая обработка записи из Ростелеком: get_record, скачивание, сохранение, запуск анализа."""
    from pathlib import Path
    def _run():
        try:
            from call_analyzer.rostelecom_connector import get_record, download_recording, make_rostelecom_filename
            from config.settings import get_config
            from sqlalchemy import create_engine, text
            cfg = get_config()
            engine = create_engine(cfg.SQLALCHEMY_DATABASE_URI)
            with engine.connect() as c:
                conn_row = c.execute(text("""
                    SELECT api_url, client_id, sign_key FROM rostelecom_ats_connections WHERE id = :cid
                """), {'cid': conn_id}).fetchone()
                r = c.execute(text("""
                    SELECT uc.base_records_path FROM user_config uc WHERE uc.user_id = :uid
                """), {'uid': user_id}).fetchone()
                base_path_str = (r.base_records_path or '').strip() if r else ''
            engine.dispose()
            if not conn_row:
                app.logger.error("Rostelecom conn not found")
                return
            if not base_path_str:
                base_path_str = str(Path(getattr(get_config(), 'BASE_RECORDS_PATH', '/var/calls')) / 'users' / str(user_id))
            base_path = Path(base_path_str)
            ts = timestamp_str or datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S.%f')
            ts_clean = ts.replace('-', '').replace(' ', '-').replace(':', '')[:15]
            filename = make_rostelecom_filename(session_id, from_number or '', request_number or '', request_pin, call_type or 'incoming', ts_clean)
            try:
                dt = datetime.strptime(ts[:19].replace('T', ' '), '%Y-%m-%d %H:%M:%S')
            except Exception:
                dt = datetime.utcnow()
            target_dir = base_path / str(dt.year) / f"{dt.month:02d}" / f"{dt.day:02d}"
            target_dir.mkdir(parents=True, exist_ok=True)
            save_path = target_dir / filename
            url, err = get_record(conn_row.api_url, conn_row.client_id, conn_row.sign_key, session_id, timeout=60)
            if err or not url:
                app.logger.error(f"Rostelecom get_record failed: {err}")
                with app.app_context():
                    RostelecomAtsConnection.query.filter_by(id=conn_id).update({'last_error': err})
                    db.session.commit()
                return
            if not download_recording(url, save_path, timeout=120):
                app.logger.error("Rostelecom download failed")
                return
            app.logger.info(f"Rostelecom запись сохранена: {save_path}. Watchdog call_analyzer обработает файл.")
            with app.app_context():
                RostelecomAtsConnection.query.filter_by(id=conn_id).update({
                    'last_webhook_at': datetime.utcnow(),
                    'last_error': None
                })
                db.session.commit()
        except Exception as e:
            app.logger.error(f"Rostelecom _process_rostelecom_recording: {e}", exc_info=True)
            try:
                with app.app_context():
                    RostelecomAtsConnection.query.filter_by(id=conn_id).update({'last_error': str(e)})
                    db.session.commit()
            except Exception:
                pass
    threading.Thread(target=_run, daemon=True, name=f"Rostelecom-{session_id}").start()


@app.route('/api/ats/rostelecom/webhook', methods=['POST'])
def api_rostelecom_webhook():
    """
    Webhook для получения событий call_events от облачной АТС Ростелеком.
    Не требует авторизации — проверка по X-Client-ID и X-Client-Sign.
    """
    try:
        client_id = request.headers.get('X-Client-ID') or request.headers.get('x-client-id')
        client_sign = request.headers.get('X-Client-Sign') or request.headers.get('x-client-sign')
        body = request.get_data(as_text=True)
        if not client_id or not client_sign or not body:
            return jsonify({'error': 'Missing X-Client-ID or X-Client-Sign or body'}), 400
        conn = RostelecomAtsConnection.query.filter_by(client_id=client_id, is_active=True).first()
        if not conn:
            app.logger.warning(f"Rostelecom webhook: unknown client_id {client_id[:16]}...")
            return jsonify({'error': 'Unknown client'}), 404
        from call_analyzer.rostelecom_connector import verify_sign
        if not verify_sign(client_id, body, conn.sign_key, client_sign):
            app.logger.warning("Rostelecom webhook: invalid signature")
            return jsonify({'error': 'Invalid signature'}), 403
        data = request.get_json(force=True, silent=True) or {}
        state = data.get('state')
        session_id = data.get('session_id')
        call_type = data.get('type', 'incoming')
        timestamp = data.get('timestamp', '')
        from_num = data.get('from_number', '')
        req_num = data.get('request_number', '')
        req_pin = data.get('request_pin', '')
        is_record = (data.get('is_record') or '').lower() == 'true'
        allowed = conn.allowed_directions if conn.allowed_directions else None
        if allowed is not None and len(allowed) > 0 and call_type not in allowed:
            app.logger.info(f"Rostelecom webhook: пропуск звонка type={call_type} (разрешены: {allowed})")
            return jsonify({'result': '0', 'resultMessage': 'OK'}), 200
        if conn.start_from and timestamp:
            try:
                ts_clean = (timestamp or '')[:19].replace('T', ' ')
                call_dt = datetime.strptime(ts_clean, '%Y-%m-%d %H:%M:%S')
                start_dt = conn.start_from.replace(tzinfo=None) if conn.start_from.tzinfo else conn.start_from
                if call_dt < start_dt:
                    app.logger.info(f"Rostelecom webhook: пропуск звонка до start_from {conn.start_from}")
                    return jsonify({'result': '0', 'resultMessage': 'OK'}), 200
            except Exception:
                pass
        if state == 'end' and is_record and session_id:
            _process_rostelecom_recording(conn.id, conn.user_id, session_id, from_num, req_num, req_pin, call_type, timestamp)
        return jsonify({'result': '0', 'resultMessage': 'OK'}), 200
    except Exception as e:
        app.logger.error(f"Rostelecom webhook error: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/ats/rostelecom/connections', methods=['GET'])
@login_required
def api_rostelecom_connections():
    """Список подключений Ростелеком для текущего пользователя"""
    try:
        conns = RostelecomAtsConnection.query.filter_by(user_id=current_user.id).all()
        return jsonify([{
            'id': c.id,
            'name': c.name,
            'api_url': c.api_url,
            'client_id': c.client_id[:16] + '...' if len(c.client_id) > 16 else c.client_id,
            'is_active': c.is_active,
            'pin_to_station': c.pin_to_station or {},
            'allowed_directions': c.allowed_directions if c.allowed_directions else ['incoming', 'outbound', 'internal'],
            'start_from': c.start_from.isoformat() + 'Z' if c.start_from else None,
            'last_sync': c.last_sync.isoformat() + 'Z' if c.last_sync else None,
            'last_webhook_at': c.last_webhook_at.isoformat() + 'Z' if c.last_webhook_at else None,
            'last_error': c.last_error,
            'created_at': c.created_at.isoformat() + 'Z',
            'updated_at': c.updated_at.isoformat() + 'Z',
        } for c in conns])
    except Exception as e:
        app.logger.error(f"Rostelecom connections: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/ats/rostelecom/connections', methods=['POST'])
@login_required
def api_rostelecom_create():
    """Создание подключения Ростелеком"""
    try:
        data = request.get_json()
        for f in ['api_url', 'client_id', 'sign_key']:
            if not data.get(f):
                return jsonify({'success': False, 'message': f'Поле {f} обязательно'}), 400
        allowed = data.get('allowed_directions')
        if allowed is not None:
            allowed = list(allowed) if isinstance(allowed, list) and len(allowed) > 0 else None
        start_from = _parse_datetime_field(data.get('start_from'))
        conn = RostelecomAtsConnection(
            user_id=current_user.id,
            name=data.get('name', 'Ростелеком'),
            api_url=data['api_url'].rstrip('/'),
            client_id=data['client_id'].strip(),
            sign_key=data['sign_key'].strip(),
            is_active=data.get('is_active', True),
            pin_to_station=data.get('pin_to_station') or {},
            allowed_directions=allowed,
            start_from=start_from,
        )
        db.session.add(conn)
        db.session.commit()
        return jsonify({'success': True, 'id': conn.id, 'message': 'Подключение создано'})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Rostelecom create: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/ats/rostelecom/connections/<int:conn_id>', methods=['PUT'])
@login_required
def api_rostelecom_update(conn_id):
    """Обновление подключения Ростелеком"""
    conn = RostelecomAtsConnection.query.filter_by(id=conn_id, user_id=current_user.id).first()
    if not conn:
        return jsonify({'error': 'Not found'}), 404
    try:
        data = request.get_json()
        if 'name' in data:
            conn.name = data['name']
        if 'api_url' in data:
            conn.api_url = data['api_url'].rstrip('/')
        if 'client_id' in data and data['client_id']:
            conn.client_id = data['client_id'].strip()
        if 'sign_key' in data and data['sign_key']:
            conn.sign_key = data['sign_key'].strip()
        if 'is_active' in data:
            conn.is_active = bool(data['is_active'])
        if 'pin_to_station' in data:
            conn.pin_to_station = data['pin_to_station'] if isinstance(data['pin_to_station'], dict) else {}
        if 'allowed_directions' in data:
            val = data['allowed_directions']
            conn.allowed_directions = val if isinstance(val, list) else None
        if 'start_from' in data:
            conn.start_from = _parse_datetime_field(data.get('start_from'))
        db.session.commit()
        return jsonify({'success': True, 'message': 'Обновлено'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/ats/rostelecom/connections/<int:conn_id>/test', methods=['POST'])
@login_required
def api_rostelecom_test(conn_id):
    """Проверка подключения к API Ростелеком"""
    conn = RostelecomAtsConnection.query.filter_by(id=conn_id, user_id=current_user.id).first()
    if not conn:
        return jsonify({'success': False, 'message': 'Подключение не найдено'}), 404
    try:
        from call_analyzer.rostelecom_connector import test_connection
        success, message = test_connection(
            api_url=conn.api_url,
            client_id=conn.client_id,
            sign_key=conn.sign_key,
        )
        if success:
            return jsonify({'success': True, 'message': message})
        return jsonify({'success': False, 'message': message}), 400
    except Exception as e:
        app.logger.error(f"Ошибка теста Ростелеком {conn_id}: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


def _sync_rostelecom_connection(conn_id: int):
    """Фоновая синхронизация истории звонков Ростелеком."""
    def _run():
        try:
            conn = RostelecomAtsConnection.query.get(conn_id)
            if not conn:
                return
            from call_analyzer.rostelecom_connector import fetch_call_history
            date_to = datetime.utcnow()
            date_from = (conn.start_from or (date_to - timedelta(days=7)))
            if date_from.tzinfo:
                date_from = date_from.replace(tzinfo=None)
            success, message, calls = fetch_call_history(
                conn.api_url, conn.client_id, conn.sign_key,
                date_from, date_to, timeout=60
            )
            with app.app_context():
                RostelecomAtsConnection.query.filter_by(id=conn_id).update({
                    'last_sync': datetime.utcnow(),
                    'last_error': None if success else message
                })
                db.session.commit()
            if success and calls:
                for c in calls:
                    sid = c.get('session_id') or c.get('sessionId')
                    is_rec = (c.get('is_record') or c.get('isRecord') or '').lower() == 'true'
                    if sid and is_rec:
                        _process_rostelecom_recording(
                            conn_id, conn.user_id, sid,
                            c.get('from_number') or c.get('fromNumber', ''),
                            c.get('request_number') or c.get('requestNumber', ''),
                            c.get('request_pin') or c.get('requestPin', ''),
                            c.get('type') or c.get('call_type', 'incoming'),
                            c.get('timestamp') or c.get('start_time', '')
                        )
        except Exception as e:
            app.logger.error(f"Rostelecom sync {conn_id}: {e}", exc_info=True)
            try:
                with app.app_context():
                    RostelecomAtsConnection.query.filter_by(id=conn_id).update({
                        'last_sync': datetime.utcnow(),
                        'last_error': str(e)
                    })
                    db.session.commit()
            except Exception:
                pass
    threading.Thread(target=_run, daemon=True, name=f"Rostelecom-sync-{conn_id}").start()


@app.route('/api/ats/rostelecom/connections/<int:conn_id>/sync', methods=['POST'])
@login_required
def api_rostelecom_sync(conn_id):
    """Запуск синхронизации истории звонков Ростелеком"""
    conn = RostelecomAtsConnection.query.filter_by(id=conn_id, user_id=current_user.id).first()
    if not conn:
        return jsonify({'success': False, 'message': 'Подключение не найдено'}), 404
    _sync_rostelecom_connection(conn_id)
    return jsonify({'success': True, 'message': 'Синхронизация запущена'})


@app.route('/api/ats/rostelecom/connections/<int:conn_id>', methods=['DELETE'])
@login_required
def api_rostelecom_delete(conn_id):
    conn = RostelecomAtsConnection.query.filter_by(id=conn_id, user_id=current_user.id).first()
    if not conn:
        return jsonify({'error': 'Not found'}), 404
    UserConfig.query.filter_by(rostelecom_ats_connection_id=conn_id).update({'rostelecom_ats_connection_id': None, 'source_type': 'local'})
    db.session.delete(conn)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/ftp')
@login_required
def ftp_page():
    """Страница управления FTP подключениями"""
    return render_template('ftp.html', active_page='ftp')

@app.route('/api/ftp/connections', methods=['GET'])
@login_required
def api_ftp_connections():
    """API для получения списка FTP подключений"""
    try:
        connections = FtpConnection.query.filter_by(user_id=current_user.id).all()
        result = []
        for conn in connections:
            result.append({
                'id': conn.id,
                'name': conn.name,
                'host': conn.host,
                'port': conn.port,
                'username': conn.username,
                'remote_path': conn.remote_path,
                'protocol': conn.protocol,
                'is_active': conn.is_active,
                'sync_interval': conn.sync_interval,
                'start_from': conn.start_from.isoformat() + 'Z' if conn.start_from else None,
                'last_processed_mtime': conn.last_processed_mtime.isoformat() + 'Z' if conn.last_processed_mtime else None,
                'last_processed_filename': conn.last_processed_filename,
                'last_sync': conn.last_sync.isoformat() + 'Z' if conn.last_sync else None,
                'last_error': conn.last_error,
                'download_count': conn.download_count,
                'created_at': conn.created_at.isoformat() + 'Z',
                'updated_at': conn.updated_at.isoformat() + 'Z'
            })
        return jsonify(result)
    except Exception as e:
        app.logger.error(f"Ошибка получения FTP подключений: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/ftp/connections', methods=['POST'])
@login_required
def api_ftp_create():
    """API для создания нового FTP подключения"""
    try:
        data = request.get_json()
        
        # Валидация
        required_fields = ['name', 'host', 'username', 'password', 'remote_path', 'protocol']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'Поле {field} обязательно'}), 400
        
        start_from = _parse_datetime_field(data.get('start_from'))

        # Создаем подключение
        conn = FtpConnection(
            user_id=current_user.id,
            name=data['name'],
            host=data['host'],
            port=int(data.get('port', 21 if data['protocol'] == 'ftp' else 22)),
            username=data['username'],
            password=data['password'],
            remote_path=data['remote_path'],
            protocol=data['protocol'],
            is_active=data.get('is_active', True),
            sync_interval=int(data.get('sync_interval', 300)),
            start_from=start_from
        )
        
        db.session.add(conn)
        db.session.commit()
        
        # Запускаем синхронизацию только если это подключение выбрано в конфигурации пользователя
        # Проверяем настройки пользователя
        from database.models import UserSettings
        from common.user_settings import default_config_template
        
        user_settings = UserSettings.query.filter_by(user_id=current_user.id).first()
        if user_settings and user_settings.data:
            config_data = user_settings.data.get('config') or default_config_template()
            paths_cfg = config_data.get('paths') or {}
            if paths_cfg.get('source_type') == 'ftp' and paths_cfg.get('ftp_connection_id') == conn.id:
                if conn.is_active:
                    try:
                        from call_analyzer.ftp_sync_manager import start_ftp_sync
                        start_ftp_sync(conn.id)
                    except Exception as e:
                        app.logger.warning(f"Не удалось запустить синхронизацию для FTP {conn.id}: {e}")
        
        return jsonify({'success': True, 'id': conn.id, 'message': 'FTP подключение создано'})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Ошибка создания FTP подключения: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/ftp/connections/<int:conn_id>', methods=['PUT'])
@login_required
def api_ftp_update(conn_id):
    """API для обновления FTP подключения"""
    try:
        conn = FtpConnection.query.filter_by(id=conn_id, user_id=current_user.id).first()
        if not conn:
            return jsonify({'success': False, 'message': 'Подключение не найдено'}), 404
        
        data = request.get_json()
        was_active = conn.is_active
        
        # Обновляем поля
        if 'name' in data:
            conn.name = data['name']
        if 'host' in data:
            conn.host = data['host']
        if 'port' in data:
            conn.port = int(data['port'])
        if 'username' in data:
            conn.username = data['username']
        if 'password' in data and data['password']:  # Обновляем пароль только если он не пустой
            conn.password = data['password']
        if 'remote_path' in data:
            conn.remote_path = data['remote_path']
        if 'protocol' in data:
            conn.protocol = data['protocol']
        if 'is_active' in data:
            conn.is_active = data['is_active']
        if 'sync_interval' in data:
            conn.sync_interval = int(data['sync_interval'])
        if 'start_from' in data:
            conn.start_from = _parse_datetime_field(data.get('start_from'))
        
        db.session.commit()
        
        # Управляем синхронизацией
        # Всегда останавливаем синхронизацию перед возможным перезапуском для применения новых настроек
        try:
            from call_analyzer.ftp_sync_manager import start_ftp_sync, stop_ftp_sync
            
            stop_ftp_sync(conn.id)
            
            if conn.is_active:
                from database.models import UserSettings
                from common.user_settings import default_config_template
                
                user_settings = UserSettings.query.filter_by(user_id=current_user.id).first()
                is_selected = False
                paths_cfg = {}
                if user_settings and user_settings.data:
                    config_data = user_settings.data.get('config') or default_config_template()
                    paths_cfg = config_data.get('paths') or {}
                    is_selected = (paths_cfg.get('source_type') == 'ftp' and 
                                  paths_cfg.get('ftp_connection_id') == conn.id)
                
                if is_selected:
                    app.logger.info(f"FTP подключение {conn.id} ({conn.name}) активно и выбрано в конфигурации, запускаем синхронизацию")
                    start_ftp_sync(conn.id)
                else:
                    app.logger.info(f"FTP подключение {conn.id} ({conn.name}) активно, но не выбрано в конфигурации (source_type={paths_cfg.get('source_type')}, ftp_id={paths_cfg.get('ftp_connection_id')})")
        except Exception as e:
            app.logger.warning(f"Не удалось обновить синхронизацию для FTP {conn.id}: {e}", exc_info=True)
        
        return jsonify({'success': True, 'message': 'FTP подключение обновлено'})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Ошибка обновления FTP подключения: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/ftp/connections/<int:conn_id>', methods=['DELETE'])
@login_required
def api_ftp_delete(conn_id):
    """API для удаления FTP подключения"""
    try:
        conn = FtpConnection.query.filter_by(id=conn_id, user_id=current_user.id).first()
        if not conn:
            return jsonify({'success': False, 'message': 'Подключение не найдено'}), 404
        
        # Останавливаем синхронизацию
        try:
            from call_analyzer.ftp_sync_manager import stop_ftp_sync
            stop_ftp_sync(conn.id)
        except Exception as e:
            app.logger.warning(f"Не удалось остановить синхронизацию для FTP {conn.id}: {e}")
        
        db.session.delete(conn)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'FTP подключение удалено'})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Ошибка удаления FTP подключения: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/ftp/connections/<int:conn_id>/test', methods=['POST'])
@login_required
def api_ftp_test(conn_id):
    """API для тестирования FTP подключения"""
    try:
        conn = FtpConnection.query.filter_by(id=conn_id, user_id=current_user.id).first()
        if not conn:
            return jsonify({'success': False, 'message': 'Подключение не найдено'}), 404
        
        from call_analyzer.ftp_sync import FtpSync
        
        ftp = FtpSync(
            host=conn.host,
            port=conn.port,
            username=conn.username,
            password=conn.password,
            remote_path=conn.remote_path,
            protocol=conn.protocol
        )
        
        success, message = ftp.test_connection()
        
        if success:
            return jsonify({'success': True, 'message': message})
        else:
            return jsonify({'success': False, 'message': message}), 400
            
    except Exception as e:
        app.logger.error(f"Ошибка тестирования FTP подключения: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/ftp/connections/<int:conn_id>/sync', methods=['POST'])
@login_required
def api_ftp_sync(conn_id):
    """API для ручной синхронизации FTP подключения"""
    try:
        conn = FtpConnection.query.filter_by(id=conn_id, user_id=current_user.id).first()
        if not conn:
            return jsonify({'success': False, 'message': 'Подключение не найдено'}), 404
        
        # Запускаем синхронизацию в фоне
        from call_analyzer.ftp_sync_manager import sync_ftp_connection
        app.logger.info(f"Запуск ручной синхронизации для FTP подключения {conn.id} ({conn.name})")
        threading.Thread(target=sync_ftp_connection, args=(conn.id,), daemon=True).start()
        
        return jsonify({'success': True, 'message': 'Синхронизация запущена'})
    except Exception as e:
        app.logger.error(f"Ошибка запуска синхронизации FTP: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/service/start', methods=['POST'])
@login_required
def service_start():
    """Запуск сервиса"""
    try:
        import platform
        if platform.system() == 'Linux':
            ok, msg = run_systemd_command('start')
        else:
            ok, msg = run_script('start_service.bat', wait=False)
        if ok:
            service_status['running'] = True
            service_status['last_start'] = datetime.now()
            return jsonify({'success': True, 'message': 'Сервис запущен'})
        return jsonify({'success': False, 'message': msg})
            
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/service/stop', methods=['POST'])
@login_required
def service_stop():
    """Остановка сервиса"""
    try:
        import platform
        if platform.system() == 'Linux':
            ok, msg = run_systemd_command('stop')
        else:
            ok, msg = run_script('stop_service.bat', wait=True)
        if ok:
            service_status['running'] = False
            service_status['last_stop'] = datetime.now()
            return jsonify({'success': True, 'message': 'Сервис остановлен'})
        return jsonify({'success': False, 'message': msg})
            
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/service/restart', methods=['POST'])
@login_required
def service_restart():
    """Перезапуск сервиса"""
    try:
        import platform
        if platform.system() == 'Linux':
            ok, msg = run_systemd_command('restart')
        else:
            ok, msg = run_script('restart_service.bat', wait=False)
        if ok:
            service_status['running'] = True
            service_status['last_start'] = datetime.now()
            return jsonify({'success': True, 'message': 'Сервис перезапущен'})
        return jsonify({'success': False, 'message': msg})
            
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/logs')
@login_required
def logs_page():
    """Страница логов"""
    return render_template('logs.html', active_page='logs')

@app.route('/api/logs')
@login_required
def api_logs():
    """API ??? ????????? ???????????? ?????."""
    try:
        logs = get_user_logs()
        return jsonify({
            'logs': logs,
            'total': len(logs)
        })
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/config/test')
@login_required
def api_config_test():
    """API ??? ???????? ???????????? ?????? ? ??????????."""
    try:
        runtime_cfg = build_user_runtime_config()
        api_keys = runtime_cfg.get('api_keys', {})
        telegram_cfg = runtime_cfg.get('telegram', {})
        test_results = {}

        # TheB.ai
        try:
            import requests
            headers = {
                'Authorization': f"Bearer {api_keys.get('thebai_api_key', '')}",
                'Content-Type': 'application/json'
            }
            payload = {
                'model': api_keys.get('thebai_model', 'deepseek-reasoner'),
                'messages': [{'role': 'user', 'content': 'ping'}],
                'max_tokens': 10
            }
            url = api_keys.get('thebai_url') or 'https://api.deepseek.com/v1/chat/completions'
            response = requests.post(url, headers=headers, json=payload, timeout=10)
            if response.status_code == 200:
                test_results['thebai'] = {'status': 'success', 'message': 'API ????????'}
            else:
                test_results['thebai'] = {'status': 'error', 'message': f'HTTP {response.status_code}'}
        except Exception as e:
            test_results['thebai'] = {'status': 'error', 'message': str(e)}

        # Telegram
        try:
            import requests
            bot_token = api_keys.get('telegram_bot_token', '')
            if not bot_token:
                raise ValueError('?? ?????? ????? ????')
            url = f"https://api.telegram.org/bot{bot_token}/getMe"
            response = requests.get(url, timeout=10)
            if response.status_code == 200:
                test_results['telegram'] = {'status': 'success', 'message': '??? ????????'}
            else:
                test_results['telegram'] = {'status': 'error', 'message': f'HTTP {response.status_code}'}
        except Exception as e:
            test_results['telegram'] = {'status': 'error', 'message': str(e)}

        # Speechmatics
        try:
            import requests
            speech_key = api_keys.get('speechmatics_api_key', '')
            if not speech_key:
                raise ValueError('?? ?????? Speechmatics API key')
            url = 'https://asr.api.speechmatics.com/v2/jobs'
            headers = {'Authorization': f'Bearer {speech_key}'}
            response = requests.get(url, headers=headers, timeout=10)
            if response.status_code == 200:
                test_results['speechmatics'] = {'status': 'success', 'message': 'Speechmatics ????????'}
            else:
                test_results['speechmatics'] = {'status': 'error', 'message': f'HTTP {response.status_code}'}
        except Exception as e:
            test_results['speechmatics'] = {'status': 'error', 'message': str(e)}

        working_apis = sum(1 for result in test_results.values() if result['status'] == 'success')
        total_apis = len(test_results)

        return jsonify({
            'test_results': test_results,
            'summary': {
                'working_apis': working_apis,
                'total_apis': total_apis,
                'all_working': working_apis == total_apis
            }
        })
    except Exception as e:
        logging.error(f'?????? ???????? ????????????: {e}')
        return jsonify({'error': str(e)})

@app.route('/api/config/save_all', methods=['POST'])
@login_required
def api_config_save_all():
    """API для сохранения всей конфигурации одним запросом (атомарно)."""
    try:
        data = request.get_json() or {}
        config_data = get_user_config_data()

        # Обновляем все секции
        if 'business_profile' in data and data['business_profile']:
            config_data['business_profile'] = data['business_profile']
        if 'api_keys' in data:
            config_data['api_keys'] = data['api_keys']
        
        if 'telegram' in data:
            config_data['telegram'] = data['telegram']
            
        if 'paths' in data:
            old_paths = config_data.get('paths', {})
            new_paths = data['paths']
            config_data['paths'] = new_paths
            
            # Управляем FTP (как в старом методе)
            old_source_type = old_paths.get('source_type', 'local')
            old_ftp_id = old_paths.get('ftp_connection_id')
            new_source_type = new_paths.get('source_type', 'local')
            new_ftp_id = new_paths.get('ftp_connection_id')
            
            if old_source_type != new_source_type or old_ftp_id != new_ftp_id:
                try:
                    from call_analyzer.ftp_sync_manager import start_ftp_sync, stop_ftp_sync
                    if old_source_type == 'ftp' and old_ftp_id:
                        stop_ftp_sync(old_ftp_id)
                    if new_source_type == 'ftp' and new_ftp_id:
                        ftp_conn = FtpConnection.query.filter_by(id=new_ftp_id, user_id=current_user.id, is_active=True).first()
                        if ftp_conn:
                            start_ftp_sync(new_ftp_id)
                except Exception as e:
                    app.logger.error(f"FTP Sync toggle error: {e}")

        if 'transcription' in data:
            config_data['transcription'] = data['transcription']
            
        if 'filename' in data:
            config_data['filename'] = data['filename']
            
        if 'stations' in data:
            st_data = data['stations']
            if 'stations' in st_data and isinstance(st_data['stations'], dict):
                config_data['stations'] = st_data['stations']
            elif 'stations' in st_data and isinstance(st_data['stations'], list):
                config_data['stations'] = {item['code']: item['name'] for item in st_data['stations'] if item.get('code')}
            
            if 'station_chat_ids' in st_data:
                config_data['station_chat_ids'] = st_data['station_chat_ids']
            if 'station_mapping' in st_data:
                config_data['station_mapping'] = st_data['station_mapping']
            if 'nizh_station_codes' in st_data:
                config_data['nizh_station_codes'] = st_data['nizh_station_codes']

        if 'employee_by_extension' in data:
            config_data['employee_by_extension'] = data['employee_by_extension']

        # Сохраняем всё сразу
        save_user_config_data(config_data)
        
        # Перезапуск воркера если нужно
        if 'paths' in data:
            try:
                request_reload(current_user.id)
            except: pass
            
        append_user_log('Сохранена полная конфигурация', module='config')
        return jsonify({'success': True, 'message': 'Все настройки сохранены'})
    except Exception as e:
        app.logger.error(f'Ошибка сохранения всей конфигурации: {e}', exc_info=True)
        return jsonify({'success': False, 'message': str(e)})

@app.route('/api/config/save', methods=['POST'])
@login_required
def api_config_save():
    """API ??? ?????????? ????????? ?????? ???????????????? ????????."""
    try:
        data = request.get_json() or {}
        config_type = data.get('type')
        config_payload = data.get('data') or {}

        if not config_type:
            return jsonify({'success': False, 'message': '?? ?????? ??? ????????????'}), 400

        config_data = get_user_config_data()

        if config_type == 'api_keys':
            config_data['api_keys'] = config_payload
        elif config_type == 'paths':
            # Сохраняем пути
            old_paths = config_data.get('paths', {})
            config_data['paths'] = config_payload
            
            # Управляем FTP синхронизацией при изменении источника
            old_source_type = old_paths.get('source_type', 'local')
            old_ftp_id = old_paths.get('ftp_connection_id')
            new_source_type = config_payload.get('source_type', 'local')
            new_ftp_id = config_payload.get('ftp_connection_id')
            
            # Если изменился источник или FTP подключение
            if old_source_type != new_source_type or old_ftp_id != new_ftp_id:
                try:
                    from call_analyzer.ftp_sync_manager import start_ftp_sync, stop_ftp_sync
                    
                    # Останавливаем старое подключение, если было
                    if old_source_type == 'ftp' and old_ftp_id:
                        stop_ftp_sync(old_ftp_id)
                        app.logger.info(f"Остановлена FTP синхронизация для подключения {old_ftp_id}")
                    
                    # Запускаем новое подключение, если выбрано FTP
                    if new_source_type == 'ftp' and new_ftp_id:
                        # Проверяем, что подключение существует и активно
                        ftp_conn = FtpConnection.query.filter_by(
                            id=new_ftp_id,
                            user_id=current_user.id,
                            is_active=True
                        ).first()
                        if ftp_conn:
                            start_ftp_sync(new_ftp_id)
                            app.logger.info(f"Запущена FTP синхронизация для подключения {new_ftp_id}")
                        else:
                            app.logger.warning(f"FTP подключение {new_ftp_id} не найдено или неактивно")
                except Exception as e:
                    app.logger.error(f"Ошибка управления FTP синхронизацией: {e}")
        elif config_type == 'telegram':
            config_data['telegram'] = config_payload
        elif config_type == 'employee_by_extension':
            config_data['employee_by_extension'] = config_payload
        elif config_type == 'transcription':
            config_data['transcription'] = config_payload
        elif config_type == 'filename':
            # Пользовательские форматы имен файлов
            config_data['filename'] = {
                'enabled': bool(config_payload.get('enabled', False)),
                'patterns': config_payload.get('patterns') or [],
                'extensions': config_payload.get('extensions') or ['.mp3', '.wav']
            }
        elif config_type == 'stations':
            config_data['stations'] = {item['code']: item['name'] for item in config_payload.get('stations', []) if item.get('code')}
            config_data['station_chat_ids'] = config_payload.get('station_chat_ids', {})
            config_data['station_mapping'] = config_payload.get('station_mapping', {})
            config_data['nizh_station_codes'] = config_payload.get('nizh_station_codes', [])
        else:
            return jsonify({'success': False, 'message': f'??????????? ??? ????????????: {config_type}'}), 400

        save_user_config_data(config_data)
        if config_type == 'paths':
            try:
                request_reload(current_user.id)
            except Exception as exc:
                app.logger.warning(f"Не удалось запросить перезапуск воркера: {exc}")
        if config_type == 'stations':
            sync_prompts_from_config()
        append_user_log(f'Обновлён блок конфигурации: {config_type}', module='config')
        return jsonify({'success': True, 'message': '????????? ?????????'})
    except Exception as e:
        app.logger.error(f'?????? ?????????? ????????: {e}')
        return jsonify({'success': False, 'message': str(e)})

def update_api_keys(content, data):
    """Обновляет API ключи в конфигурации"""
    # TheB.ai
    if 'thebai_api_key' in data:
        content = re.sub(
            r'THEBAI_API_KEY = "[^"]*"',
            f'THEBAI_API_KEY = "{data["thebai_api_key"]}"',
            content
        )
    if 'thebai_url' in data:
        content = re.sub(
            r'THEBAI_URL = "[^"]*"',
            f'THEBAI_URL = "{data["thebai_url"]}"',
            content
        )
    if 'thebai_model' in data:
        content = re.sub(
            r'THEBAI_MODEL = "[^"]*"',
            f'THEBAI_MODEL = "{data["thebai_model"]}"',
            content
        )
    
    # Speechmatics
    if 'speechmatics_api_key' in data:
        content = re.sub(
            r'SPEECHMATICS_API_KEY = "[^"]*"',
            f'SPEECHMATICS_API_KEY = "{data["speechmatics_api_key"]}"',
            content
        )
    
    # Telegram Bot Token (отправляется как часть api_keys)
    if 'telegram_bot_token' in data:
        content = re.sub(
            r'TELEGRAM_BOT_TOKEN = "[^"]*"',
            f'TELEGRAM_BOT_TOKEN = "{data["telegram_bot_token"]}"',
            content
        )
    
    return content

def update_paths(content, data):
    """Обновляет пути в конфигурации"""
    if 'base_records_path' in data:
        # Экранируем обратные слеши для регулярного выражения
        escaped_path = data["base_records_path"].replace("\\", "\\\\")
        content = re.sub(
            r'BASE_RECORDS_PATH = Path\("[^"]*"\)',
            f'BASE_RECORDS_PATH = Path("{escaped_path}")',
            content
        )
    if 'prompts_file' in data:
        # Экранируем обратные слеши для регулярного выражения
        escaped_path = data["prompts_file"].replace("\\", "\\\\")
        content = re.sub(
            r'PROMPTS_FILE = Path\("[^"]*"\)',
            f'PROMPTS_FILE = Path("{escaped_path}")',
            content
        )
    if 'additional_vocab_file' in data:
        # Экранируем обратные слеши для регулярного выражения
        escaped_path = data["additional_vocab_file"].replace("\\", "\\\\")
        content = re.sub(
            r'ADDITIONAL_VOCAB_FILE = Path\("[^"]*"\)',
            f'ADDITIONAL_VOCAB_FILE = Path("{escaped_path}")',
            content
        )
    
    return content

def update_telegram(content, data):
    """Обновляет Telegram настройки в конфигурации"""
    if 'telegram_bot_token' in data:
        content = re.sub(
            r'TELEGRAM_BOT_TOKEN = "[^"]*"',
            f'TELEGRAM_BOT_TOKEN = "{data["telegram_bot_token"]}"',
            content
        )
    if 'alert_chat_id' in data:
        content = re.sub(
            r'ALERT_CHAT_ID = "[^"]*"',
            f'ALERT_CHAT_ID = "{data["alert_chat_id"]}"',
            content
        )
    if 'tg_channel_nizh' in data:
        content = re.sub(
            r"TG_CHANNEL_NIZH = '[^']*'",
            f"TG_CHANNEL_NIZH = '{data['tg_channel_nizh']}'",
            content
        )
    if 'tg_channel_other' in data:
        content = re.sub(
            r"TG_CHANNEL_OTHER = '[^']*'",
            f"TG_CHANNEL_OTHER = '{data['tg_channel_other']}'",
            content
        )
    
    return content

def update_employee_by_extension(content, data):
    """Обновляет EMPLOYEE_BY_EXTENSION в конфигурации"""
    employees = data if isinstance(data, dict) else data.get('employee_by_extension', {})
    # Формируем словарь в виде Python-кода
    lines = ["EMPLOYEE_BY_EXTENSION = {"]
    for ext, name in employees.items():
        safe_ext = str(ext).replace("'", "\'")
        safe_name = str(name).replace("'", "\'")
        lines.append(f"    '{safe_ext}': '{safe_name}',")
    lines.append("}")
    block = "\n".join(lines)

    if 'EMPLOYEE_BY_EXTENSION' in content:
        content = re.sub(r'EMPLOYEE_BY_EXTENSION\s*=\s*\{[\s\S]*?\}', block, content)
    else:
        content = content.rstrip() + "\n\n" + block + "\n"
    return content

def update_transcription(content, data):
    """Обновляет настройки транскрипции (TBANK_STEREO_ENABLED)."""
    stereo = bool(data.get('tbank_stereo_enabled', False))
    pattern = r'TBANK_STEREO_ENABLED\s*=\s*(True|False)'
    repl = f'TBANK_STEREO_ENABLED = {str(stereo)}'
    if 'TBANK_STEREO_ENABLED' in content:
        content = re.sub(pattern, repl, content)
    else:
        content = content.rstrip() + f"\n\n{repl}\n"
    return content

def update_stations(content, data):
    """Обновляет станции в конфигурации"""
    app.logger.info(f"update_stations called with data: {data}")
    
    if 'station_mapping' in data:
        # Получаем существующие станции из конфигурации
        existing_stations = {}
        existing_names = {}
        existing_mapping = {}
        
        # Извлекаем существующие STATION_CHAT_IDS
        chat_ids_match = re.search(r'STATION_CHAT_IDS = \{([^}]*)\}', content, re.DOTALL)
        if chat_ids_match:
            chat_ids_content = chat_ids_match.group(1)
            for line in chat_ids_content.split('\n'):
                if "':" in line and "[" in line:
                    station_id = line.split("':")[0].strip().strip("'\"")
                    if station_id:
                        existing_stations[station_id] = True
        
        # Извлекаем существующие STATION_NAMES
        names_match = re.search(r'STATION_NAMES = \{([^}]*)\}', content, re.DOTALL)
        if names_match:
            names_content = names_match.group(1)
            for line in names_content.split('\n'):
                if '":' in line:
                    station_id = line.split('":')[0].strip().strip('"')
                    if station_id:
                        existing_names[station_id] = True
        
        # Извлекаем существующие STATION_MAPPING
        mapping_match = re.search(r'STATION_MAPPING = \{([^}]*)\}', content, re.DOTALL)
        if mapping_match:
            mapping_content = mapping_match.group(1)
            for line in mapping_content.split('\n'):
                if "':" in line and "[" in line:
                    station_id = line.split("':")[0].strip().strip("'\"")
                    if station_id:
                        existing_mapping[station_id] = True
        
        # Обновляем словари, добавляя новые станции к существующим
        mapping_dict = data['station_mapping']
        
        # Формируем STATION_CHAT_IDS (сохраняем существующие + добавляем новые)
        stations_str = "STATION_CHAT_IDS = {\n"
        
        # Добавляем существующие станции (исключая удаляемые и обновляемые)
        if chat_ids_match:
            chat_ids_content = chat_ids_match.group(1)
            for line in chat_ids_content.split('\n'):
                if "':" in line and "[" in line:
                    # Извлекаем ID станции из строки
                    station_id = line.split("':")[0].strip().strip("'\"")
                    # Проверяем, не удаляется ли эта станция и не обновляется ли
                    if station_id not in mapping_dict:
                        stations_str += line.strip() + "\n"
        
        # Добавляем новые/обновленные станции
        for station_id, station_data in mapping_dict.items():
            if station_data is None:
                # Удаляем станцию - пропускаем её
                continue
            if isinstance(station_data, dict) and 'chat_ids' in station_data:
                chat_ids = station_data['chat_ids']
                chat_ids_str = "', '".join(chat_ids)
                stations_str += f"    '{station_id}': ['{chat_ids_str}'],\n"
        
        stations_str += "}"
        
        content = re.sub(
            r'STATION_CHAT_IDS = \{.*?\}',
            stations_str,
            content,
            flags=re.DOTALL
        )
        
        # Формируем STATION_NAMES (сохраняем существующие + добавляем новые)
        names_str = "STATION_NAMES = {\n"
        
        # Добавляем существующие имена станций (исключая удаляемые и обновляемые)
        if names_match:
            names_content = names_match.group(1)
            for line in names_content.split('\n'):
                if '":' in line:
                    # Извлекаем ID станции из строки
                    station_id = line.split('":')[0].strip().strip('"')
                    # Проверяем, не удаляется ли эта станция и не обновляется ли
                    if station_id not in mapping_dict:
                        names_str += line.strip() + "\n"
        
        # Добавляем новые/обновленные имена станций
        for station_id, station_data in mapping_dict.items():
            if station_data is None:
                # Удаляем станцию - пропускаем её
                continue
            if isinstance(station_data, dict) and 'name' in station_data:
                station_name = station_data['name']
                names_str += f'    "{station_id}": "{station_name}",\n'
        
        names_str += "}"
        
        content = re.sub(
            r'STATION_NAMES = \{.*?\}',
            names_str,
            content,
            flags=re.DOTALL
        )
        
        # Формируем STATION_MAPPING (сохраняем существующие + добавляем новые)
        mapping_str = "STATION_MAPPING = {\n"
        
        # Добавляем существующие маппинги (исключая удаляемые и обновляемые)
        if mapping_match:
            mapping_content = mapping_match.group(1)
            for line in mapping_content.split('\n'):
                if "':" in line and "[" in line:
                    # Извлекаем ID станции из строки
                    station_id = line.split("':")[0].strip().strip("'\"")
                    # Проверяем, не удаляется ли эта станция и не обновляется ли
                    if station_id not in mapping_dict:
                        mapping_str += line.strip() + "\n"
        
        # Добавляем новые/обновленные маппинги
        for station_id, station_data in mapping_dict.items():
            if station_data is None:
                # Удаляем станцию - пропускаем её
                continue
            if isinstance(station_data, dict) and 'sub_stations' in station_data:
                sub_stations = station_data['sub_stations']
                sub_stations_str = "', '".join(sub_stations)
                mapping_str += f"    '{station_id}': ['{sub_stations_str}'],\n"
        
        mapping_str += "}"
        
        content = re.sub(
            r'STATION_MAPPING = \{.*?\}',
            mapping_str,
            content,
            flags=re.DOTALL
        )
    
    if 'nizh_station_codes' in data:
        # Формируем новый список кодов станций
        codes = data['nizh_station_codes']
        codes_str = "NIZH_STATION_CODES = [\n"
        for code in codes:
            codes_str += f"    '{code}',\n"
        codes_str += "]"
        
        # Заменяем старый список (включая закомментированные)
        content = re.sub(
            r'#?NIZH_STATION_CODES = \[.*?\]',
            codes_str,
            content,
            flags=re.DOTALL
        )
    
    return content

if __name__ == '__main__':
    # Настраиваем логирование
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(name)s - %(message)s',
        handlers=[
            logging.FileHandler('web_interface.log', encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    
    print("Запуск веб-интерфейса Call Analyzer...")
    print("Откройте браузер и перейдите по адресу: http://localhost:5000")
    print("Для остановки нажмите Ctrl+C")
    
    # Выполняем инициализацию перед запуском
    initialize_app()

    host = os.getenv('FLASK_HOST', '0.0.0.0')
    port = int(os.getenv('FLASK_PORT', 5000))
    debug = app.config.get('DEBUG', False)

    try:
        app.run(host=host, port=port, debug=debug, use_reloader=False)
    except KeyboardInterrupt:
        print("\nОстановка веб-интерфейса...")
        sys.exit(0)