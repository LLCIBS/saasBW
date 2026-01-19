import os
import re
import datetime
from datetime import date, timedelta
import requests
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import pandas as pd
import yaml

try:
    from call_analyzer.utils import ensure_telegram_ready  # type: ignore
except ImportError:
    from utils import ensure_telegram_ready

def load_config(config_path):
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)

def send_telegram_report(file_path, text_message, config):
    if not ensure_telegram_ready('отправка отчёта rr_3'):
        return
    token = config['telegram']['token']
    chat_ids = [chat_id for chat_id in config['telegram']['chats'] if chat_id]
    if not token or not chat_ids:
        print('Telegram не настроен для отчёта rr_3, сообщение пропущено.')
        return
    url_base = f"https://api.telegram.org/bot{token}/sendDocument"
    if not os.path.exists(file_path):
        print(f'Файл {file_path} не найден.')
        return
    for chat_id in chat_ids:
        print(f'Отправка отчёта в чат {chat_id}...')
        with open(file_path, 'rb') as f:
            files = {'document': f}
            data = {'chat_id': chat_id, 'caption': text_message}
            try:
                resp = requests.post(url_base, files=files, data=data)
                if resp.status_code == 200:
                    print(f'Отчёт успешно доставлен в чат {chat_id}.')
                else:
                    print(f'Ошибка {resp.status_code}: {resp.text}')
            except Exception as e:
                print(f'Исключение при отправке отчёта: {e}')
    return
    for chat_id in chat_ids:
        print(f"Отправляем отчет в чат {chat_id}...")
        with open(file_path, 'rb') as f:
            files = {'document': f}
            data = {'chat_id': chat_id, 'caption': text_message}
            try:
                resp = requests.post(url_base, files=files, data=data)
                if resp.status_code == 200:
                    print(f"Отправлено в чат {chat_id}.")
                else:
                    print(f"Ошибка {resp.status_code}: {resp.text}")
            except Exception as e:
                print(f"Ошибка при отправке: {e}")

STATION_MAP = {
    "9309": "Ретрак Сервис",
    "9310": "Ретрак-Центр Коновалова",
    "9311": "Ретрак-Центр Дзержинск",
    "9328": "Ретрак-Центр Невинномысск",
    "9336": "Ретрак Сервис",
    "9339": "Ретрак-Центр Краснодар",
    "9341": "Ретрак-Центр Ярославль",
    "9342": "Ретрак-Юг Волгоград",
    "9345": "Ретрак-Центр Екатеринбург",
    "9346": "Ретрак-Центр Кстово",
    "9348": "Ретрак-Центр Самара",
    "9329": "Ретрак-Центр Балашиха",
    "9349": "Ретрак-Центр Воронеж",
    "9343": "Ретрак-Центр Ставрополь"
}

ALL_STATIONS = list(STATION_MAP.values())
QUESTION_WEIGHTS = [5, 10, 5, 10, 10, 5, 10, 25, 10, 10]
ANSWER_REGEX = re.compile(r'(\d+)\.\s.*?\[ОТВЕТ:\s*(ДА|НЕТ)\]', flags=re.IGNORECASE | re.DOTALL)
ANALYSIS_REGEX = re.compile(r'Результат анализа:\s*([\d\.]+)%', flags=re.IGNORECASE)

def parse_tg_calls(tg_calls_file):
    calls = []
    if not os.path.exists(tg_calls_file):
        print(f"Файл {tg_calls_file} не найден.")
        return calls
    with open(tg_calls_file, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split("-")
            if len(parts) < 9:
                continue
            phone = parts[2].lstrip("+")
            c_full = parts[4].strip()
            surname = c_full  # Используем полное ФИО
            day_str = parts[5]
            month_str = parts[6]
            year_hour = parts[7].split()
            if len(year_hour) < 2:
                continue
            year_str = year_hour[0]
            hour_str = year_hour[1]
            minute_str = parts[8].replace(".mp3", "")
            dt_str = f"{day_str}-{month_str}-{year_str} {hour_str}:{minute_str}"
            try:
                dt_obj = datetime.datetime.strptime(dt_str, "%d-%m-%Y %H:%M")
            except:
                continue
            calls.append({"phone": phone, "datetime": dt_obj, "surname": surname})
    return calls

def find_surname(phone, dt_obj, tg_calls):
    phone = phone.lstrip("+")
    t0 = dt_obj.replace(second=0, microsecond=0)
    for c in tg_calls:
        if c["phone"] == phone:
            t1 = c["datetime"].replace(second=0, microsecond=0)
            diff = abs((t0 - t1).total_seconds())
            if diff <= 300:
                return c["surname"]
    return None

def get_last_week_dates():
    today = date.today()
    start_date = today - timedelta(days=7)
    end_date = today - timedelta(days=1)
    return start_date, end_date

def date_range_list(start_date, end_date):
    delta = end_date - start_date
    return [start_date + timedelta(days=i) for i in range(delta.days + 1)]

def parse_transcript_file(file_path, tg_calls):
    base_name = os.path.basename(file_path)
    if not os.path.exists(file_path):
        return None
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    found_answers = ANSWER_REGEX.findall(content)
    if not found_answers:
        return None
    ans_dict = {}
    for q_idx_str, yesno_str in found_answers:
        try:
            q_idx = int(q_idx_str)
        except:
            continue
        if 1 <= q_idx <= 10 and q_idx not in ans_dict:
            ans_dict[q_idx] = yesno_str.upper()
    if not all(i in ans_dict for i in range(1, 11)):
        return None
    analysis_match = ANALYSIS_REGEX.search(content)
    analysis_percent = 0.0
    if analysis_match:
        try:
            analysis_percent = float(analysis_match.group(1))
        except:
            analysis_percent = 0.0
    parts = base_name.split("_")
    if len(parts) < 4:
        return None
    station_code = parts[2]
    station_name = STATION_MAP.get(station_code, f"Неизвестная станция {station_code}")
    phone_str = parts[1].strip()
    dt_str = parts[3].strip()
    try:
        dt_obj = datetime.datetime.strptime(dt_str, "%Y-%m-%d-%H-%M-%S")
    except:
        dt_obj = None
    final_score = 0
    final_answers = []
    for i in range(1, 11):
        yesno = ans_dict[i]
        binval = 1 if yesno == "ДА" else 0
        final_score += binval * QUESTION_WEIGHTS[i-1]
        final_answers.append(binval)
    consultant_surname = None
    if dt_obj:
        consultant_surname = find_surname(phone_str, dt_obj, tg_calls)
    if not consultant_surname:
        consultant_surname = base_name
    return {
        "filename": consultant_surname,
        "station_name": station_name,
        "answers": final_answers,
        "score": final_score,
        "analysis_percent": analysis_percent
    }

def collect_weekly_data(base_folder, tg_calls):
    last_monday, last_sunday = get_last_week_dates()
    date_list = date_range_list(last_monday, last_sunday)
    results = []
    for d in date_list:
        folder = os.path.join(base_folder, d.strftime("%Y"), d.strftime("%m"), d.strftime("%d"), "transcriptions", "retrack")
        if not os.path.isdir(folder):
            continue
        for fname in os.listdir(folder):
            if fname.lower().endswith(".txt"):
                fpath = os.path.join(folder, fname)
                parsed = parse_transcript_file(fpath, tg_calls)
                if parsed:
                    results.append(parsed)
    return results

def create_weekly_report(data_list, output_folder, last_monday, last_sunday):
    report_filename = f"Ретрак_Отчёт_{last_monday.strftime('%d.%m')}-{last_sunday.strftime('%d.%m')}.xlsx"
    full_path = os.path.join(output_folder, report_filename)

    rows = []
    for item in data_list:
        row = {
            "файл": item["filename"],
            "станция": item["station_name"],
            "баллы": item["score"],
            # Было "Процент анализа", теперь назовем "процент"
            "процент": item["analysis_percent"]
        }
        for i, val in enumerate(item["answers"], start=1):
            row[f"{i}"] = val
        rows.append(row)

    # Было: "файл", "станция", "баллы", "Процент анализа", 1..10
    # Теперь "процент" вместо "Процент анализа":
    cols = ["файл", "станция", "баллы", "процент"] + [f"{i}" for i in range(1, 11)]
    df = pd.DataFrame(rows, columns=cols)

    if not df.empty:
        known = df["станция"].unique()
    else:
        known = []
    all_st = sorted(list(set(ALL_STATIONS).union(known)))

    wb = Workbook()
    ws_rating = wb.active
    ws_rating.title = "Рейтинг"
    ws_summary = wb.create_sheet("Сводный отчёт", 1)
    ws_data = wb.create_sheet("Данные", 2)

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    # ------------------------
    # 1) Лист "Рейтинг"
    # ------------------------
    ws_rating.cell(row=1, column=1, value="Станция")
    ws_rating.cell(row=1, column=2, value="План")
    ws_rating.cell(row=1, column=3, value="Факт")
    ws_rating.cell(row=1, column=4, value="%")

    for c in range(1, 5):
        cell = ws_rating.cell(row=1, column=c)
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border

    rating_rows = []
    total_plan = 0
    total_fact = 0
    for st_name in all_st:
        sub_df = df[df["станция"] == st_name]
        call_count = len(sub_df)
        plan = call_count * 100
        fact = sub_df["баллы"].sum() if call_count > 0 else 0
        prc = fact / plan if plan else 0
        total_plan += plan
        total_fact += fact
        rating_rows.append((st_name, plan, fact, prc))

    rating_rows.sort(key=lambda x: x[3], reverse=True)

    row_num = 2
    for st, pl, fc, pr in rating_rows:
        ws_rating.cell(row=row_num, column=1, value=st)
        ws_rating.cell(row=row_num, column=2, value=pl)
        ws_rating.cell(row=row_num, column=3, value=fc)
        ws_rating.cell(row=row_num, column=4, value=pr)

        for col_i in range(1, 5):
            cl = ws_rating.cell(row=row_num, column=col_i)
            cl.alignment = center_alignment
            cl.border = thin_border
            if col_i == 4:
                cl.number_format = "0%"

        row_num += 1

    ws_rating.cell(row=row_num, column=1, value="Итого")
    ws_rating.cell(row=row_num, column=1).font = bold_font
    ws_rating.cell(row=row_num, column=1).alignment = center_alignment
    ws_rating.cell(row=row_num, column=1).border = thin_border

    ws_rating.cell(row=row_num, column=2, value=total_plan)
    ws_rating.cell(row=row_num, column=3, value=total_fact)

    total_prc = total_fact / total_plan if total_plan else 0
    cellp = ws_rating.cell(row=row_num, column=4, value=total_prc)
    cellp.number_format = "0%"

    for c in range(2, 5):
        c_ = ws_rating.cell(row=row_num, column=c)
        c_.font = bold_font
        c_.alignment = center_alignment
        c_.border = thin_border

    ws_rating.column_dimensions["A"].width = 35
    ws_rating.column_dimensions["B"].width = 10
    ws_rating.column_dimensions["C"].width = 10
    ws_rating.column_dimensions["D"].width = 6

    # ------------------------
    # 2) Лист "Сводный отчёт"
    # ------------------------
    # Было: df.sort_values(by=["станция","Процент анализа"])
    # Теперь: df.sort_values(by=["станция","процент"])
    if not df.empty:
        df_srt = df.sort_values(by=["станция", "процент"], ascending=[True, False])
    else:
        df_srt = df

    row_summary = 1
    # В шапке каждой станции выводим столбцы
    # "файл","баллы","процент", и 1..10
    headers = ["файл", "баллы", "процент"] + [f"{i}" for i in range(1, 11)]

    for station_name in df_srt["станция"].unique():
        ws_summary.cell(row=row_summary, column=1, value=f"Станция: {station_name}")
        ws_summary.cell(row=row_summary, column=1).font = bold_font
        row_summary += 1

        # Вывод шапки
        for col_i, head in enumerate(headers, start=1):
            cc = ws_summary.cell(row=row_summary, column=col_i, value=head)
            cc.font = bold_font
            cc.alignment = center_alignment
            cc.border = thin_border
        row_summary += 1

        subdf = df_srt[df_srt["станция"] == station_name]

        # Вывод строк. Нужно «процент» показывать с символом '%'
        for _, row_data in subdf.iterrows():
            col_i = 1
            for head in headers:
                val = row_data[head]
                cell = ws_summary.cell(row=row_summary, column=col_i)

                # Если head == "процент", добавим символ '%'
                if head == "процент":
                    cell.value = f"{val:.2f}%"  # форматируем число + '%'
                else:
                    cell.value = val

                cell.alignment = center_alignment
                cell.border = thin_border
                col_i += 1

            row_summary += 1

        row_summary += 2

    ws_summary.column_dimensions["A"].width = 50

    # Вставляем **список вопросов** после таблицы
    questions = [
        "1. Сотрудник назвал станцию, представился сам и поприветствовал клиента?",
        "2. Сотрудник выяснил имя клиента и обращался к нему по имени не менее двух раз?",
        "3. Выяснил причину обращения?",
        "4. Сотрудник кратко ответил на вопрос клиента, а затем вежливо перехватил инициативу?",
        "5. Сотрудник выражал позитивный настрой?",
        "6. Сказал 'спасибо за ожидание', если была пауза?",
        "7. Демонстрировал тех. грамотность и профессионализм?",
        "8. Был ли достигнут результат разговора?",
        "9. Подвёл итог разговора и спросил, остались ли вопросы?",
        "10. Поблагодарил клиента и пожелал хорошего дня?"
    ]

    # Допустим, выводим их начиная с row_summary + 1
    for q_text in questions:
        ws_summary.cell(row=row_summary, column=1, value=q_text)
        ws_summary.cell(row=row_summary, column=1).alignment = Alignment(horizontal='left')
        row_summary += 1

    # ------------------------
    # 3) Лист "Данные"
    # ------------------------
    all_cols = list(df.columns)
    for ci, cn in enumerate(all_cols, start=1):
        c = ws_data.cell(row=1, column=ci, value=cn)
        c.font = bold_font
        c.border = thin_border
        c.alignment = center_alignment

    for i, rd in df.iterrows():
        rr = i + 2
        for ci, cn in enumerate(all_cols, start=1):
            val = rd[cn]
            cl = ws_data.cell(row=rr, column=ci)

            # Если это столбец "процент", тоже хотим символ '%':
            if cn == "процент":
                cl.value = f"{val:.2f}%"
            else:
                cl.value = val

            cl.border = thin_border
            cl.alignment = center_alignment

    ws_data.column_dimensions["A"].width = 50
    ws_data.column_dimensions["B"].width = 30

    wb.save(full_path)
    print(f"Отчет сохранен в {full_path}")
    return full_path

def run_rr_3():
    # Используем основную конфигурацию вместо ReTruck
    import config as main_config
    base_folder = str(main_config.BASE_RECORDS_PATH)
    tg_log_file = os.path.join(base_folder, "tg_calls.txt")
    tg_calls = parse_tg_calls(tg_log_file)
    last_monday, last_sunday = get_last_week_dates()
    period_text = f"Отчет за неделю {last_monday.strftime('%d.%m')} - {last_sunday.strftime('%d.%m')}"
    calls_data = collect_weekly_data(base_folder, tg_calls)
    if not calls_data:
        print("Нет файлов с полным набором из 10 ответов.")
        return
    today = date.today()
    year_str = today.strftime("%Y")
    month_str = today.strftime("%m")
    day_str = today.strftime("%d")
    output_folder = os.path.join(base_folder, year_str, month_str, day_str, "transcriptions", "retrack", "R-reports")
    os.makedirs(output_folder, exist_ok=True)
    report_path = create_weekly_report(calls_data, output_folder, last_monday, last_sunday)
    print(f"Отчет сформирован: {report_path}")
    # Создаем конфиг для отправки в Telegram
    telegram_config = {
        'telegram': {
            'token': main_config.TELEGRAM_BOT_TOKEN,
            'chats': [main_config.ALERT_CHAT_ID]
        }
    }
    send_telegram_report(report_path, period_text, telegram_config)

if __name__ == "__main__":
    run_rr_3()
