import os
import sys
import pandas as pd
from datetime import datetime, date
import requests
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from pathlib import Path
from collections import Counter, defaultdict
import re
import io
import contextlib

try:
    from call_analyzer.utils import ensure_telegram_ready  # type: ignore
except ImportError:
    from utils import ensure_telegram_ready

# Добавляем путь к родительской папке для импорта config
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config

# === Whynot Analysis Configuration ===
BASE_DIR = Path(str(config.BASE_RECORDS_PATH))
DEEPSEEK_KEY = config.THEBAI_API_KEY
MODEL_URL = config.THEBAI_URL
MODEL_ID = config.THEBAI_MODEL
PROMPT = (
    "Ты специалист контроля качества кол-центра.\n"
    "Мне нужно чтобы ты на основании транскрипции диалога определил причину, "
    "по которой клиент не был записан на приём. Я дам список возможных причин, "
    "верни только одну строку с точной причиной и ничего более.\n"
    "Вот список возможных причин:\n"
    "нет времени/занято/долгая запись;\n"
    "подумает/отказ клиента;\n"
    "высокая стоимость/дорого;\n"
    "не выполняем работы;\n"
    "свои запчасти клиента;\n"
    "другое\n"
    "Определи, какая из них подходит к диалогу и напиши только цитатную причину."
)
MISSED_CALLBACK_REASON = "не перезвонили за отведённое время"
POSSIBLE_REASONS = [
    "нет времени/занято/долгая запись",
    "подумает/отказ клиента",
    "высокая стоимость/дорого",
    "не выполняем работы",
    "свои запчасти клиента",
    "другое",
    MISSED_CALLBACK_REASON
]
# Тег может быть записан в виде [ТИПЗВОНКА:ЦЕЛЕВОЙ], [ТИП ЗВОНКА:ЦЕЛЕВОЙ] и т.д.
TAG_TARGET = re.compile(r"\[ТИП\s*ЗВОНКА\s*:\s*ЦЕЛЕВОЙ\]", re.I)
TAG_SIGNUP = re.compile(r"\[РЕЗУЛЬТАТ\s*:\s*ЗАПИСЬ\]", re.I)
TAG_NO = re.compile(r"\[РЕЗУЛЬТАТ\s*:\s*НЕТ\]", re.I)
TAG_REFUSAL = re.compile(r"\[РЕЗУЛЬТАТ\s*:\s*ОТКАЗ\]", re.I)
TAG_RECALL = re.compile(r"\[РЕЗУЛЬТАТ\s*:\s*ПЕРЕЗВОНИТЬ\]", re.I)
TAG_TRANSFER = re.compile(r"\[РЕЗУЛЬТАТ\s*:\s*ПЕРЕВОД\]", re.I)
TAG_MESSENGER = re.compile(r"\[РЕЗУЛЬТАТ\s*:\s*МЕССЕНДЖЕР\]", re.I)
# Универсальный паттерн для любых тегов [РЕЗУЛЬТАТ: ...]
TAG_RESULT_ANY = re.compile(r"\[результат\s*:\s*(запись|нет|перезвонить|перевод|мессенджер|отказ)\]", re.I)
TAG_PRIMARY_RECALL_HOUR = re.compile(r"\[перезвонить\s*:\s*час\]", re.I)
TAG_PRIMARY_RECALL_WHEN = re.compile(r"\[перезвонить\s*:\s*когда=([^\]]+)\]", re.I)
TAG_PRIMARY_TRANSFER_HOUR = re.compile(r"\[перевод\s*:\s*условия\s*=\s*час\]", re.I)
TAG_PRIMARY_TRANSFER_WHEN = re.compile(r"\[перевод\s*:\s*условия\s*=\s*([^\]]+)\]", re.I)


def get_main_station_code(station_code):
    """
    Преобразует код подстанции в основной код станции.
    
    Args:
        station_code (str): Код станции (может быть основной или подстанция)
        
    Returns:
        str: Основной код станции или None, если не найден
    """
    # Сначала проверяем, есть ли код в основных станциях
    if station_code in config.STATION_NAMES:
        return station_code
    
    # Ищем в маппинге подстанций
    for main_code, sub_codes in config.STATION_MAPPING.items():
        if station_code in sub_codes:
            return main_code
    
    return None


def ask_model(dialog: str) -> str:
    print(f"ask_model: отправляю запрос LLM, длина диалога = {len(dialog)}")
    headers = {"Authorization": f"Bearer {DEEPSEEK_KEY}", "Content-Type": "application/json"}
    payload = {"model": MODEL_ID, "messages": [{"role": "user", "content": f"{PROMPT}\n\nДиалог:\n{dialog}"}]}
    r = requests.post(MODEL_URL, headers=headers, json=payload, timeout=90)
    r.raise_for_status()
    return r.json()["choices"][0]["message"]["content"].strip().lower()


def parse_day(day: date) -> dict:
    print(f"parse_day: формирование данных по причинам для даты {day}")
    print(f"parse_day: дата = {day}")
    folder = BASE_DIR / f"{day:%Y}" / f"{day:%m}" / f"{day:%d}" / "transcriptions"
    print(f"parse_day: папка = {folder}")
    data = defaultdict(lambda: {"target": 0, "signup": 0, "reasons": Counter()})
    if not folder.exists(): return data

    # Обработка primary файлов (recall and transfer) в поддиректориях recall_analysis и transfer_analysis
    primary_phones = {'recall': set(), 'transfer': set()}
    for follow_up_type in ['recall', 'transfer']:
        analysis_dir = folder / f"{follow_up_type}_analysis"
        tag_hour = TAG_PRIMARY_RECALL_HOUR if follow_up_type == 'recall' else TAG_PRIMARY_TRANSFER_HOUR
        tag_when = TAG_PRIMARY_RECALL_WHEN if follow_up_type == 'recall' else TAG_PRIMARY_TRANSFER_WHEN
        prefix = f"primary_{follow_up_type}_"
        if analysis_dir.exists():
            for prim in analysis_dir.glob(f"{prefix}*.txt"):
                ptext = prim.read_text(encoding="utf-8", errors="ignore")
                parts = prim.name.split("_")
                st = parts[3] if len(parts) > 3 else None
                phone = parts[2] if len(parts) > 2 else None
                
                # Преобразуем код подстанции в основной код станции для primary файлов
                if st:
                    main_station = get_main_station_code(st)
                    if main_station:
                        original_st = st
                        st = main_station
                        print(f"parse_day: primary файл - подстанция {original_st} преобразована в основную станцию {main_station}")
                
                if phone:
                    primary_phones[follow_up_type].add(phone)
                # CHAS tag: count missed only if no follow-up
                if tag_hour.search(ptext):
                    if phone and any(find_follow_up_result(phone, folder, follow_up_type)):
                        print(f"parse_day: primary {follow_up_type} CHAS -> follow-up найден, station={st}")
                    else:
                        data[st]["reasons"][MISSED_CALLBACK_REASON] += 1
                        print(f"parse_day: primary {follow_up_type} CHAS -> пропущенный follow-up, station={st}")
                # WHEN tag: schedule callback/transfer
                m = tag_when.search(ptext)
                if m:
                    val = m.group(1).strip()
                    # skip scheduling 'час' only for primary transfer
                    if not (follow_up_type == 'transfer' and val.lower() == 'час'):
                        reason = f"запланирован {'перезвон' if follow_up_type=='recall' else 'перевод'}: {val}"
                        data[st]["reasons"][reason] += 1
                        print(f"parse_day: primary {follow_up_type} WHEN -> {reason}, station={st}")

    print(f"parse_day: найдено {len(list(folder.glob('*.txt')))} файлов транскрипций")
    for fpath in folder.glob("*.txt"):
        print(f"parse_day: обрабатывается файл {fpath.name}")
        text = fpath.read_text(encoding="utf-8", errors="ignore")
        name_lower = fpath.name.lower()
        # пропустить все primary файлы во втором цикле
        if name_lower.startswith("primary_"):
            continue
        parts = fpath.name.split('_')
        
        # Определяем код станции, поддерживая оба формата
        if len(parts) < 3:
            continue
            
        first_part = parts[1]
        second_part = parts[2]
        
        # Проверяем, является ли first_part известным кодом станции (включаем основные станции и подстанции)
        all_station_codes = set(config.STATION_NAMES.keys())
        # Добавляем все коды подстанций
        for sub_codes in config.STATION_MAPPING.values():
            all_station_codes.update(sub_codes)
        
        if first_part in all_station_codes:
            # Формат: fs_[station_code]_[phone_number]_[datetime]_...
            st = first_part
        elif second_part in all_station_codes:
            # Формат: fs_[phone_number]_[station_code]_[datetime]_...
            st = second_part
        else:
            # Если ни first_part, ни second_part не являются известными кодами станций,
            # пропускаем такой файл
            print(f"parse_day: неизвестные коды станций в файле {fpath.name}: {first_part}, {second_part}")
            continue
        
        # Преобразуем код подстанции в основной код станции
        original_st = st
        main_station = get_main_station_code(st)
        if main_station:
            st = main_station
            if original_st != main_station:
                print(f"parse_day: подстанция {original_st} преобразована в основную станцию {main_station}")
        else:
            print(f"parse_day: неизвестный код станции {st}, используем как есть")
            
        print(f"parse_day: файл {fpath.name} отнесён к станции {st}")
        # Проверяем наличие целевого тега
        if TAG_TARGET.search(text) is None:
            print(f"parse_day: пропускаем нецелевой звонок {fpath.name}")
            continue

        # Проверяем наличие корректных тегов класса и результата
        lower_text = text.lower()
        has_class_tag = any(t in lower_text for t in ("[класс:а]", "[класс: а]", "[класс:б]", "[класс: б]"))
        has_result_tag = any([
            TAG_SIGNUP.search(text), TAG_NO.search(text), TAG_REFUSAL.search(text),
            TAG_MESSENGER.search(text), TAG_RECALL.search(text), TAG_TRANSFER.search(text)
        ])
        if not (has_class_tag and has_result_tag):
            print(f"parse_day: пропускаем файл без корректных class/result тегов {fpath.name}")
            continue

        data[st]["target"] += 1
        print(f"parse_day: найден ТИПЗВОНКА, station={st}, total_target={data[st]['target']}")

        # 1. Запись сразу
        if TAG_SIGNUP.search(text):
            data[st]["signup"] += 1
            print(f"parse_day: найден ЗАПИСЬ, station={st}, total_signup={data[st]['signup']}")
            continue
        # 2. Явный отказ
        if TAG_REFUSAL.search(text):
            data[st]["reasons"]["подумает/отказ клиента"] += 1
            print(f"parse_day: найден ОТКАЗ, station={st}, total_reasons={data[st]['reasons']['подумает/отказ клиента']}")
            continue
        # 3. Перешли в мессенджер
        if TAG_MESSENGER.search(text):
            data[st]["reasons"]["перешли в мессенджер"] += 1
            print(f"parse_day: найден МЕССЕНДЖЕР, station={st}, total_reasons={data[st]['reasons']['перешли в мессенджер']}")
            continue
        # 4. Теги перезвон / перевод – проверяем follow-up
        follow_handled = False
        phone_number = None
        follow_type = None
        if TAG_RECALL.search(text):
            phone_number = get_phone_number_from_filename(fpath.name)
            print(f"parse_day: найден ПЕРЕЗВОНИТЬ в файле {fpath.name}, phone={phone_number}")
            follow_type = "recall"
        elif TAG_TRANSFER.search(text):
            phone_number = get_phone_number_from_filename(fpath.name)
            print(f"parse_day: найден ПЕРЕВОД в файле {fpath.name}, phone={phone_number}")
            follow_type = "transfer"
        # Пропустить root TAG_RECALL/TAG_TRANSFER для первичных номеров
        if phone_number and follow_type and phone_number in primary_phones.get(follow_type, set()):
            print(f"parse_day: пропускаем root {follow_type} для primary phone={phone_number}, station={st}")
            continue
        if phone_number and follow_type:
            follow = find_follow_up_result(phone_number, folder, follow_type)
            print(f"parse_day: follow-up result for phone={phone_number}, type={follow_type}: {follow}")
            # follow = (is_record, is_no, is_connected, is_messenger, is_refusal)
            is_record, is_no, is_connected, is_messenger, is_refusal = follow
            if is_record or is_connected:
                data[st]["signup"] += 1
                print(f"parse_day: follow-up -> запись, station={st}, total_signup={data[st]['signup']}")
                follow_handled = True
            elif is_messenger:
                data[st]["reasons"]["перешли в мессенджер"] += 1
                print(f"parse_day: follow-up -> мессенджер, station={st}, total_reasons={data[st]['reasons']['перешли в мессенджер']}")
                follow_handled = True
            elif is_refusal:
                data[st]["reasons"]["подумает/отказ клиента"] += 1
                print(f"parse_day: follow-up -> отказ, station={st}, total_reasons={data[st]['reasons']['подумает/отказ клиента']}")
                follow_handled = True
            elif is_no:
                # follow-up: нет -> отправляем LLM на анализ причины
                raw = ask_model(text)
                reason = raw if raw in POSSIBLE_REASONS else "другое"
                data[st]["reasons"][reason] += 1
                print(f"parse_day: follow-up -> нет, LLM причина={reason}, station={st}, total_reasons={data[st]['reasons'][reason]}")
                follow_handled = True
            elif not any(follow):
                # пропущенный follow-up -> учитываем отдельно
                data[st]["reasons"][MISSED_CALLBACK_REASON] += 1
                print(f"parse_day: follow-up -> пропущенный callback, reason={MISSED_CALLBACK_REASON}, station={st}, total_reasons={data[st]['reasons'][MISSED_CALLBACK_REASON]}")
                follow_handled = True
        if follow_handled:
            continue
        # 4. Обычное 'нет' – запрашиваем модель
        if TAG_NO.search(text):
            raw = ask_model(text)
            reason = raw if raw in POSSIBLE_REASONS else "другое"
            data[st]["reasons"][reason] += 1
            print(f"parse_day: найден ТЭГ НЕТ в файле {fpath.name}, LLM причина={reason}, station={st}, total_reasons={data[st]['reasons'][reason]}")
    return data


def build_detailed_report_sheet(wb: Workbook, date_label: str, data: dict):
    print(f"build_detailed_report_sheet: формирование листа 'Отчёт' для {date_label}")
    ws = wb.active;
    ws.title = "Отчёт"
    stations = sorted(data)
    station_labels = [get_station_name(st) for st in stations]
    header = ["№", "Причина", "Кол-во"] + station_labels
    for c, v in enumerate(header, 1):
        cell = ws.cell(row=1, column=c, value=v);
        cell.fill = PatternFill("solid", "0066CC", "0066CC");
        cell.font = Font(color="FFFFFF");
        cell.alignment = Alignment(horizontal="center", vertical="center")
    print(f"build_detailed_report_sheet: заголовки: {header}")
    tot_t = sum(data[s]["target"] for s in stations);
    tot_s = sum(data[s]["signup"] for s in stations);
    tot_n = sum(data[s]["reasons"].get(reason, 0) for s in stations for reason in POSSIBLE_REASONS)
    print(f"build_detailed_report_sheet: итоги – цели={tot_t}, записаны={tot_s}, незаписанные={tot_n}")
    for i, lab, val in [(2, "Всего целевых звонков", tot_t), (3, "Количество записанных", tot_s), (4, "Количество незаписанных", tot_n)]:
        ws.cell(row=i, column=2, value=lab).font = Font(color="FF0000");
        ws.cell(row=i, column=3, value=val).font = Font(color="FF0000")
        print(f"build_detailed_report_sheet: записал '{lab}' со значением {val} в строку {i}")
    for idx, st in enumerate(stations, 4):
        ws.cell(row=2, column=idx, value=data[st]["target"]).font = Font(color="FF0000")
        ws.cell(row=3, column=idx, value=data[st]["signup"]).font = Font(color="FF0000")
        unsched = sum(data[st]["reasons"].get(reason, 0) for reason in POSSIBLE_REASONS)
        ws.cell(row=4, column=idx, value=unsched).font = Font(color="FF0000")
        print(f"build_detailed_report_sheet: станция {st} – цели={data[st]['target']}, записаны={data[st]['signup']}, незаписанные={unsched} в колонке {idx}")
    # собираем все причины, включая динамические primary
    reason_keys = set()
    for s in stations:
        reason_keys.update(data[s]["reasons"].keys())
    sorted_r = sorted(reason_keys, key=lambda r: (-sum(data[s]["reasons"].get(r, 0) for s in stations), r))
    totals = {r: sum(data[s]["reasons"].get(r, 0) for s in stations) for r in sorted_r}
    for i, r in enumerate(sorted_r, 5):
        ws.cell(row=i, column=1, value=i - 4);
        ws.cell(row=i, column=2, value=r);
        ws.cell(row=i, column=3, value=totals[r])
        print(f"build_detailed_report_sheet: причина '{r}', общий счёт={totals[r]} на строке {i}")
        # если пропущенный follow-up, выделяем всю строку красным
        if r == MISSED_CALLBACK_REASON:
            for col in range(1, 4 + len(stations)):
                ws.cell(row=i, column=col).font = Font(color="FF0000")
        # Если причина «Перешли в мессенджер» – выделяем строку жирным шрифтом
        if isinstance(r, str) and r.lower() == "перешли в мессенджер":
            for col in range(1, 4 + len(stations)):
                ws.cell(row=i, column=col).font = Font(bold=True)
        for j, st in enumerate(stations, 4):
            ws.cell(row=i, column=j, value=data[st]["reasons"].get(r, 0))
            print(f"build_detailed_report_sheet: станция '{st}', причина '{r}', значение={data[st]['reasons'].get(r, 0)} в ячейке ({i},{j})")
    ws.row_dimensions[1].height = 30;
    ws.column_dimensions["A"].width = 3;
    ws.column_dimensions["B"].width = 40
    for c in range(4, 4 + len(stations)):
        ws.column_dimensions[get_column_letter(c)].width = 6
    thin = Side("thin", "000000");
    border = Border(thin, thin, thin, thin)
    for row in ws.iter_rows(min_row=1, max_row=4 + len(sorted_r), min_col=1, max_col=3 + len(stations)):
        for cell in row:
            if cell.value is not None: cell.border = border


# =====================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# =====================================================

def get_station_name(station_code):
    """Преобразует код станции в её человекочитаемое название."""
    station_names = {
        '9322': 'Брн',
        '9326': 'Рпб',
        '9347': 'Род',
        '9327': 'Чон',
        '9325': 'Чон К',
        '9300': 'Сах',
        '9321': 'Кмн',
        '9324': 'Хлз',
        '9344': 'Кбш',
        '9304': 'Дзр',
        '9308': 'Арз',
        '9301': 'Влд',
        '9302': 'КзнИ',
        '9307': 'КзнС',
        '9350': 'Рдн',
        '9316': 'Меч',
        '9319': 'Тгн'
    }
    return station_names.get(station_code, station_code)


def get_phone_number_from_filename(filename):
    """
    Извлекает номер телефона из названия файла.
    Поддерживает форматы из конфигурации FILENAME_FORMATS:
    - fs_79209153888_9301_2025-03-27-09-09-32_... (телефон первым)
    - fs_9301_79209153888_2025-03-27-09-09-32_... (станция первой)
    Возвращает, например, '79209153888'
    """
    parts = filename.split("_")
    if len(parts) < 3:
        return None
    
    first_id = parts[1]
    second_id = parts[2]
    
    # Проверяем, является ли first_id известным кодом станции (включаем основные станции и подстанции)
    all_station_codes = set(config.STATION_NAMES.keys())
    # Добавляем все коды подстанций
    for sub_codes in config.STATION_MAPPING.values():
        all_station_codes.update(sub_codes)
    
    if first_id in all_station_codes:
        # Формат: fs_[station_code]_[phone_number]_...
        return second_id
    else:
        # Формат: fs_[phone_number]_[station_code]_...
        return first_id


# =====================================================
#  ПАРСИНГ ОДНОГО ФАЙЛА: БЕРЁМ ПЕРВЫЙ RESULT ПОСЛЕ TARGET
# =====================================================

def analyze_file(filepath):
    """
    Анализ текстового файла и возврат статистики по звонкам.

    Возвращает кортеж булевых значений:
    (is_target, class_a, class_b,
     result_record, result_no, result_recall, result_transfer, result_messenger, result_refusal)

    is_target = содержит ли файл тег [ТИПЗВОНКА: ЦЕЛЕВОЙ].
    """
    with open(filepath, 'r', encoding='utf-8') as file:
        content = file.read().lower()

    # Ищем тег [ТИПЗВОНКА:ЦЕЛЕВОЙ]
    m_target = TAG_TARGET.search(content)
    is_target = m_target is not None

    # Поиск тегов класса выполняем по всему тексту
    class_a = '[класс:а]' in content or '[класс: а]' in content
    class_b = '[класс:б]' in content or '[класс: б]' in content

    # По умолчанию все result-флаги False
    result_record = result_no = result_recall = result_transfer = result_messenger = result_refusal = False

    if is_target:
        # Находим ПЕРВЫЙ тег [РЕЗУЛЬТАТ: ...] ПОСЛЕ целевого тега
        pos_after_target = m_target.end()
        m_res = TAG_RESULT_ANY.search(content, pos_after_target)
        if m_res:
            res_type = m_res.group(1)
            if res_type == 'запись':
                result_record = True
            elif res_type == 'нет':
                result_no = True
            elif res_type == 'перезвонить':
                result_recall = True
            elif res_type == 'перевод':
                result_transfer = True
            elif res_type == 'мессенджер':
                result_messenger = True
            elif res_type == 'отказ':
                result_refusal = True
    # Подробный вывод всех найденных тегов для файла
    print(f"analyze_file: {os.path.basename(filepath)} -> "
          f"target={is_target} "
          f"class_a={class_a} class_b={class_b} "
          f"record={result_record} no={result_no} recall={result_recall} "
          f"transfer={result_transfer} messenger={result_messenger} refusal={result_refusal}")

    return (
        is_target,
        class_a, class_b,
        result_record, result_no, result_recall, result_transfer, result_messenger, result_refusal)


def analyze_follow_up_file(filepath, follow_up_type):
    """
    Анализирует файл follow-up.
    follow_up_type: 'recall' или 'transfer'
    Возвращает кортеж из 5 булевых:
    (is_record, is_no, is_connected, is_messenger, is_refusal)
    Где для recall проверяются:
      [перезвонить:запись], [перезвонить:нет], [перезвонить:связались],
      [перезвонить:мессенджер], [перезвон:отказ]
    А для transfer – аналогично с тегами [перевод:...]
    """
    with open(filepath, 'r', encoding='utf-8') as f:
        text = f.read().lower()

    if follow_up_type == 'recall':
        tag_record = '[перезвонить:запись]'
        tag_no = '[перезвонить:нет]'
        tag_connected = '[перезвонить:связались]'
        tag_messenger = '[перезвонить:мессенджер]'
        tag_refusal = '[перезвон:отказ]'
    else:  # transfer
        tag_record = '[перевод:запись]'
        tag_no = '[перевод:нет]'
        tag_connected = '[перевод:связались]'
        tag_messenger = '[перевод:мессенджер]'
        tag_refusal = '[перевод:отказ]'

    is_record = tag_record in text
    is_no = tag_no in text
    is_connected = tag_connected in text
    is_messenger = tag_messenger in text
    is_refusal = tag_refusal in text
    # Подробный вывод для follow-up файла
    print(f"analyze_follow_up_file: {os.path.basename(filepath)} type={follow_up_type} -> "
          f"record={is_record} no={is_no} connected={is_connected} "
          f"messenger={is_messenger} refusal={is_refusal}")
    return (is_record, is_no, is_connected, is_messenger, is_refusal)


def find_follow_up_result(phone_number, base_folder, follow_up_type):
    """
    Ищет в папке base_folder/<follow_up_type>_analysis файл, имя которого начинается с
    '{follow_up_type}_{phone_number}_'. Если найден, анализирует и возвращает tuple из 5 булевых,
    иначе возвращает (False, False, False, False, False).
    """
    folder = os.path.join(base_folder, f"{follow_up_type}_analysis")
    print(f"find_follow_up_result: searching folder={folder} pattern={follow_up_type}_{phone_number}_*.txt")
    if not os.path.isdir(folder):
        print("find_follow_up_result: folder does not exist")
        return (False, False, False, False, False)

    for root, _, files in os.walk(folder):
        for file in files:
            if file.startswith(f"{follow_up_type}_{phone_number}_"):
                filepath = os.path.join(root, file)
                print(f"find_follow_up_result: found follow-up file {filepath}")
                result_tuple = analyze_follow_up_file(filepath, follow_up_type)
                print(f"find_follow_up_result: analysis tuple {result_tuple}")
                return result_tuple
    return (False, False, False, False, False)


# =====================================================
# ГЛАВНЫЕ ФУНКЦИИ ДЛЯ ОТЧЁТА
# =====================================================

def gather_station_files(station_code, base_folder):
    """
    Собирает файлы для станции, поддерживая оба формата имен файлов:
    - fs_[phone_number]_[station_code]_[datetime]_... (станция на позиции 2)
    - fs_[station_code]_[phone_number]_[datetime]_... (станция на позиции 1)
    
    Также собирает файлы с подстанций, которые привязаны к данной основной станции через STATION_MAPPING.
    """
    station_files = []
    
    # Получаем список всех кодов станций, которые должны быть включены в отчет для данной основной станции
    # Включаем саму основную станцию и все её подстанции
    target_station_codes = {station_code}
    
    # Добавляем подстанции, которые привязаны к данной основной станции
    if station_code in config.STATION_MAPPING:
        target_station_codes.update(config.STATION_MAPPING[station_code])
        print(f"gather_station_files: для основной станции {station_code} добавлены подстанции: {config.STATION_MAPPING[station_code]}")
    
    # Сканируем папку, пропуская подпапки с анализом follow-up (*_analysis)
    for root, dirs, files in os.walk(base_folder):
        # Удаляем из обхода все подпапки, заканчивающиеся на '_analysis'
        dirs[:] = [d for d in dirs if not d.endswith('_analysis')]
        for file in files:
            if not file.endswith('.txt'):
                continue
                
            file_parts = file.split("_")
            if len(file_parts) < 3:
                continue
                
            # Проверяем оба возможных положения кода станции
            first_part = file_parts[1]
            second_part = file_parts[2]
            
            # Определяем, где находится код станции и проверяем, входит ли он в целевые коды
            file_station_code = None
            if first_part in target_station_codes:
                file_station_code = first_part
            elif second_part in target_station_codes:
                file_station_code = second_part
            
            if file_station_code:
                station_files.append(os.path.join(root, file))
                if file_station_code != station_code:
                    print(f"Файл с подстанции {file_station_code} включен в отчет основной станции {station_code}: {file}")
                else:
                    print(f"Файл найден для станции {station_code}: {file}")
                
    print(f"gather_station_files: итого файлов для станции {station_code} (включая подстанции): {len(station_files)}")
    return station_files


def process_station_data(station_code, files, base_folder):
    """
    Обработка данных для станции.

    Логика:
      - Если в исходном файле найден один из тегов: [РЕЗУЛЬТАТ:ЗАПИСЬ], [РЕЗУЛЬТАТ:НЕТ],
        [РЕЗУЛЬТАТ:МЕССЕНДЖЕР], [РЕЗУЛЬТАТ:ОТКАЗ] – они учитываются напрямую.
      - Если найден тег [РЕЗУЛЬТАТ:ПЕРЕЗВОНИТЬ] (trigger recall) или [РЕЗУЛЬТАТ:ПЕРЕВОД] (trigger transfer),
        то по номеру телефона ищется follow-up файл.
          • Если файл найден, то outcome определяется по его содержимому:
              - Если найден тег "запись" – outcome = "record"
              - Если найден тег "нет" – outcome = "no"
              - Если найден тег "связались" – outcome = "connected_recall" (для recall) или "connected_transfer" (для transfer)
              - Если найден тег "мессенджер" – outcome = "messenger"
              - Если найден тег "отказ" – outcome = "refusal"
          • Если файл не найден, outcome остаётся равным "recall" или "transfer" (звонок считается потерянным).
      - Затем по итоговому outcome обновляются счётчики:
         Для группы A (исходно trigger recall):
           • "record"      → class_a_record
           • "no"          → class_a_no
           • "messenger"   → class_a_messenger
           • "refusal"     → class_a_refusal
           • "connected_recall" → class_a_connected_recall
           • "connected_transfer" → class_a_connected_transfer
           • Если outcome осталось "recall" (follow-up не найден) → class_a_recall
         Для группы B (исходно trigger transfer):
           • "record"      → class_b_record
           • "no"          → class_b_no
           • "messenger"   → class_b_messenger
           • "refusal"     → class_b_refusal
           • "connected_recall" → class_b_connected_recall
           • "connected_transfer" → class_b_connected_transfer
           • Если outcome осталось "transfer" (follow-up не найден) → class_b_transfer
    """
    # Инициализируем счетчики для группы A и группы B
    class_a_record = class_a_no = class_a_recall = class_a_messenger = class_a_refusal = 0
    class_a_connected_recall = class_a_connected_transfer = 0
    class_b_record = class_b_no = class_b_transfer = class_b_messenger = class_b_refusal = 0
    class_b_connected_recall = class_b_connected_transfer = 0

    for file_path in files:
        filename = os.path.basename(file_path)
        (is_target, is_class_a, is_class_b, r_record, r_no, r_recall, r_transfer, r_messenger, r_refusal) = analyze_file(file_path)
        # Требуем наличия целевого тега, тега класса и любого валидного тега результата
        if (not is_target) or (not (is_class_a or is_class_b)) or (not any([r_record, r_no, r_recall, r_transfer, r_messenger, r_refusal])):
            print(f"process_station_data: SKIP отсутствует обязательный тег (target/class/result) -> {os.path.basename(file_path)}")
            continue

        # Обрабатываем только целевые звонки
        if not is_target:
            continue

        # Определяем исходный outcome по приоритету:
        # Если обнаружен record, то outcome = record, далее no, messenger, refusal.
        outcome = None
        # Приоритет определения outcome:
        # 1) ЗАПИСЬ (record) 
        # 2) МЕССЕНДЖЕР (messenger) – клиент ушёл в мессенджер
        # 3) ОТКАЗ (refusal) – прямой отказ
        # 4) ПЕРЕЗВОНИТЬ / ПЕРЕВОД (recall / transfer) – запланирован follow-up
        # 5) НЕТ (no) – нет записи и не запланирован follow-up
        if r_record:
            outcome = "record"
        elif r_messenger:
            outcome = "messenger"
        elif r_refusal:
            outcome = "refusal"
        elif r_recall:
            outcome = "recall"  # trigger recall (группа A)
        elif r_transfer:
            outcome = "transfer"  # trigger transfer (группа B)
        elif r_no:
            outcome = "no"
        else:
            continue  # Если ни один тег не найден, пропускаем

        # Если trigger равен recall или transfer – ищем follow-up файл
        if outcome in ("recall", "transfer"):
            phone_number = get_phone_number_from_filename(filename)
            if phone_number:
                follow_type = outcome  # recall или transfer
                followup = find_follow_up_result(phone_number, base_folder, follow_type)
                # Если найден хотя бы один тег follow-up, определяем итоговое outcome:
                if any(followup):
                    if followup[0]:
                        outcome = "record"
                    elif followup[1]:
                        outcome = "no"
                    elif followup[2]:
                        # Различаем: если исходный trigger recall, то connected_recall;
                        # если trigger transfer, то connected_transfer.
                        outcome = "connected_recall" if follow_type == "recall" else "connected_transfer"
                    elif followup[3]:
                        outcome = "messenger"
                    elif followup[4]:
                        outcome = "refusal"
                # Если ни один тег follow-up найден, outcome остается "recall" или "transfer"

        # Подробный вывод выбранного исхода перед обновлением счётчиков
        print(f"process_station_data: {filename} => class_a={is_class_a} class_b={is_class_b} outcome={outcome}")
        # Обновляем счетчики по итоговому outcome для группы A или группы B
        if is_class_a:
            if outcome == "record":
                class_a_record += 1
            elif outcome == "no":
                class_a_no += 1
            elif outcome == "messenger":
                class_a_messenger += 1
            elif outcome == "refusal":
                class_a_refusal += 1
            elif outcome == "connected_recall":
                class_a_connected_recall += 1
            elif outcome == "connected_transfer":
                class_a_connected_transfer += 1
            elif outcome == "recall":  # follow-up не найден, считается потерянным
                class_a_recall += 1
        if is_class_b:
            if outcome == "record":
                class_b_record += 1
            elif outcome == "no":
                class_b_no += 1
            elif outcome == "messenger":
                class_b_messenger += 1
            elif outcome == "refusal":
                class_b_refusal += 1
            elif outcome == "connected_recall":
                class_b_connected_recall += 1
            elif outcome == "connected_transfer":
                class_b_connected_transfer += 1
            elif outcome == "transfer":  # follow-up не найден
                class_b_transfer += 1

        # Снимок текущих счётчиков после обработки файла
        print(f"process_station_data: counters snapshot after {filename}: "
              f"A(rec={class_a_record}, no={class_a_no}, recall={class_a_recall}, mess={class_a_messenger}, ref={class_a_refusal}, "
              f"conn_rec={class_a_connected_recall}, conn_trans={class_a_connected_transfer}) "
              f"B(rec={class_b_record}, no={class_b_no}, transfer={class_b_transfer}, mess={class_b_messenger}, ref={class_b_refusal}, "
              f"conn_rec={class_b_connected_recall}, conn_trans={class_b_connected_transfer})")

    return {
        'station_code': station_code,
        'class_a_record': class_a_record,
        'class_a_no': class_a_no,
        'class_a_recall': class_a_recall,
        'class_a_messenger': class_a_messenger,
        'class_a_refusal': class_a_refusal,
        'class_a_connected_recall': class_a_connected_recall,
        'class_a_connected_transfer': class_a_connected_transfer,
        'class_b_record': class_b_record,
        'class_b_no': class_b_no,
        'class_b_transfer': class_b_transfer,
        'class_b_messenger': class_b_messenger,
        'class_b_refusal': class_b_refusal,
        'class_b_connected_recall': class_b_connected_recall,
        'class_b_connected_transfer': class_b_connected_transfer
    }


# =====================================================
# СОЗДАНИЕ И ЗАПОЛНЕНИЕ EXCEL-ФАЙЛА
# =====================================================

def create_excel_file():
    print("create_excel_file: начало создания файла")
    output_dir = str(BASE_DIR / "conversion_bestway")
    os.makedirs(output_dir, exist_ok=True)
    print(f"create_excel_file: создана директория {output_dir}")
    excel_filename = f"Отчёт_Конверсия-Причины_Отказов_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    print(f"create_excel_file: имя файла {excel_filename}")
    return os.path.join(output_dir, excel_filename)


def make_summary_dataframe(station_results):
    print(f"make_summary_dataframe: формирую сводный отчёт для {len(station_results)} станций")
    rows = [
        'Клиент позвонил записаться и...',
        'Записаны',
        'Не записаны',
        'Нужен повторный звонок',
        'Переведен на другую станцию',
        'Перешли в мессенджер',
        'Клиент отказался от услуг',
        'Связались (перезвон)',
        'Связались (перевод)',
        '',
        'Клиент позвонил поинтересоваться и...',
        'Записаны',
        'Не записаны',
        'Нужен повторный звонок',
        'Переведен на другую станцию',
        'Перешли в мессенджер',
        'Клиент отказался от услуг',
        'Связались (перезвон)',
        'Связались (перевод)',
        '',
        'Звонков за день'
    ]

    df = pd.DataFrame(index=rows)

    for res in station_results:
        station_code = res['station_code']
        station_name = get_station_name(station_code)

        total_calls = sum([
            res['class_a_record'],
            res['class_a_no'],
            res['class_a_recall'],
            res['class_a_messenger'],
            res['class_a_refusal'],
            res['class_a_connected_recall'],
            res['class_a_connected_transfer'],
            res['class_b_record'],
            res['class_b_no'],
            res['class_b_transfer'],
            res['class_b_messenger'],
            res['class_b_refusal'],
            res['class_b_connected_recall'],
            res['class_b_connected_transfer']
        ])

        col_data = [
            '',  # Заголовок первой группы
            res['class_a_record'],
            res['class_a_no'],
            res['class_a_recall'],
            0,  # "Переведен на другую станцию" для группы A всегда 0
            res['class_a_messenger'],
            res['class_a_refusal'],
            res['class_a_connected_recall'],
            res['class_a_connected_transfer'],
            '',
            '',  # Клиент позвонил поинтересоваться и...
            res['class_b_record'],
            res['class_b_no'],
            res['class_b_transfer'],
            0,  # "Переведен на другую станцию" для группы B всегда 0
            res['class_b_messenger'],
            res['class_b_refusal'],
            res['class_b_connected_recall'],
            res['class_b_connected_transfer'],
            '',
            total_calls
        ]
        df[station_name] = col_data

    return df


def make_short_report_dataframe(station_results):
    print(f"make_short_report_dataframe: формирую краткий отчёт для {len(station_results)} станций")
    data = []
    for res in station_results:
        station_code = res['station_code']
        st_name = get_station_name(station_code)
        total_calls = sum([
            res['class_a_record'],
            res['class_a_no'],
            res['class_a_recall'],
            res['class_a_messenger'],
            res['class_a_refusal'],
            res['class_a_connected_recall'],
            res['class_a_connected_transfer'],
            res['class_b_record'],
            res['class_b_no'],
            res['class_b_transfer'],
            res['class_b_messenger'],
            res['class_b_refusal'],
            res['class_b_connected_recall'],
            res['class_b_connected_transfer']
        ])
        # "Записано" включает прямые записи и успешные follow-up
        total_record = (
            res['class_a_record'] +
            res['class_a_connected_recall'] +
            res['class_a_connected_transfer'] +
            res['class_b_record'] +
            res['class_b_connected_recall'] +
            res['class_b_connected_transfer']
        )
        # "Потеряно" включает все случаи, когда запись не состоялась
        # «Потеряно» не должно включать переходы в мессенджер – такие звонки считаем отдельно
        lost_calls = (
            res['class_a_no'] +
            res['class_a_recall'] +
            res['class_a_refusal'] +
            res['class_b_no'] +
            res['class_b_transfer'] +
            res['class_b_refusal']
        )
        data.append([st_name, total_calls, total_record, lost_calls])

    df = pd.DataFrame(data, columns=["Станция", "Всего звонков", "Записано", "Потеряно"])

    # Добавляем итоговую строку, суммируя данные по каждому числовому столбцу
    sum_calls = df["Всего звонков"].sum()
    sum_record = df["Записано"].sum()
    sum_lost = df["Потеряно"].sum()

    # Добавляем строку "Итог" в конец DataFrame
    df.loc[len(df.index)] = ["Итог", sum_calls, sum_record, sum_lost]

    return df


def style_summary_sheet(wb):
    print("style_summary_sheet: применяю стили к листу «Сводный отчёт»")
    sheetname = "Сводный отчёт"
    if sheetname not in wb.sheetnames:
        return
    ws = wb[sheetname]

    ws.column_dimensions['A'].width = 37
    for col_num in range(2, ws.max_column + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 16
        for cell in ws[col_letter]:
            cell.alignment = Alignment(wrap_text=True)

    thin_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    max_row = ws.max_row
    max_col = ws.max_column
    for row in range(2, max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val in ["Клиент позвонил записаться и...",
                   "Клиент позвонил поинтересоваться и...",
                   "Звонков за день"]:
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).font = Font(bold=True)
        else:
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).font = Font(bold=False)


def style_report_sheet(wb):
    print("style_report_sheet: применяю стили к листу «Отчёт»")
    sheetname = "Отчёт"
    if sheetname not in wb.sheetnames:
        return
    ws = wb[sheetname]

    # Установим ширину для первых 4 столбцов
    widths = [30, 15, 15, 15]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Зададим жирный шрифт для шапки (первая строка)
    header_row = 1
    max_row = ws.max_row
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True)

    # Создадим объект тонкой границы
    thin_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    # Применим тонкую границу ко всем НЕ пустым ячейкам
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border

    # Для столбцов B, C, D настроим выравнивание по центру
    for row in range(1, max_row + 1):
        for col in [2, 3, 4]:  # B=2, C=3, D=4
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.alignment = Alignment(horizontal='center')

    # Жирным шрифтом выделяем строку «Итог» и строку «Перешли в мессенджер»
    for row in range(2, max_row + 1):  # начиная со второй строки, чтобы не трогать заголовок
        first_col_val = ws.cell(row=row, column=1).value
        reason_val = ws.cell(row=row, column=2).value  # во второй колонке хранится причина отказа
        if first_col_val == "Итог" or (isinstance(reason_val, str) and reason_val.lower() == "перешли в мессенджер"):
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).font = Font(bold=True)


def send_excel_report_to_telegram(excel_file):
    print(f"send_excel_report_to_telegram: отправка файла {excel_file}")
    if not ensure_telegram_ready('отправка отчёта skolko_52'):
        return
    token = config.TELEGRAM_BOT_TOKEN
    chat_ids = [chat_id for chat_id in [config.ALERT_CHAT_ID] if chat_id]
    if not token or not chat_ids:
        print('Telegram не настроен для skolko_52, сообщение пропущено.')
        return
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    if not os.path.exists(excel_file):
        print(f'Файл {excel_file} не найден.')
        return
    for chat_id in chat_ids:
        print(f"send_excel_report_to_telegram: отправка в чат {chat_id}")
        with open(excel_file, 'rb') as file:
            response = requests.post(url, data={'chat_id': chat_id}, files={'document': file})
            if response.status_code == 200:
                print(f'Отчёт успешно отправлен в чат {chat_id}.')
            else:
                print(f'Ошибка при отправке в чат {chat_id}: {response.status_code}, {response.text}')
    return


# =====================================================
# ОСНОВНАЯ ФУНКЦИЯ
# =====================================================

def run_skolko_52():
    print("run_skolko_52: начало обработки")
    base_folder = str(config.BASE_RECORDS_PATH / f"{datetime.now():%Y/%m/%d}" / "transcriptions")
    station_codes = [
        '9322', '9300', '9321', '9326', '9347', '9327', '9325', '9324',
        '9304', '9308', '9301', '9302', '9307', '9316', '9319', '9350', '9344'
    ]

    excel_file = create_excel_file()
    print(f"run_skolko_52: файл Excel = {excel_file}")
    station_results = []
    # ---------- Сбор данных для сводного отчёта (без лишних логов) ----------
    verbose_summary = True  # поставить True, чтобы видеть логи при генерации сводного отчёта
    if verbose_summary:
        for code in station_codes:
            print(f"run_skolko_52: обрабатываю станцию {code}")
            files = gather_station_files(code, base_folder)
            # Обрабатываем станцию даже если файлов нет — получим нулевые показатели
            res = process_station_data(code, files, base_folder)
            print(f"run_skolko_52: результат для станции {code}: {res}")
            station_results.append(res)
    else:
        with contextlib.redirect_stdout(io.StringIO()):
            for code in station_codes:
                print(f"run_skolko_52: обрабатываю станцию {code}")
                files = gather_station_files(code, base_folder)
                res = process_station_data(code, files, base_folder)
                print(f"run_skolko_52: результат для станции {code}: {res}")
                station_results.append(res)

    # ---------- Анализ причин по скрипту whynot (parse_day) ----------
    today = datetime.now().date()
    verbose_reasons = False  # поставить True, чтобы видеть логи при анализе причин
    if verbose_reasons:
        reasons_data = parse_day(today)
    else:
        with contextlib.redirect_stdout(io.StringIO()):
            reasons_data = parse_day(today)
    # Добавляем станции без данных, чтобы они отобразились в отчёте с нулями
    for code in station_codes:
        reasons_data.setdefault(code, {"target": 0, "signup": 0, "reasons": Counter()})
    print(f"run_skolko_52: коды станций в reasons_data = {list(reasons_data.keys())}")
    date_label = today.strftime('%d.%m.%Y')
    stations = sorted(reasons_data)
    # собираем уникальные причины
    reasons_set = set()
    for inf in reasons_data.values():
        reasons_set.update(inf['reasons'].keys())
    sorted_reasons = sorted(reasons_set,
                            key=lambda r: (-sum(reasons_data[st]['reasons'].get(r, 0) for st in stations), r))
    # формируем DataFrame с анализом причин
    reasons_rows = []
    for reason in sorted_reasons:
        row = {'Причина': reason, 'Всего': sum(reasons_data[st]['reasons'].get(reason, 0) for st in stations)}
        for st in stations:
            row[st] = reasons_data[st]['reasons'].get(reason, 0)
        reasons_rows.append(row)
    reasons_df = pd.DataFrame(reasons_rows)

    # ---------- Формируем данные для листа «Сводный отчёт» без логов ----------
    with contextlib.redirect_stdout(io.StringIO()):
        df_summary = make_summary_dataframe(station_results)  # Лист "Сводный отчёт"

    wb = Workbook()
    build_detailed_report_sheet(wb, date_label, reasons_data)
    ws2 = wb.create_sheet("Сводный отчёт")
    # Запись сводного отчёта вручную
    df_s = df_summary.reset_index()
    for c_idx, col in enumerate(df_s.columns, start=1): ws2.cell(row=1, column=c_idx, value=col)
    for r_idx, row in enumerate(df_s.values, start=2):
        for c_idx, val in enumerate(row, start=1): ws2.cell(row=r_idx, column=c_idx, value=val)
    with contextlib.redirect_stdout(io.StringIO()):
        style_summary_sheet(wb)
    wb.save(excel_file)
    print(f"run_skolko_52: сохранён файл {excel_file}")
    send_excel_report_to_telegram(excel_file)
    print("run_skolko_52: отправка в Телеграм завершена")
    print("Скрипт run_skolko_52 завершён успешно.")
    print("run_skolko_52: конец обработки")


# Для запуска скрипта можно раскомментировать ниже:
if __name__ == "__main__":
    run_skolko_52()
