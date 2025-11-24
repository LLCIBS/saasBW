import os
import re
import json
import requests
import pandas as pd
import openpyxl

# Отключаем FutureWarning pandas о downcasting
pd.set_option('future.no_silent_downcasting', True)

from datetime import datetime, timedelta
import time
from openpyxl.styles import Font, Border, Side, Alignment
import shutil
import yaml
try:
    from call_analyzer.utils import ensure_telegram_ready  # type: ignore
except ImportError:
    from utils import ensure_telegram_ready

# Импортируем конфигурацию из основного модуля
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
import config as main_config
try:
    from call_analyzer.utils import parse_filename  # type: ignore
except ImportError:
    from utils import parse_filename
import config as cfg
# Импортируем функции для извлечения имени оператора
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
try:
    from exental_alert import get_operator_name, extract_dialog_from_txt
except ImportError:
    # Fallback если не удалось импортировать
    def get_operator_name(dialog_text=None, station_code=None):
        if station_code:
            employee_full = main_config.EMPLOYEE_BY_EXTENSION.get(station_code)
            if employee_full:
                return employee_full.split()[0] if employee_full else 'Не указано'
        return 'Не указано'
    def extract_dialog_from_txt(txt_path):
        return ""

# Настройки для Telegram
telegram_bot_token = main_config.TELEGRAM_BOT_TOKEN
# Настройки API
thebai_url = main_config.THEBAI_URL
thebai_api_key = main_config.THEBAI_API_KEY


def load_prompt(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        data = yaml.safe_load(f)
        return data.get('prompt', '')

def get_num_questions_from_yaml(script_prompt_path: str = None) -> int:
    try:
        prompt_path = script_prompt_path or str(main_config.SCRIPT_PROMPT_8_PATH)
        with open(prompt_path, 'r', encoding='utf-8') as f:
            data = yaml.safe_load(f) or {}
            checklist = data.get('checklist') or []
            if isinstance(checklist, list) and len(checklist) > 0:
                return len(checklist)
    except Exception:
        pass
    return 8

def get_checklist_titles_from_yaml() -> list:
    """Возвращает список строк вида "N. <Название пункта>" из YAML чек-листа.
    Если чек-лист пуст/ошибка — вернёт 8 дефолтных пунктов.
    """
    try:
        with open(str(main_config.SCRIPT_PROMPT_8_PATH), 'r', encoding='utf-8') as f:
            data = yaml.safe_load(f) or {}
            checklist = data.get('checklist') or []
            titles = []
            for idx, item in enumerate(checklist, start=1):
                title = str((item or {}).get('title', '')).strip()
                if title:
                    titles.append(f"{idx}. {title}")
            if titles:
                return titles
    except Exception:
        pass
    # fallback на 8 старых вопросов, если YAML пуст
    return [
        "1. Сотрудник поприветствовал клиента, представился по имени, назвал отдел?",
        "2. Сотрудник спросил имя клиента?",
        "3. Имя клиента звучало в диалоге не менее 2 раз? ",
        "4. Сотрудник предложил клиенту конкретное решение или следующий шаг?",
        "5. Записан ли клиент на СТО в результате звонка?",
        "6. Сотрудник уточнил, не осталось ли у Клиента вопросов?",
        "7. Сотрудник резюмировал результаты переговоров с клиентом или зафиксировал достигнутые договоренности? ",
        "8. Сотрудник поблагодарил клиента за звонок и попрощался? "
    ]

# Функция для отправки текста на анализ
def analyze_content(transcript):
    print("Отправка текста на анализ в TheBai...")
    # Путь к YAML-файлу с prompt из конфига
    prompt_file_path = str(main_config.SCRIPT_PROMPT_8_PATH)

    # Загрузка prompt из YAML-файла
    prompt_template = load_prompt(prompt_file_path)

    if not prompt_template:
        print("Промпт не найден или пуст в YAML-файле.")
        return None

    # Формирование полного prompt с подстановкой транскрипта
    prompt = f"{prompt_template}\n\nВот диалог:\n{transcript}"

    payload = json.dumps({
        "model": main_config.THEBAI_MODEL,
        "messages": [
            {
                "role": "user",
                "content": prompt
            }
        ],
        "stream": False
    })

    headers = {
        'Authorization': f'Bearer {thebai_api_key}',
        'Content-Type': 'application/json'
    }

    try:
        response = requests.post(thebai_url, headers=headers, data=payload)
        time.sleep(30)  # Задержка 30 секунд между запросами
        if response.status_code == 200:
            response_data = response.json()
            print("Успешный анализ, результат получен.")
            return response_data['choices'][0]['message']['content']
        else:
            print(f'Ошибка API TheBai: {response.status_code}, {response.text}')
            return None
    except Exception as e:
        print(f"Ошибка при подключении к API: {e}")
        return None

def get_last_week_dates():
    # Устанавливаем значения по умолчанию
    offset_days = 6
    days = 6

    # Проверяем наличие файла с настройками
    settings_file = "days.txt"
    if os.path.exists(settings_file):
        with open(settings_file, "r") as file:
            for line in file:
                if line.startswith("offset_days="):
                    try:
                        offset_days = int(line.split("=")[1].strip())
                    except ValueError:
                        print("Некорректное значение offset_days в файле. Используется значение по умолчанию: 6.")
                elif line.startswith("days="):
                    try:
                        days = int(line.split("=")[1].strip())
                    except ValueError:
                        print("Некорректное значение days в файле. Используется значение по умолчанию: 6.")

    today = datetime.today()
    start_date = today - timedelta(days=offset_days)
    end_date = start_date + timedelta(days=days)
    return start_date, end_date, days

def generate_report_name_and_message(period_start: datetime, period_end: datetime):
    report_name = f"Отчет_по_скрипту_{period_start.strftime('%d')}.{period_start.strftime('%m')}-{period_end.strftime('%d')}.{period_end.strftime('%m')}.xlsx"
    telegram_message = f"Отчет по скрипту за период с {period_start.strftime('%d.%m.%Y')} по {period_end.strftime('%d.%m.%Y')}."
    return report_name, telegram_message


def extract_transcript(content):
    dialog_start = content.find("Диалог:")
    analysis_start = content.find("Анализ:")
    if dialog_start == -1:
        print("Диалог не найден.")
        return None
    # Если "Диалог:" найден, но "Анализ:" нет
    if analysis_start == -1:
        # Возвращаем всё, что идёт после "Диалог:"
        return content[dialog_start + len("Диалог:"):].strip()
    # Если и "Диалог:", и "Анализ:" найдены
    return content[dialog_start + len("Диалог:"):analysis_start].strip()


def get_daily_transcriptions_folder(base_folder, period_start: datetime, period_end: datetime):
    today_folder = datetime.today().strftime('%Y/%m/%d')
    week_folder_name = f"{period_start.strftime('%d')}-{period_end.strftime('%d')}_script"
    daily_folder = os.path.join(base_folder, today_folder, 'transcriptions', week_folder_name)
    os.makedirs(daily_folder, exist_ok=True)
    return daily_folder

def check_and_copy_existing_analysis(file_path, month_folder, target_folder):
    base_filename = os.path.basename(file_path)
    analysis_filename = f"{base_filename[:-4]}_analysis.txt"

    for root, dirs, files in os.walk(month_folder):
        print(f"Проверка директории: {root}")
        if analysis_filename in files:
            existing_analysis_path = os.path.join(root, analysis_filename)
            print(f"Найден существующий анализ: {existing_analysis_path}")

            target_path = os.path.join(target_folder, analysis_filename)
            shutil.copy2(existing_analysis_path, target_path)
            print(f"Скопирован анализ в: {target_path}")
            return True

    print(f"Анализ {analysis_filename} не найден в месячной папке.")
    return False

def parse_tg_bw_calls(file_path):
    call_records = []
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            match = re.match(r'^(?:[^-]+-){2}(\+?\d+)-(.+?)-(.+?)-(\d{2}-\d{2}-\d{4}) (\d{2}-\d{2})\.mp3$', line)
            if match:
                phone_number = match.group(1).lstrip('+')
                station_code = str(match.group(2)).strip()
                consultant_full_name = str(match.group(3)).strip()
                date_str = match.group(4)
                time_str = match.group(5)
                date_time_str = f"{date_str} {time_str.replace('-', ':')}"
                try:
                    date_time_obj = datetime.strptime(date_time_str, '%d-%m-%Y %H:%M')
                    consultant_surname = consultant_full_name.split()[0]
                    call_records.append({
                        'phone_number': phone_number,
                        'station_code': station_code,
                        'consultant_surname': consultant_surname,
                        'datetime': date_time_obj
                    })
                except ValueError:
                    print(f"Некорректный формат даты и времени в строке: {line}")
            else:
                print(f"Строка не соответствует шаблону: {line}")
    return call_records

def analyze_files(period_start: datetime, period_end: datetime, base_folder=None):
    print(f"Начало анализа файлов за выбранный период: {period_start.strftime('%d.%m.%Y')} - {period_end.strftime('%d.%m.%Y')}...")

    # Используем переданный base_folder или берём из конфига
    folder_path = str(base_folder) if base_folder else str(main_config.BASE_RECORDS_PATH)
    print(f"DEBUG: Используем BASE_RECORDS_PATH: {folder_path}")
    transcriptions_folder = get_daily_transcriptions_folder(folder_path, period_start, period_end)
    # Чистим целевую папку отчета, чтобы не смешивались файлы от прошлых запусков
    try:
        for entry in os.listdir(transcriptions_folder):
            entry_path = os.path.join(transcriptions_folder, entry)
            if os.path.isfile(entry_path) and (entry.endswith('.txt') or entry.endswith('.xlsx')):
                os.remove(entry_path)
    except Exception as e:
        print(f"Не удалось очистить папку отчета {transcriptions_folder}: {e}")
    report_name, telegram_message = generate_report_name_and_message(period_start, period_end)

    # tg_bw_calls теперь формируем в папке текущего отчета периода
    tg_bw_calls_file = os.path.join(transcriptions_folder, 'tg_bw_calls.txt')
    # Всегда генерируем tg_bw_calls.txt заново для выбранного периода
    print(f"Генерация tg_bw_calls.txt за выбранный интервал: {period_start} - {period_end}")
    generate_tg_bw_calls_for_period(period_start, period_end, tg_bw_calls_file, base_folder=folder_path)
    if not os.path.exists(tg_bw_calls_file):
        print(f"Не удалось создать {tg_bw_calls_file}. Завершение работы.")
        return None
    call_records = parse_tg_bw_calls(tg_bw_calls_file)
    print(f"Найдено записей в tg_bw_calls.txt: {len(call_records)}")

    # Ограничиваем записи выбранным периодом
    call_records = [r for r in call_records if period_start <= r['datetime'] <= period_end + timedelta(days=1)]

    # Список папок по датам в выбранном диапазоне
    num_days = (period_end.date() - period_start.date()).days
    last_week_dates = [(period_start + timedelta(days=i)).strftime('%Y/%m/%d') for i in range(num_days + 1)]

    failed_files = []

    # Копируем готовые анализы из script_8 вместо повторного анализа
    for date_str in last_week_dates:
        transcriptions_base = os.path.join(folder_path, date_str, 'transcriptions')
        
        # Ищем все подпапки с анализами (приоритет script_8)
        script_folders = []
        if os.path.exists(transcriptions_base):
            # Сначала ищем script_8
            script_8_path_candidate = os.path.join(transcriptions_base, 'script_8')
            if os.path.exists(script_8_path_candidate) and os.path.isdir(script_8_path_candidate):
                script_folders.append(script_8_path_candidate)
            # Затем ищем другие папки со script в названии
            try:
                for item in os.listdir(transcriptions_base):
                    item_path = os.path.join(transcriptions_base, item)
                    if os.path.isdir(item_path) and 'script' in item.lower() and item != 'script_8':
                        script_folders.append(item_path)
            except (OSError, PermissionError) as e:
                print(f"Ошибка при чтении директории {transcriptions_base}: {e}")
        
        if not script_folders:
            print(f"Папки с анализами не найдены для {date_str}, пропуск.")
            continue
        
        # Используем первую найденную папку (обычно script_8)
        script_8_path = script_folders[0]
        print(f"Проверка готовых анализов в {script_8_path}")
        try:
            files_in_script_8 = os.listdir(script_8_path)
            print(f"DEBUG: Files in script_8_path: {files_in_script_8}")
            # Фильтруем только файлы анализов
            analysis_files = [f for f in files_in_script_8 if f.endswith('_analysis.txt')]
            print(f"DEBUG: Analysis files found: {len(analysis_files)} (total files: {len(files_in_script_8)})")
        except (OSError, PermissionError) as e:
            print(f"Ошибка при чтении директории {script_8_path}: {e}")
            continue

        # Подготовим словарь для быстрого фильтра по tg_bw_calls
        call_records_dict = {}
        for r in call_records:
            key = (r['phone_number'], r['datetime'].strftime('%Y-%m-%d %H:%M'))
            call_records_dict[key] = True
        
        print(f"DEBUG: call_records count: {len(call_records)}")
        print(f"DEBUG: call_records_dict keys: {list(call_records_dict.keys())}")

        # Используем список analysis_files, который мы уже отфильтровали выше
        for file in analysis_files:
            print(f"DEBUG: Processing analysis file: {file}")
            file_path = os.path.join(script_8_path, file)
            output_path = os.path.join(transcriptions_folder, file)
            
            # Предварительная фильтрация: копируем только те файлы, которые есть в tg_bw_calls
            base_name = file.replace('_analysis.txt', '')
            pre_phone = None
            pre_dt = None
            
            # Парсим имя файла для проверки соответствия tg_bw_calls
            match_fs = re.match(r'fs_(\+?\d+)_\d{3,4}_(\d{4}-\d{2}-\d{2})-(\d{2}-\d{2}-\d{2})', base_name)
            if match_fs:
                pre_phone = match_fs.group(1).lstrip('+')
                pre_dt_str = f"{match_fs.group(2)} {match_fs.group(3).replace('-', ':')}"
                try:
                    pre_dt = datetime.strptime(pre_dt_str, '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    pre_dt = None
            elif base_name.lower().startswith('external-'):
                try:
                    parts = base_name.split('-')
                    pre_phone = parts[2].lstrip('+')
                    yyyymmdd = parts[3]
                    hhmmss = parts[4]
                    pre_dt = datetime.strptime(f"{yyyymmdd} {hhmmss}", '%Y%m%d %H%M%S')
                except Exception:
                    pre_dt = None

            if pre_dt and pre_phone:
                key_exact = (pre_phone, pre_dt.strftime('%Y-%m-%d %H:%M'))
                in_calls = key_exact in call_records_dict
                print(f"DEBUG: Checking file {file}: phone={pre_phone}, dt={pre_dt}, key={key_exact}, exact_match={in_calls}")
                if not in_calls:
                    for minutes_diff in range(-5, 6):
                        adjusted_time = pre_dt + timedelta(minutes=minutes_diff)
                        key = (pre_phone, adjusted_time.strftime('%Y-%m-%d %H:%M'))
                        if key in call_records_dict:
                            in_calls = True
                            print(f"DEBUG: Found match with {minutes_diff} minutes offset")
                            break
                if not in_calls:
                    # Этот файл не входит в tg_bw_calls выбранного периода — пропускаем
                    print(f"DEBUG: File {file} not in call_records, skipping")
                    continue
                print(f"DEBUG: File {file} matches, will copy")

            # Копируем готовый анализ
            try:
                import shutil
                shutil.copy2(file_path, output_path)
                print(f"Скопирован готовый анализ: {file}")
            except Exception as e:
                print(f"Ошибка копирования файла {file}: {e}")

    print("Копирование готовых анализов завершено.")

    output_file_path = os.path.join(transcriptions_folder, report_name)
    create_excel_report(transcriptions_folder, output_file_path, telegram_message, call_records)
    send_report_to_telegram(output_file_path, telegram_message)
    print(f"Отчет успешно создан и отправлен в {output_file_path}")
    return output_file_path

def generate_tg_bw_calls_for_period(period_start: datetime, period_end: datetime, output_path: str, base_folder=None):
    """
    Строит tg_bw_calls.txt из имен файлов звонков за указанный интервал.
    Формат строки соответствует парсеру parse_tg_bw_calls: первые 2 блока произвольные.
    Консультант неизвестен — подставляется 'Не указано'.
    """
    base_dir = str(base_folder) if base_folder else str(main_config.BASE_RECORDS_PATH)
    print(f"DEBUG: generate_tg_bw_calls использует base_dir: {base_dir}")
    lines = []
    print(f"Генерация tg_bw_calls.txt за период {period_start} - {period_end} в {output_path}")
    day_count = (period_end.date() - period_start.date()).days
    for i in range(day_count + 1):
        day = (period_start + timedelta(days=i))
        day_folder = os.path.join(base_dir, day.strftime('%Y/%m/%d'))
        print(f"DEBUG: Проверяем папку: {day_folder}")
        if not os.path.exists(day_folder):
            print(f"DEBUG: Папка не существует: {day_folder}")
            continue
        print(f"DEBUG: Папка существует, содержимое:")
        try:
            all_files = os.listdir(day_folder)
            print(f"DEBUG: Все файлы в папке: {all_files}")
            audio_files = [f for f in all_files if any(f.lower().endswith(ext) for ext in main_config.FILENAME_PATTERNS['supported_extensions'])]
            print(f"DEBUG: Аудио файлы: {audio_files}")
        except Exception as e:
            print(f"DEBUG: Ошибка при чтении папки: {e}")
        
        try:
            found_for_day = 0
            for root, _, files in os.walk(day_folder):
                print(f"DEBUG: os.walk - root={root}, files={files}")
                if root == day_folder:
                    preview = [f for f in files if any(f.lower().endswith(ext) for ext in main_config.FILENAME_PATTERNS['supported_extensions'])][:10]
                    if preview:
                        print(f"Файлы аудио в {day_folder} (первые 10): {preview}")
                for entry in files:
                    name_lower = entry.lower()
                    if any(name_lower.endswith(ext) for ext in main_config.FILENAME_PATTERNS['supported_extensions']):
                        phone, station, call_time = parse_filename(entry)
                        if not phone or not station or not call_time:
                            print(f"DEBUG: Не удалось распарсить файл: {entry}")
                            print(f"  phone={phone}, station={station}, call_time={call_time}")
                            continue
                        if not (period_start <= call_time <= period_end + timedelta(days=1)):
                            print(f"DEBUG: Файл {entry} не попадает в период")
                            print(f"  call_time={call_time}, period_start={period_start}, period_end={period_end}")
                            continue
                        ddmmYYYY = call_time.strftime('%d-%m-%Y')
                        hhmm = call_time.strftime('%H-%M')
                        phone_out = phone if phone.startswith('+') else f"+{phone}"
                        employee = main_config.EMPLOYEE_BY_EXTENSION.get(station, 'Не указано')
                        line = f"fs-bw-{phone_out}-{station}-{employee}-{ddmmYYYY} {hhmm}.mp3"
                        lines.append(line)
                        found_for_day += 1
                        print(f"DEBUG: Добавлен файл: {entry} -> {line}")

            # Дополнительно: берем из подкаталога transcript имена .txt
            transcript_dir = os.path.join(day_folder, 'transcript')
            if os.path.exists(transcript_dir):
                try:
                    preview_txt = [f for f in os.listdir(transcript_dir) if f.lower().endswith('.txt')][:10]
                    if preview_txt:
                        print(f"Файлы transcript в {transcript_dir} (первые 10): {preview_txt}")
                except Exception:
                    pass
                for entry in os.listdir(transcript_dir):
                    name_lower = entry.lower()
                    if name_lower.endswith('.txt'):
                        # Убираем .txt для парсинга
                        base_name = entry[:-4]
                        phone, station, call_time = parse_filename(base_name)
                        if not phone or not station or not call_time:
                            continue
                        if not (period_start <= call_time <= period_end + timedelta(days=1)):
                            continue
                        ddmmYYYY = call_time.strftime('%d-%m-%Y')
                        hhmm = call_time.strftime('%H-%M')
                        phone_out = phone if phone.startswith('+') else f"+{phone}"
                        employee = main_config.EMPLOYEE_BY_EXTENSION.get(station, 'Не указано')
                        line = f"fs-bw-{phone_out}-{station}-{employee}-{ddmmYYYY} {hhmm}.mp3"
                        lines.append(line)
                        found_for_day += 1

            # Также ищем файлы анализа в подкаталогах transcriptions
            transcriptions_dir = os.path.join(day_folder, 'transcriptions')
            if os.path.exists(transcriptions_dir):
                # Ищем во всех подпапках, включая script_8 и другие
                for root, dirs, files in os.walk(transcriptions_dir):
                    print(f"DEBUG: Проверка папки анализов: {root}")
                    for entry in files:
                        if entry.endswith('_analysis.txt'):
                            # Убираем _analysis.txt для парсинга
                            base_name = entry[:-13]  # убираем '_analysis.txt'
                            phone, station, call_time = parse_filename(base_name)
                            if not phone or not station or not call_time:
                                print(f"DEBUG: Не удалось распарсить файл анализа: {entry}")
                                continue
                            if not (period_start <= call_time <= period_end + timedelta(days=1)):
                                print(f"DEBUG: Файл анализа {entry} не попадает в период")
                                continue
                            ddmmYYYY = call_time.strftime('%d-%m-%Y')
                            hhmm = call_time.strftime('%H-%M')
                            phone_out = phone if phone.startswith('+') else f"+{phone}"
                            employee = main_config.EMPLOYEE_BY_EXTENSION.get(station, 'Не указано')
                            line = f"fs-bw-{phone_out}-{station}-{employee}-{ddmmYYYY} {hhmm}.mp3"
                            lines.append(line)
                            found_for_day += 1
                            print(f"DEBUG: Добавлен файл анализа: {entry} -> {line}")
                        elif entry.endswith('.txt') and not entry.endswith('_analysis.txt'):
                            # Убираем .txt, т.к. парсер ожидает имя без расширения
                            base_name = entry[:-4]
                            phone, station, call_time = parse_filename(base_name)
                            if not phone or not station or not call_time:
                                continue
                            if not (period_start <= call_time <= period_end + timedelta(days=1)):
                                continue
                            ddmmYYYY = call_time.strftime('%d-%m-%Y')
                            hhmm = call_time.strftime('%H-%M')
                            phone_out = phone if phone.startswith('+') else f"+{phone}"
                            employee = main_config.EMPLOYEE_BY_EXTENSION.get(station, 'Не указано')
                            line = f"fs-bw-{phone_out}-{station}-{employee}-{ddmmYYYY} {hhmm}.mp3"
                            lines.append(line)
                            found_for_day += 1
            print(f"Найдено для {day.strftime('%Y-%m-%d')}: {found_for_day}")
        except Exception:
            continue
    if not lines:
        print("Файлы для генерации tg_bw_calls.txt не найдены в выбранном периоде.")
        return
    try:
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("\n".join(sorted(set(lines))))
        print(f"Сгенерирован {output_path}: {len(lines)} строк")
    except Exception as e:
        print(f"Не удалось записать {output_path}: {e}")

def replace_yes_no_in_excel(file_path):
    print(f"Открываем Excel-файл для замены 'ДА' на 1 и 'НЕТ' на 0: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == "ДА":
                    cell.value = 1
                elif cell.value == "НЕТ":
                    cell.value = 0
    wb.save(file_path)
    print(f"Все 'ДА' заменены на 1, а 'НЕТ' на 0 в файле {file_path}")

def send_report_to_telegram(file_path, message):
    """Отправка отчёта в Telegram."""
    if not ensure_telegram_ready('отправка отчёта week_full'):
        return
    chat_id = getattr(main_config, 'ALERT_CHAT_ID', None) or getattr(main_config, 'TG_CHANNEL_NIZH', None)
    if not chat_id:
        print('Telegram chat_id не задан, отправка отчёта пропущена')
        return
    token = getattr(main_config, 'TELEGRAM_BOT_TOKEN', '')
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    try:
        with open(file_path, 'rb') as file:
            files = {'document': file}
            data = {'chat_id': chat_id, 'caption': message}
            resp = requests.post(url, data=data, files=files)
        if resp.status_code == 200:
            print(f'Отчёт {file_path} успешно отправлен в чат {chat_id}')
        else:
            print(f'Ошибка при отправке отчёта в Telegram: {resp.text}')
    except Exception as e:
        print(f'Исключение при отправке отчёта: {e}')
    return
    url = f"https://api.telegram.org/bot{telegram_bot_token}/sendDocument"
    try:
        with open(file_path, "rb") as file:
            response = requests.post(url, data={"chat_id": chat_id, "caption": message}, files={"document": file})
        if response.status_code == 200:
            print(f"Отчет успешно отправлен в чат {chat_id}")
        else:
            print(f"Ошибка при отправке в Telegram: {response.text}")
    except Exception as e:
        print(f"Исключение при отправке в Telegram: {e}")

def get_station_groups(station_names=None, station_mapping=None, employee_map=None):
    station_names = station_names or getattr(main_config, 'STATION_NAMES', {})
    station_mapping = station_mapping or getattr(main_config, 'STATION_MAPPING', {})
    employee_map = employee_map or getattr(main_config, 'EMPLOYEE_BY_EXTENSION', {})
    station_groups = {}

    for station, name in station_names.items():
        station_groups[station] = name

    for main_station, substations in station_mapping.items():
        main_station_name = station_names.get(main_station)
        if main_station_name:
            for substation in substations:
                station_groups[substation] = main_station_name

    for station in employee_map.keys():
        if station not in station_groups:
            consultant = employee_map.get(station, "Не указано")
            station_groups[station] = f"Станция {consultant}"

    station_groups["Не указано"] = "Не указано"
    
    return station_groups


def create_excel_report(transcriptions_folder, output_file_path, telegram_message, call_records):
    if not os.path.exists(transcriptions_folder):
        os.makedirs(transcriptions_folder)

    # Получаем конфигурацию станций
    station_names = getattr(main_config, 'STATION_NAMES', {})
    station_mapping = getattr(main_config, 'STATION_MAPPING', {})
    employee_by_extension = getattr(main_config, 'EMPLOYEE_BY_EXTENSION', {})

    # Создаем словарь для быстрого поиска консультанта
    call_records_dict = {}
    for record in call_records:
        key = (record['phone_number'], record['datetime'].strftime('%Y-%m-%d %H:%M'))
        call_records_dict[key] = {
            'consultant_surname': record['consultant_surname'],
            'station_code': record['station_code']
        }

    data_for_excel = []
    for file in os.listdir(transcriptions_folder):
        if file.endswith("_analysis.txt"):
            file_path = os.path.join(transcriptions_folder, file)
            print(f"Извлечение данных из файла: {file_path}")

            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()

            # Извлекаем номер телефона и дату/время из имени файла
            base_name = file.replace('_analysis.txt', '')
            mp3_filename = base_name + '.mp3'


            consultant_surname = 'Не указано'
            station_code = 'Неизвестно'

            phone_number = None
            date_time_obj = None
            # Поддержка исходного формата fs_...
            match_fs = re.match(r'fs_(\+?\d+)_\d{3,4}_(\d{4}-\d{2}-\d{2})-(\d{2}-\d{2}-\d{2})', base_name)
            if match_fs:
                phone_number = match_fs.group(1).lstrip('+')
                date_str = match_fs.group(2)
                time_str = match_fs.group(3)
                date_time_str = f"{date_str} {time_str.replace('-', ':')}"
                try:
                    date_time_obj = datetime.strptime(date_time_str, '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    date_time_obj = None
            # Поддержка формата external-<station>-<phone>-<YYYYMMDD>-<HHMMSS>-...
            if not date_time_obj:
                if base_name.lower().startswith('external-'):
                    try:
                        parts = base_name.split('-')
                        station_code = parts[1]
                        ph = parts[2].lstrip('+')
                        yyyymmdd = parts[3]
                        hhmmss = parts[4]
                        date_time_obj = datetime.strptime(f"{yyyymmdd} {hhmmss}", '%Y%m%d %H%M%S')
                        phone_number = ph
                    except Exception:
                        date_time_obj = None
            # Поддержка формата вход_<station>_с_<phone>_на_<phone>_от_<YYYY>_<MM>_<DD>
            if not date_time_obj:
                if base_name.lower().startswith('вход_'):
                    try:
                        # Формат: вход_EkbFocusMal128801_с_79536098664_на_73432260822_от_2025_10_21
                        match_vhod = re.match(r'вход_([^_]+)_с_(\+?\d+)_на_\d+_от_(\d{4})_(\d{1,2})_(\d{1,2})', base_name, re.IGNORECASE)
                        if match_vhod:
                            station_code = match_vhod.group(1)
                            phone_number = match_vhod.group(2).lstrip('+')
                            year = match_vhod.group(3)
                            month = match_vhod.group(4)
                            day = match_vhod.group(5)
                            # Время устанавливаем в 00:00:00, так как в файлах анализа время не указано
                            date_time_obj = datetime.strptime(f"{year}-{month.zfill(2)}-{day.zfill(2)} 00:00:00", '%Y-%m-%d %H:%M:%S')
                    except Exception:
                        date_time_obj = None

            # Поддержка формата out-<phone>-<station>-<YYYYMMDD>-<HHMMSS>-...
            if not date_time_obj:
                if base_name.lower().startswith('out-'):
                    try:
                        parts = base_name.split('-')
                        # out-89196552973-203-20251120-173809-...
                        phone_number = parts[1].lstrip('+')
                        station_code = parts[2]
                        yyyymmdd = parts[3]
                        hhmmss = parts[4]
                        date_time_obj = datetime.strptime(f"{yyyymmdd} {hhmmss}", '%Y%m%d %H%M%S')
                    except Exception:
                        date_time_obj = None

            # Инициализируем значения по умолчанию
            consultant_surname = 'Не указано'
            station_code = 'Неизвестно'
            
            print(f"DEBUG: Обрабатываем файл {file}")
            print(f"DEBUG: base_name = {base_name}")
            print(f"DEBUG: phone_number = {phone_number}, date_time_obj = {date_time_obj}")
            
            # Сначала пытаемся получить station_code из call_records_dict
            if date_time_obj and phone_number:
                key_exact = (phone_number, date_time_obj.strftime('%Y-%m-%d %H:%M'))
                print(f"DEBUG: Ищем точное соответствие: {key_exact}")
                print(f"DEBUG: Доступные ключи в call_records_dict: {list(call_records_dict.keys())}")
                
                if key_exact in call_records_dict:
                    # consultant_surname из call_records_dict - это fallback
                    consultant_surname_fallback = call_records_dict[key_exact]['consultant_surname']
                    if consultant_surname_fallback and consultant_surname_fallback != 'Не указано':
                        consultant_surname = consultant_surname_fallback
                    
                    station_code = call_records_dict[key_exact]['station_code']
                    print(f"DEBUG: Найдено точное соответствие: consultant={consultant_surname_fallback}, station={station_code}")
                else:
                    # Поиск с допуском в 5 минут
                    matched = False
                    for minutes_diff in range(-5, 6):
                        adjusted_time = date_time_obj + timedelta(minutes=minutes_diff)
                        key = (phone_number, adjusted_time.strftime('%Y-%m-%d %H:%M'))
                        if key in call_records_dict:
                            consultant_surname_fallback = call_records_dict[key]['consultant_surname']
                            if consultant_surname_fallback and consultant_surname_fallback != 'Не указано':
                                consultant_surname = consultant_surname_fallback
                            
                            station_code = call_records_dict[key]['station_code']
                            matched = True
                            print(f"DEBUG: Найдено соответствие с допуском {minutes_diff} мин: consultant={consultant_surname_fallback}, station={station_code}")
                            break
                    if not matched:
                        print(f"DEBUG: Не найдено соответствия для файла {file}")
            else:
                print(f"DEBUG: Не удалось извлечь phone_number или date_time_obj из файла {file}")
            
            # Приоритет 1: Извлекаем имя оператора из транскрипции (диалога)
            # Диалог может быть в файле анализа или в отдельном txt файле
            dialog_text = None
            try:
                # Пытаемся извлечь диалог из файла анализа
                # Структура файла: "Диалог (из исходного TXT):\n\n{dialog_text}\n\nРаспознавание по чек-листу:\n\n"
                dialog_marker = "Диалог (из исходного TXT):"
                dialog_start = content.find(dialog_marker)
                if dialog_start == -1:
                    # Fallback: ищем просто "Диалог:"
                    dialog_start = content.find("Диалог:")
                    if dialog_start != -1:
                        dialog_marker = "Диалог:"
                
                if dialog_start != -1:
                    # Находим начало самого диалога (после заголовка и двух переносов строк)
                    dialog_content_start = dialog_start + len(dialog_marker)
                    # Пропускаем возможные пробелы и переносы строк
                    while dialog_content_start < len(content) and content[dialog_content_start] in [' ', '\n', '\r', ':']:
                        dialog_content_start += 1
                    
                    # Находим конец диалога (начало "Распознавание по чек-листу:")
                    dialog_end = content.find("Распознавание по чек-листу:", dialog_content_start)
                    if dialog_end == -1:
                        # Fallback: ищем "Анализ:"
                        dialog_end = content.find("Анализ:", dialog_content_start)
                    
                    if dialog_end != -1:
                        dialog_text = content[dialog_content_start:dialog_end].strip()
                    else:
                        # Если нет маркера конца, берем все после начала диалога
                        dialog_text = content[dialog_content_start:].strip()
                    
                    if dialog_text:
                        print(f"DEBUG: Извлечен диалог из файла анализа (длина: {len(dialog_text)} символов)")
                    else:
                        print(f"DEBUG: Диалог найден, но текст пуст")
            except Exception as e:
                print(f"DEBUG: Не удалось извлечь диалог из файла анализа: {e}")
            
            # Получаем имя оператора с приоритетом: из транскрипции, затем из таблицы
            if station_code and station_code != 'Неизвестно':
                consultant_surname_before = consultant_surname  # Сохраняем для сравнения
                
                # get_operator_name может вернуть "Не указано", если не найдет
                extracted_name = get_operator_name(dialog_text, station_code)
                
                # Если удалось извлечь новое имя, и оно не "Не указано" - обновляем
                if extracted_name and extracted_name != 'Не указано':
                    consultant_surname = extracted_name
                    print(f"DEBUG: Имя оператора обновлено с '{consultant_surname_before}' на '{consultant_surname}' (извлечено из транскрипции/конфига)")
                elif consultant_surname == 'Не указано' and extracted_name == 'Не указано':
                    # Если было "Не указано" и осталось "Не указано", пробуем поискать в конфиге напрямую
                    # (get_operator_name уже должна это делать, но на всякий случай)
                    if station_code in employee_by_extension:
                        consultant_surname = employee_by_extension[station_code]
                        print(f"DEBUG: Имя оператора взято из конфига по коду станции: {consultant_surname}")
                else:
                     print(f"DEBUG: Оставили имя '{consultant_surname}' (get_operator_name вернул '{extracted_name}')")

            # Применяем маппинг станций: если это подстанция, находим основную станцию
            if station_code and station_code != 'Неизвестно':
                # Проверяем, является ли это подстанцией
                main_station = None
                for main_st, substations in main_config.STATION_MAPPING.items():
                    if station_code in substations:
                        main_station = main_st
                        break
                
                # Если это подстанция, используем основную станцию
                if main_station:
                    station_code = main_station
                
                # Получаем название станции из STATION_NAMES
                station_name = station_names.get(station_code, station_code)
                station_code = station_name

            # Извлекаем ответы по числу пунктов чек-листа
            answers = []
            total_q = get_num_questions_from_yaml()
            
            # Ищем ответы в формате "N. ... — ДА/НЕТ"
            for i in range(1, total_q + 1):
                # Формат: "N. ... — ДА/НЕТ"
                dash_match = re.search(rf'^{i}\.\s.*?—\s*(ДА|НЕТ)\s*$', content, re.MULTILINE | re.IGNORECASE)
                if dash_match:
                    answers.append(dash_match.group(1).upper().strip())
                else:
                    # Fallback: ищем в формате [ОТВЕТ: ДА/НЕТ] по порядку
                    answer_matches = re.findall(r'\[ОТВЕТ:\s*(ДА|НЕТ)\]', content, re.IGNORECASE)
                    if i <= len(answer_matches):
                        answers.append(answer_matches[i-1].upper().strip())
                    else:
                        answers.append(None)

            # Получаем название станции из кода
            station_name = main_config.STATION_NAMES.get(station_code, station_code)
            data_for_excel.append([file, consultant_surname, station_name] + answers)

    if data_for_excel:
        total_q = get_num_questions_from_yaml()
        df = pd.DataFrame(
            data_for_excel,
            columns=['Название файла', 'Консультант', 'Название станции'] + [f'Вопрос {i}' for i in range(1, total_q + 1)]
        )
        df.to_excel(output_file_path, sheet_name="Данные", index=False)

        # Замена "ДА" и "НЕТ" на 1 и 0
        replace_yes_no_in_excel(output_file_path)

        create_summary_report(output_file_path)
        print(f'Отчет успешно создан: {output_file_path}')
    else:
        print("Не было найдено данных для записи в Excel.")

def create_summary_report(output_file_path):
    # Получаем конфигурацию станций
    station_names = getattr(main_config, 'STATION_NAMES', {})
    station_mapping = getattr(main_config, 'STATION_MAPPING', {})
    employee_by_extension = getattr(main_config, 'EMPLOYEE_BY_EXTENSION', {})
    
    excel_data = pd.read_excel(output_file_path, sheet_name="Данные")
    total_q = get_num_questions_from_yaml()
    column_names = ['Название файла', 'Консультант', 'Название станции'] + [f'Вопрос {i}' for i in range(1, total_q + 1)]
    # Подгоняем число колонок к фактическим данным на случай расхождений
    if len(excel_data.columns) != len(column_names):
        # Если данных больше — обрежем лишние столбцы справа (это старые хвосты)
        if len(excel_data.columns) > len(column_names):
            excel_data = excel_data.iloc[:, :len(column_names)]
        # Если данных меньше — добавим недостающие столбцы вопросов пустыми
        else:
            for _ in range(len(column_names) - len(excel_data.columns)):
                excel_data[excel_data.columns.size] = None
            excel_data = excel_data.iloc[:, :len(column_names)]
    excel_data.columns = column_names

    exclude_consultant_names = ["Не указано"]
    excel_data = excel_data[~excel_data['Консультант'].isin(exclude_consultant_names)]

    # Заменяем числовые коды станций на полные названия
    # Объединяем основную станцию и подстанцию в одну группу
    station_groups_map = get_station_groups(station_names, station_mapping, employee_by_extension)
    
    # Создаем новую колонку с полными названиями станций
    excel_data['Полное название станции'] = excel_data['Название станции'].astype(str).map(
        lambda x: station_groups_map.get(x, x)
    )

    grouped = excel_data.groupby('Полное название станции')

    workbook = openpyxl.load_workbook(output_file_path)

    for sheet_name in ['Общий процент по станциям', 'Сводный отчет']:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    workbook.create_sheet('Общий процент по станциям', 0)
    sheet_overall = workbook['Общий процент по станциям']
    workbook.create_sheet('Сводный отчет', 1)
    sheet = workbook['Сводный отчет']

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')
    bold_font = Font(bold=True)

    station_totals = []

    row_num = 1
    station_names = get_station_groups()

    # Готовим список вопросов для группировки по размеру чек-листа
    total_q = get_num_questions_from_yaml()
    question_cols = [f'Вопрос {i}' for i in range(1, total_q + 1)]

    for station, group in grouped:
        station_name_full = station  # Теперь station уже содержит полное название

        call_counts = group.groupby('Консультант').size()
        call_counts.name = 'Кол-во звонков'

        pivot_table = group.groupby("Консультант")[question_cols].mean()
        pivot_table = pivot_table.fillna(0)

        pivot_table.insert(0, "Кол-во звонков", call_counts)
        pivot_table['Итог'] = pivot_table[question_cols].mean(axis=1)

        for col in question_cols + ['Итог']:
            pivot_table[col] = pivot_table[col] * 100

        total_calls = pivot_table['Кол-во звонков'].sum()
        weighted_totals = {}
        for col in question_cols + ['Итог']:
            weighted_sum = (pivot_table[col] * pivot_table['Кол-во звонков']).sum()
            weighted_mean = weighted_sum / total_calls
            weighted_totals[col] = weighted_mean

        overall_mean = pd.DataFrame({
            'Кол-во звонков': [total_calls],
            **weighted_totals
        }, index=['Общий итог'])

        station_total_percentage = overall_mean['Итог'].values[0]
        station_totals.append([station_name_full, station_total_percentage])

        final_table = pd.concat([pivot_table, overall_mean])

        sheet.cell(row=row_num, column=1, value=f"Станция: {station_name_full}")
        sheet.cell(row=row_num, column=1).font = bold_font
        row_num += 1

        headers = ['Консультант'] + list(final_table.columns)
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=row_num, column=col_num, value=header)
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = center_alignment

        row_num += 1

        for idx, row_data in final_table.iterrows():
            sheet.cell(row=row_num, column=1, value=idx)
            sheet.cell(row=row_num, column=1).border = thin_border
            sheet.cell(row=row_num, column=1).alignment = center_alignment
            for col_num, (col_name, value) in enumerate(row_data.items(), 2):
                cell = sheet.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = center_alignment
                if col_name != 'Кол-во звонков':
                    cell.value = value / 100
                    cell.number_format = "0%"
            row_num += 1

        row_num += 2

    station_totals_sorted = sorted(station_totals, key=lambda x: x[1], reverse=True)
    station_totals_ranked = [(idx + 1, name, perc) for idx, (name, perc) in enumerate(station_totals_sorted)]

    sheet_overall.cell(row=1, column=1, value='Место')
    sheet_overall.cell(row=1, column=2, value='Название станции')
    sheet_overall.cell(row=1, column=3, value='% выполнения')
    for col in range(1, 4):
        cell = sheet_overall.cell(row=1, column=col)
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = thin_border

    for row_idx, (rank, station_name, total_percentage) in enumerate(station_totals_ranked, start=2):
        sheet_overall.cell(row=row_idx, column=1, value=rank)
        sheet_overall.cell(row=row_idx, column=2, value=station_name)
        sheet_overall.cell(row=row_idx, column=3, value=total_percentage / 100)
        for col in range(1, 4):
            cell = sheet_overall.cell(row=row_idx, column=col)
            cell.alignment = center_alignment
            cell.border = thin_border
        sheet_overall.cell(row=row_idx, column=3).number_format = "0%"

        # >>> NEW CODE <<<
        # Добавляем строку "Общий процент" после перечисления станций
    if station_totals_sorted:
        overall_avg = sum(perc for _, perc in station_totals_sorted) / len(station_totals_sorted)
    else:
        overall_avg = 0

        # row_idx сейчас указывает на строку после последней станции
    row_idx = sheet_overall.max_row + 1

    # Пустая ячейка в колонке "Место" (первая колонка)
    cell_place = sheet_overall.cell(row=row_idx, column=1, value=None)
    cell_place.border = thin_border
    cell_place.alignment = center_alignment

    # Текст "Общий процент" во второй колонке
    cell_title = sheet_overall.cell(row=row_idx, column=2, value="Общий процент")
    cell_title.font = bold_font
    cell_title.alignment = center_alignment
    cell_title.border = thin_border

    # Среднее значение в третьей колонке
    cell_value = sheet_overall.cell(row=row_idx, column=3, value=overall_avg / 100)
    cell_value.number_format = "0%"
    cell_value.border = thin_border
    cell_value.alignment = center_alignment
    # >>> END NEW CODE <<<

    sheet_overall.column_dimensions['A'].width = 10
    sheet_overall.column_dimensions['B'].width = 35
    sheet_overall.column_dimensions['C'].width = 15

    for row in sheet_overall.iter_rows(min_row=1, max_row=sheet_overall.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = thin_border

    notes_start_row = row_num + 1
    # Динамический список пунктов чек-листа из YAML
    notes = get_checklist_titles_from_yaml()

    sheet.column_dimensions['A'].width = 33

    for i, note in enumerate(notes, start=notes_start_row):
        cell = sheet.cell(row=i, column=1, value=note)
        cell.alignment = Alignment(horizontal='left', vertical='top')
        cell.font = Font(size=10)

    if 'Данные' in workbook.sheetnames:
        data_sheet = workbook['Данные']
        workbook.remove(data_sheet)
        workbook._add_sheet(data_sheet, index=2)

    workbook.save(output_file_path)
    print(f"Сводный отчет добавлен в {output_file_path}")

def resolve_consultant_name(dialog_text, station_code, employee_map):
    if employee_map and station_code in employee_map:
        employee_full = employee_map.get(station_code)
        if isinstance(employee_full, str) and employee_full.strip():
            parts = employee_full.strip().split()
            return parts[0] if parts else employee_full.strip()
        return str(employee_full)
    return get_operator_name(dialog_text, station_code)

def compute_realtime_summary(
    period_start: datetime,
    period_end: datetime,
    allowed_stations: list = None,
    base_folder=None,
    station_names=None,
    station_mapping=None,
    employee_by_extension=None,
    script_prompt_path=None
) -> dict:
    """
    Строит онлайн-сводку (как вкладка "Сводный отчет") без генерации Excel.
    Источник данных: готовые *_analysis.txt из папок transcriptions/script_8 за период.
    Консультант берётся из EMPLOYEE_BY_EXTENSION по коду станции.
    Возвращает JSON-подобную структуру с агрегацией по станциям и консультантам.
    allowed_stations: список разрешенных кодов станций. Если None - все станции.
    """
    if not base_folder:
        base_folder = str(main_config.BASE_RECORDS_PATH)
    else:
        base_folder = str(base_folder)
    station_names = station_names or getattr(main_config, 'STATION_NAMES', {})
    station_mapping = station_mapping or getattr(main_config, 'STATION_MAPPING', {})
    employee_by_extension = employee_by_extension or getattr(main_config, 'EMPLOYEE_BY_EXTENSION', {})
    records = []

    total_q = get_num_questions_from_yaml(script_prompt_path)

    day_count = (period_end.date() - period_start.date()).days
    for i in range(day_count + 1):
        day = (period_start + timedelta(days=i)).strftime('%Y/%m/%d')
        script_8_path = os.path.join(base_folder, day, 'transcriptions', 'script_8')
        if not os.path.exists(script_8_path):
            continue
        for file in os.listdir(script_8_path):
            if not file.endswith('_analysis.txt'):
                continue
            base_name = file[:-13]
            station_code = 'Неизвестно'
            phone_number = None
            date_time_obj = None

            # fs_ формат
            match_fs = re.match(r'fs_(\+?\d+)_\d{3,4}_(\d{4}-\d{2}-\d{2})-(\d{2}-\d{2}-\d{2})', base_name)
            if match_fs:
                phone_number = match_fs.group(1).lstrip('+')
                date_str = match_fs.group(2)
                time_str = match_fs.group(3)
                try:
                    date_time_obj = datetime.strptime(f"{date_str} {time_str.replace('-', ':')}", '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    date_time_obj = None

            # external- формат
            if not date_time_obj and base_name.lower().startswith('external-'):
                try:
                    parts = base_name.split('-')
                    station_code = parts[1]
                    phone_number = parts[2].lstrip('+')
                    yyyymmdd = parts[3]
                    hhmmss = parts[4]
                    date_time_obj = datetime.strptime(f"{yyyymmdd} {hhmmss}", '%Y%m%d %H%M%S')
                except Exception:
                    date_time_obj = None

            # Если станция не была определена из fs_, попытаемся достать её из parse_filename
            if station_code == 'Неизвестно':
                try:
                    ph2, st2, dt2 = parse_filename(base_name)
                    if st2:
                        station_code = st2
                    if not date_time_obj:
                        date_time_obj = dt2
                    if not phone_number:
                        phone_number = ph2.lstrip('+') if ph2 else None
                except Exception:
                    pass

            # Файл содержимого
            try:
                with open(os.path.join(script_8_path, file), 'r', encoding='utf-8') as f:
                    content = f.read()
            except Exception:
                continue

            # Извлекаем ответы
            answers = []
            for i_q in range(1, total_q + 1):
                dash_match = re.search(rf'^{i_q}\.\s.*?—\s*(ДА|НЕТ)\s*$', content, re.MULTILINE | re.IGNORECASE)
                if dash_match:
                    answers.append(1 if dash_match.group(1).upper().strip() == 'ДА' else 0)
                else:
                    seq = re.findall(r'\[ОТВЕТ:\s*(ДА|НЕТ)\]', content, re.IGNORECASE)
                    if i_q <= len(seq):
                        answers.append(1 if seq[i_q - 1].upper().strip() == 'ДА' else 0)
                    else:
                        answers.append(None)

            # Приоритет 1: Извлекаем имя оператора из транскрипции (диалога)
            # Пытаемся извлечь диалог из файла анализа
            dialog_text = None
            try:
                # Структура файла: "Диалог (из исходного TXT):\n\n{dialog_text}\n\nРаспознавание по чек-листу:\n\n"
                dialog_marker = "Диалог (из исходного TXT):"
                dialog_start = content.find(dialog_marker)
                if dialog_start == -1:
                    # Fallback: ищем просто "Диалог:"
                    dialog_start = content.find("Диалог:")
                    if dialog_start != -1:
                        dialog_marker = "Диалог:"
                
                if dialog_start != -1:
                    # Находим начало самого диалога (после заголовка и двух переносов строк)
                    dialog_content_start = dialog_start + len(dialog_marker)
                    # Пропускаем возможные пробелы и переносы строк
                    while dialog_content_start < len(content) and content[dialog_content_start] in [' ', '\n', '\r', ':']:
                        dialog_content_start += 1
                    
                    # Находим конец диалога (начало "Распознавание по чек-листу:")
                    dialog_end = content.find("Распознавание по чек-листу:", dialog_content_start)
                    if dialog_end == -1:
                        # Fallback: ищем "Анализ:"
                        dialog_end = content.find("Анализ:", dialog_content_start)
                    
                    if dialog_end != -1:
                        dialog_text = content[dialog_content_start:dialog_end].strip()
                    else:
                        # Если нет маркера конца, берем все после начала диалога
                        dialog_text = content[dialog_content_start:].strip()
            except Exception:
                pass
            
            # Получаем имя оператора с приоритетом: из транскрипции, затем из таблицы
            consultant = resolve_consultant_name(dialog_text, station_code, employee_by_extension)

            # Получаем название станции из кода
            station_name = main_config.STATION_NAMES.get(station_code, station_code)

            # Фильтрация по разрешенным станциям
            if allowed_stations is not None and station_code not in allowed_stations:
                continue

            records.append({
                'consultant': consultant,
                'station_code': station_code,
                'station_name': station_name,
                'answers': answers
            })

    if not records:
        return {
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'stations': [],
            'ranking': []
        }

    # Преобразуем в DataFrame
    rows = []
    for r in records:
        row = {
            'Консультант': r['consultant'],
            'Название станции': r['station_name']
        }
        for idx in range(total_q):
            row[f'Вопрос {idx+1}'] = r['answers'][idx]
        rows.append(row)

    df = pd.DataFrame(rows)

    # Приводим столбцы с вопросами к числовому типу, чтобы избежать ошибок при агрегации (особенно если там None)
    question_cols = [f'Вопрос {i}' for i in range(1, total_q + 1)]
    for col in question_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # Группировка станций
    station_names = get_station_groups()
    df['Полное название станции'] = df['Название станции'].astype(str).map(lambda x: station_names.get(x, x))

    question_cols = [f'Вопрос {i}' for i in range(1, total_q + 1)]
    grouped = df.groupby('Полное название станции')

    stations_out = []
    station_totals = []
    for station_name, group in grouped:
        call_counts = group.groupby('Консультант').size()
        pivot = group.groupby('Консультант')[question_cols].mean().fillna(0)
        pivot.insert(0, 'Кол-во звонков', call_counts)
        pivot['Итог'] = pivot[question_cols].mean(axis=1)

        # В проценты
        pivot_percent = pivot.copy()
        for c in question_cols + ['Итог']:
            pivot_percent[c] = (pivot_percent[c] * 100).round(2)

        total_calls = pivot['Кол-во звонков'].sum()
        weighted_totals = {}
        for c in question_cols + ['Итог']:
            weighted_totals[c] = ((pivot[c] * pivot['Кол-во звонков']).sum() / max(total_calls, 1)) * 100
        station_total = weighted_totals['Итог']
        station_totals.append((station_name, station_total))

        stations_out.append({
            'station': station_name,
            'consultants': [
                {
                    'name': row['Консультант'],
                    'calls': int(row['Кол-во звонков']),
                    'percent_total': float(row['Итог'])
                } | {f'q{i+1}': float(row[f'Вопрос {i+1}']) for i in range(total_q)}
                for idx, row in pivot_percent.reset_index().iterrows()
            ],
            'station_percent_total': round(float(station_total), 2)
        })

    ranking_sorted = sorted(station_totals, key=lambda x: x[1], reverse=True)
    ranking = [
        {'rank': i+1, 'station': name, 'percent_total': round(float(perc), 2)}
        for i, (name, perc) in enumerate(ranking_sorted)
    ]

    return {
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'stations': stations_out,
        'ranking': ranking,
        'total_questions': total_q
    }

def run_week_full(start_date: datetime = None, end_date: datetime = None, base_folder=None):
    # Интервал по умолчанию: прошлая неделя
    if not start_date or not end_date:
        last_monday, last_sunday, _ = get_last_week_dates()
        start_date = start_date or last_monday
        end_date = end_date or last_sunday
    
    # Сбрасываем время на начало дня для start_date и конец дня для end_date
    if start_date:
        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
    if end_date:
        end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
    
    return analyze_files(start_date, end_date, base_folder=base_folder)
