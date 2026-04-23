#!/usr/bin/env python3
"""
Движок классификации звонков с интеграцией обучения
"""

import os
import re
import requests
import pandas as pd
import time
import hashlib
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
try:
    from .training_examples import TrainingExamplesManager
    from .classification_rules import ClassificationRulesManager
except ImportError:
    from training_examples import TrainingExamplesManager
    from classification_rules import ClassificationRulesManager
import json
import shutil
import tempfile

# Если у пользователя нет станций в ЛК — порядок колонок по умолчанию (автосервис BW)
FALLBACK_REPORT_STATIONS_ORDER = (
    "Чон",
    "Чон К",
    "Сах",
    "Род",
    "Брн",
    "Кмн",
    "Кбш",
    "Дзр",
    "Меч",
    "Тгн",
    "Влд",
    "КзнИ",
    "КзнС",
    "Рдн",
)


class CallClassificationEngine:
    """Движок классификации звонков с дообучением"""
    
    def __init__(
        self,
        api_key=None,
        base_url=None,
        model=None,
        user_id=None,
        classification_root=None,
        station_names=None,
        station_mapping=None,
        station_report_names=None,
        station_report_order=None,
    ):
        """
        Инициализация движка классификации.

        Параметры:
        - user_id: владелец данных в PostgreSQL (обязателен, иначе CLASSIFICATION_USER_ID)
        - classification_root: каталог .../classification (лог, uploads, артефакты), обязателен
        - api_key: ключ доступа к LLM (если не передан, берётся из переменной окружения THEBAI_API_KEY)
        - base_url: базовый URL LLM (если не передан, берётся из THEBAI_URL или используется https://api.deepseek.com/v1)
        - model: имя модели (если не передано, берётся из THEBAI_MODEL или используется deepseek-chat)
        """
        # Настройка API клиента (по умолчанию — DeepSeek совместимый endpoint)
        api_key = api_key or os.getenv("THEBAI_API_KEY", "")
        base_url = base_url or os.getenv("THEBAI_URL", "https://api.deepseek.com/v1/chat/completions")
        self.model = model or os.getenv("THEBAI_MODEL", "deepseek-chat")
        self.api_key = api_key
        self.base_url = base_url

        uid = user_id
        if uid is None:
            env_uid = os.getenv("CLASSIFICATION_USER_ID")
            uid = int(env_uid) if env_uid and str(env_uid).isdigit() else None
        if uid is None or int(uid) <= 0:
            raise ValueError("CallClassificationEngine: укажите user_id= или CLASSIFICATION_USER_ID")
        self.user_id = int(uid)

        if classification_root is None:
            raise ValueError("CallClassificationEngine: укажите classification_root= (каталог .../classification)")
        root = Path(classification_root)
        self.classification_root = root
        self.debug_log_path = root / "classification_llm_debug.log"
        self.uploads_dir = root / "uploads"
        self.training_manager = TrainingExamplesManager(user_id=self.user_id, classification_root=root)
        self.rules_manager = ClassificationRulesManager(user_id=self.user_id, classification_root=root)
        
        # Словарь для преобразования номеров станций в названия
        self.STATION_CODES = {
            '9322': 'Брн', '4231': 'Брн', '4230': 'Брн',
            '9326': 'Рпб', '4160': 'Рпб',
            '9347': 'Род', '4254': 'Род', '4255': 'Род',
            '9327': 'Чон', '4210': 'Чон', '4211': 'Чон',
            '9325': 'Чон К', '4217': 'Чон К',
            '9300': 'Сах', '4222': 'Сах', '4221': 'Сах',
            '9321': 'Кмн', '4200': 'Кмн', '4201': 'Кмн',
            '9324': 'Хлз', '4240': 'Хлз',
            '9344': 'Кбш', '4253': 'Кбш', '4256': 'Кбш',
            '9304': 'Дзр', '4100': 'Дзр', '4101': 'Дзр',
            '9308': 'Арз', '4110': 'Арз', '4111': 'Арз',
            '9301': 'Влд', '4140': 'Влд', '4141': 'Влд',
            '9302': 'КзнИ', '4155': 'КзнИ', '4156': 'КзнИ',
            '9307': 'КзнС', '4150': 'КзнС', '5151': 'КзнС',
            '9350': 'Рдн', '4257': 'Рдн', '4258': 'Рдн',
            '9316': 'Меч', '4170': 'Меч', '4172': 'Меч',
            '9319': 'Тгн', '4181': 'Тгн', '4180': 'Тгн'
        }

        # Пользовательские станции из ЛК имеют приоритет.
        user_station_codes = {}
        if isinstance(station_names, dict):
            for code, name in station_names.items():
                code_s = str(code).strip()
                name_s = str(name or code).strip()
                if code_s and name_s:
                    user_station_codes[code_s] = name_s
            if user_station_codes:
                self.STATION_CODES.update(user_station_codes)
        
        # Загружаем дополнительные коды станций из настроек (если заданы)
        try:
            extra_codes_json = self.rules_manager.get_setting('station_codes_extra', '')
            if extra_codes_json:
                extra_codes = json.loads(extra_codes_json)
                if isinstance(extra_codes, dict):
                    # Ключи всегда приводим к строкам, чтобы не зависеть от типа в JSON
                    self.STATION_CODES.update({str(k): v for k, v in extra_codes.items()})
        except Exception:
            # При ошибке настроек продолжаем с базовым набором кодов
            pass
        
        # Станции, которые нужно исключить из проекта
        # Арз  - Арзамас
        # Хлз  - Хальзовская
        # Рпб  - Республиканская
        self.EXCLUDED_STATIONS = ['Арз', 'Хлз', 'Рпб']

        # Привязка подстанций к основным станциям пользователя.
        if isinstance(station_mapping, dict):
            for main_code, sub_codes in station_mapping.items():
                main_code_s = str(main_code).strip()
                if not main_code_s:
                    continue
                main_name = self.STATION_CODES.get(main_code_s, main_code_s)
                for sub_code in (sub_codes or []):
                    sub_code_s = str(sub_code).strip()
                    if sub_code_s:
                        self.STATION_CODES[sub_code_s] = main_name

        # Для пользовательских профилей убираем legacy-исключения (Арз/Хлз/Рпб)
        if user_station_codes:
            self.EXCLUDED_STATIONS = []

        # Порядок и подписи колонок «Станция» в Excel — из ЛК (Название для отчётов + порядок карточек)
        sn = dict(station_names or {})
        srn = dict(station_report_names or {}) if station_report_names else {}
        sro = list(station_report_order or []) if station_report_order else []

        self._station_display_by_code = {str(k): str(v) for k, v in sn.items()}
        self._report_label_by_code = {}
        labels_order = []

        for code in sro:
            cs = str(code).strip()
            if not cs or cs not in sn:
                continue
            lbl = (srn.get(cs) or sn.get(cs) or cs).strip()
            if not lbl:
                continue
            self._report_label_by_code[cs] = lbl
            if lbl not in labels_order:
                labels_order.append(lbl)

        for cs in sorted(sn.keys(), key=lambda x: str(x)):
            if cs in self._report_label_by_code:
                continue
            lbl = (srn.get(cs) or sn.get(cs) or cs).strip()
            self._report_label_by_code[cs] = lbl
            if lbl not in labels_order:
                labels_order.append(lbl)

        if isinstance(station_mapping, dict):
            for main_code, sub_codes in station_mapping.items():
                main_s = str(main_code).strip()
                ml = self._report_label_by_code.get(main_s)
                if not ml:
                    continue
                for sub_code in (sub_codes or []):
                    sc = str(sub_code).strip()
                    if sc:
                        self._report_label_by_code[sc] = ml

        if labels_order:
            self.ALL_STATIONS = labels_order
        else:
            self.ALL_STATIONS = list(FALLBACK_REPORT_STATIONS_ORDER)

        self._report_label_by_display = {}
        for code, lbl in self._report_label_by_code.items():
            disp = str(self.STATION_CODES.get(code, "")).strip()
            if disp:
                self._report_label_by_display[disp] = lbl
        for code, nm in self._station_display_by_code.items():
            nm_clean = str(nm).strip()
            if nm_clean:
                lbl = self._report_label_by_code.get(code)
                if lbl:
                    self._report_label_by_display[nm_clean] = lbl
        for lbl in self.ALL_STATIONS:
            self._report_label_by_display[lbl] = lbl

        self._has_user_station_list = bool(self._station_display_by_code)
        
        # Новая схема категорий согласно Google Sheets
        # Новая схема категорий согласно ТЗ v1.0
        self.NEW_CATEGORIES = {
            "IN.NE": "Входящие - Не целевые",
            "IN.CONS.MSG": "Входящие - Целевые - Консультация - ПЕРЕШЛИ В МЕССЕНДЖЕР",
            "IN.CONS.REDIR": "Входящие - Целевые - Консультация - ПЕРЕАДРЕСАЦИЯ",
            "IN.CONS.OWN": "Входящие - Целевые - Консультация - СВОИ ЗАПЧАСТИ",
            "IN.CONS.THINK": "Входящие - Целевые - Консультация - ПОДУМАЕТ/ОТКАЗ",
            "IN.CONS.BUSY": "Входящие - Целевые - Консультация - НЕТ ВРЕМЕНИ/ЗАНЯТО",
            "IN.CONS.COST": "Входящие - Целевые - Консультация - ВЫСОКАЯ СТОИМОСТЬ",
            "IN.CONS.NODO": "Входящие - Целевые - Консультация - НЕ ВЫПОЛНЯЕМ РАБОТЫ",
            "IN.CONS.CB": "Входящие - Целевые - Консультация - ЗАПЛАНИРОВАН ПЕРЕЗВОН",
            "IN.CONS.OTHER": "Входящие - Целевые - Консультация - Общая",
            "IN.BOOK": "Входящие - Целевые - Запись",
            "IN.FU.BOOK": "Входящие - Целевые - Последующий контакт с записью",
            "IN.INFO.FU.NOBOOK": "Входящие - Справочные - Последующий контакт без записи",
            "OUT.NE": "Исходящие - Не целевые",
            "OUT.CONS.MSG": "Исходящие - Целевые - Консультация - ПЕРЕШЛИ В МЕССЕНДЖЕР",
            "OUT.CONS.REDIR": "Исходящие - Целевые - Консультация - ПЕРЕАДРЕСАЦИЯ",
            "OUT.CONS.OWN": "Исходящие - Целевые - Консультация - СВОИ ЗАПЧАСТИ",
            "OUT.CONS.THINK": "Исходящие - Целевые - Консультация - ПОДУМАЕТ/ОТКАЗ",
            "OUT.CONS.BUSY": "Исходящие - Целевые - Консультация - НЕТ ВРЕМЕНИ/ЗАНЯТО",
            "OUT.CONS.COST": "Исходящие - Целевые - Консультация - ВЫСОКАЯ СТОИМОСТЬ",
            "OUT.CONS.NODO": "Исходящие - Целевые - Консультация - НЕ ВЫПОЛНЯЕМ РАБОТЫ",
            "OUT.CONS.CB": "Исходящие - Целевые - Консультация - ЗАПЛАНИРОВАН ПЕРЕЗВОН",
            "OUT.CONS.OTHER": "Исходящие - Целевые - Консультация - Общая",
            "OUT.BOOK": "Исходящие - Целевые - Запись",
            "OUT.FU.BOOK": "Исходящие - Целевые - Последующий контакт с записью",
            "OUT.INFO.FU.NOBOOK": "Исходящие - Справочные - Последующий контакт без записи",
            "OUT.OBZ.BOOK": "Исходящие - Обзвон - С записью",
            "OUT.OBZ.NOBOOK": "Исходящие - Обзвон - Без записи"
        }
        
        # Соответствие старых категорий новым (для обратной совместимости)
        self.LEGACY_TO_NEW_MAPPING = {
            "1": "1",   # НЕ ЦЕЛЕВОЙ -> Входящие - Не целевые
            "2": "11",  # ЗАПИСЬ НА СЕРВИС -> Входящие - Целевые - Запись
            "3": "2",   # КОНСУЛЬТАЦИЯ -> Входящие - Целевые - Консультация
            "4": "6",   # ПОДУМАЕТ/ОТКАЗ -> Входящие - Целевые - ПОДУМАЕТ/ОТКАЗ
            "5": "7",   # НЕТ ВРЕМЕНИ/ЗАНЯТО -> Входящие - Целевые - НЕТ ВРЕМЕНИ/ЗАНЯТО
            "6": "8",   # ВЫСОКАЯ СТОИМОСТЬ -> Входящие - Целевые - ВЫСОКАЯ СТОИМОСТЬ
            "7": "5",   # СВОИ ЗАПЧАСТИ -> Входящие - Целевые - СВОИ ЗАПЧАСТИ
            "8": "9",   # НЕ ВЫПОЛНЯЕМ РАБОТЫ -> Входящие - Целевые - НЕ ВЫПОЛНЯЕМ РАБОТЫ
            "9": "3",   # ПЕРЕШЛИ В МЕССЕНДЖЕР -> Входящие - Целевые - ПЕРЕШЛИ В МЕССЕНДЖЕР
            "10": "10", # ЗАПЛАНИРОВАН ПЕРЕЗВОН -> Входящие - Целевые - ЗАПЛАНИРОВАН ПЕРЕЗВОН
            "11": "4",  # ПЕРЕАДРЕСАЦИЯ -> Входящие - Целевые - ПЕРЕАДРЕСАЦИЯ
            "12": "14", # ОБЗВОН -> Исходящие - Не целевые
            "13": "12", # ПОСЛЕДУЮЩИЙ КОНТАКТ -> Входящие - Целевые - Последующий контакт с записью
            "14": "1"   # ДРУГОЕ -> Входящие - Не целевые
        }
        
        # Группировка для отображения
        # Группировка для отображения
        self.CATEGORY_GROUPS = {
            "IN.NE": "Не целевые",
            "IN.CONS.MSG": "Целевые",
            "IN.CONS.REDIR": "Целевые",
            "IN.CONS.OWN": "Целевые",
            "IN.CONS.THINK": "Целевые",
            "IN.CONS.BUSY": "Целевые",
            "IN.CONS.COST": "Целевые",
            "IN.CONS.NODO": "Целевые",
            "IN.CONS.CB": "Целевые",
            "IN.CONS.OTHER": "Целевые",
            "IN.BOOK": "Целевые",
            "IN.FU.BOOK": "Целевые",
            "IN.INFO.FU.NOBOOK": "Справочные",
            "OUT.NE": "Не целевые",
            "OUT.CONS.MSG": "Целевые",
            "OUT.CONS.REDIR": "Целевые",
            "OUT.CONS.OWN": "Целевые",
            "OUT.CONS.THINK": "Целевые",
            "OUT.CONS.BUSY": "Целевые",
            "OUT.CONS.COST": "Целевые",
            "OUT.CONS.NODO": "Целевые",
            "OUT.CONS.CB": "Целевые",
            "OUT.CONS.OTHER": "Целевые",
            "OUT.BOOK": "Целевые",
            "OUT.FU.BOOK": "Целевые",
            "OUT.INFO.FU.NOBOOK": "Справочные",
            "OUT.OBZ.BOOK": "Обзвон",
            "OUT.OBZ.NOBOOK": "Обзвон"
        }
        
        # Системный промпт будет генерироваться динамически из правил

        # Пользовательский шаблон имени файла (регулярное выражение)
        # Если включен в настройках, используется перед стандартными шаблонами
        self.custom_filename_pattern = None
        self.custom_pattern_enabled = False
        self.custom_filename_patterns = []
        try:
            pattern_str = self.rules_manager.get_setting('filename_pattern_custom', '') or ''
            enabled = self.rules_manager.get_setting('filename_pattern_custom_enabled', '0') == '1'
            patterns_json = self.rules_manager.get_setting('filename_patterns_json', '') or ''
            if patterns_json:
                loaded_patterns = json.loads(patterns_json)
                if isinstance(loaded_patterns, list):
                    self.custom_filename_patterns = loaded_patterns
            if pattern_str and enabled:
                self.custom_filename_pattern = re.compile(pattern_str)
                self.custom_pattern_enabled = True
        except Exception:
            self.custom_filename_pattern = None
            self.custom_pattern_enabled = False
            self.custom_filename_patterns = []

    def _extract_with_custom_pattern(self, filename):
        """Пытается разобрать имя файла по пользовательским паттернам из ЛК."""
        if not getattr(self, 'custom_pattern_enabled', False):
            return None

        for pattern_cfg in self.custom_filename_patterns:
            try:
                regex = str((pattern_cfg or {}).get('regex') or '').strip()
                if not regex:
                    continue
                match = re.search(regex, filename)
                if not match:
                    continue

                phone_group = int((pattern_cfg or {}).get('phone_group') or 1)
                station_group = int((pattern_cfg or {}).get('station_group') or 2)
                date_group = int((pattern_cfg or {}).get('date_group') or 3)
                time_group = int((pattern_cfg or {}).get('time_group') or 4)
                datetime_group = (pattern_cfg or {}).get('datetime_group')
                datetime_format = str((pattern_cfg or {}).get('datetime_format') or '').strip()
                direction = str((pattern_cfg or {}).get('direction') or 'auto').strip().lower()

                phone_number = match.group(phone_group)
                station_number = match.group(station_group)
                call_date = "Не распознана"
                call_time = "Не распознано"

                if datetime_group:
                    datetime_value = match.group(int(datetime_group))
                    if datetime_format:
                        dt = datetime.strptime(datetime_value, datetime_format)
                        call_date = dt.strftime('%d.%m.%Y')
                        call_time = dt.strftime('%H:%M')
                    else:
                        call_date = datetime_value
                else:
                    date_str = match.group(date_group)
                    time_str = match.group(time_group)
                    try:
                        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                        call_date = date_obj.strftime('%d.%m.%Y')
                    except Exception:
                        call_date = date_str
                    try:
                        time_obj = datetime.strptime(time_str, '%H-%M-%S')
                        call_time = time_obj.strftime('%H:%M')
                    except Exception:
                        call_time = time_str

                if direction == 'incoming':
                    call_type = "Входящий"
                elif direction == 'outgoing':
                    call_type = "Исходящий"
                else:
                    if len(str(phone_number)) >= 10 and 3 <= len(str(station_number)) <= 5:
                        call_type = "Исходящий"
                    elif len(str(station_number)) >= 10 and 3 <= len(str(phone_number)) <= 5:
                        call_type = "Входящий"
                    else:
                        call_type = "Не определен"

                station_name = self.STATION_CODES.get(station_number, f"Неизвестная станция ({station_number})")
                return phone_number, call_date, call_time, station_number, station_name, call_type
            except Exception:
                continue

        return None

    def extract_file_info(self, filename):
        """Извлечение номера телефона, даты, времени, номера станции и типа звонка из имени файла"""
        station_number = "Не распознан"
        phone_number = "Не распознан"
        call_date = "Не распознана"
        call_time = "Не распознано"
        call_type = "Не определен"

        parsed_custom = self._extract_with_custom_pattern(filename)
        if parsed_custom:
            return parsed_custom

        # 1. Сначала пробуем пользовательский шаблон, если он включен.
        #    Ожидается, что группы 1-4 соответствуют: телефон, станция, дата, время.
        if getattr(self, 'custom_pattern_enabled', False) and self.custom_filename_pattern:
            try:
                match = self.custom_filename_pattern.search(filename)
            except re.error:
                match = None
            if match:
                try:
                    phone_number = match.group(1)
                    station_number = match.group(2)
                    date_str = match.group(3)
                    time_str = match.group(4)
                    
                    try:
                        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                        call_date = date_obj.strftime('%d.%m.%Y')
                    except Exception:
                        call_date = date_str
                    
                    try:
                        time_obj = datetime.strptime(time_str, '%H-%M-%S')
                        call_time = time_obj.strftime('%H:%M')
                    except Exception:
                        call_time = time_str
                    
                    station_name = self.STATION_CODES.get(station_number, f"Неизвестная станция ({station_number})")
                    return phone_number, call_date, call_time, station_number, station_name, call_type
                except Exception:
                    # При любой ошибке падаем обратно на стандартную логику
                    phone_number = "Не распознан"
                    station_number = "Не распознан"
                    call_date = "Не распознана"
                    call_time = "Не распознано"
                    call_type = "Не определен"

        # 2. Стандартные шаблоны для известных форматов файлов
        patterns = [
            # Компактный формат: телефон_станция_YYYYMMDD-HHMMSS[-суффикс] (напр. 79673923233_201_20260313-151817-LbWWbypolYXY)
            r'(\d{10,11})_(\d{3,5})_(\d{8})-(\d{6})(?:-[\w.]+)?',
            # Новый формат без префикса fs_, варианты:
            # 1) Исходящие: телефон_станция_дата-время
            #    09278612779_401_2025-12-12-14-20-02.txt
            # 2) Входящие: станция_телефон_дата-время
            #    401_79613420826_2026-01-20-11-02-25.txt
            r'(\d{10,11})_(\d{3,5})_(\d{4}-\d{2}-\d{2})-(\d{2}-\d{2}-\d{2})',
            r'(\d{3,5})_(\d{10,11})_(\d{4}-\d{2}-\d{2})-(\d{2}-\d{2}-\d{2})',
            # Старые форматы с префиксом fs_
            r'fs_(\d{11})_(\d{4})_(\d{4}-\d{2}-\d{2})-(\d{2}-\d{2}-\d{2})',  # Входящий: fs_79084901148_9327_2025-09-25-10-11-07
            r'fs_(\d{4})_(\d{11})_(\d{4}-\d{2}-\d{2})-(\d{2}-\d{2}-\d{2})',  # Исходящий: fs_9307_79872960287_2025-09-25-09-47-23
            r'(\d{10,11})_(\d{4})_(\d{4}-\d{2}-\d{2})-(\d{2}-\d{2}-\d{2})',
            r'.*?(\d{10,11}).*?(\d{3,5}).*?(\d{4}-\d{2}-\d{2})-(\d{2}-\d{2}-\d{2})'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, filename)
            if match:
                if pattern.startswith(r'(\d{10,11})_(\d{3,5})_'):
                    # Исходящий звонок: телефон_станция_дата-время
                    phone_number = match.group(1)
                    station_number = match.group(2)
                    call_type = "Исходящий"
                elif pattern.startswith(r'(\d{3,5})_(\d{10,11})_'):
                    # Входящий звонок: станция_телефон_дата-время
                    station_number = match.group(1)
                    phone_number = match.group(2)
                    call_type = "Входящий"
                elif r'fs_(\d{11})_(\d{4})_' in pattern:
                    # Входящий звонок: fs_79084901148_9327_2025-09-25-10-11-07
                    phone_number = match.group(1)
                    station_number = match.group(2)
                    call_type = "Входящий"
                elif r'fs_(\d{4})_(\d{11})_' in pattern:
                    # Исходящий звонок: fs_9307_79872960287_2025-09-25-09-47-23
                    station_number = match.group(1)
                    phone_number = match.group(2)
                    call_type = "Исходящий"
                else:
                    # Другие форматы - определяем по длине групп:
                    # одна группа 10–11 цифр (телефон), другая 3–5 цифр (станция)
                    g1 = match.group(1)
                    g2 = match.group(2)
                    if len(g1) >= 10 and 3 <= len(g2) <= 5:
                        phone_number = g1
                        station_number = g2
                    elif len(g2) >= 10 and 3 <= len(g1) <= 5:
                        station_number = g1
                        phone_number = g2
                    else:
                        phone_number = g1
                        station_number = g2
                    call_type = "Не определен"
                
                date_str = match.group(3)
                time_str = match.group(4)
                
                # Компактный формат: YYYYMMDD и HHMMSS (без дефисов)
                if len(date_str) == 8 and len(time_str) == 6 and date_str.isdigit() and time_str.isdigit():
                    try:
                        date_obj = datetime.strptime(date_str, '%Y%m%d')
                        call_date = date_obj.strftime('%d.%m.%Y')
                    except Exception:
                        call_date = date_str
                    try:
                        time_obj = datetime.strptime(time_str, '%H%M%S')
                        call_time = time_obj.strftime('%H:%M')
                    except Exception:
                        call_time = time_str
                else:
                    try:
                        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                        call_date = date_obj.strftime('%d.%m.%Y')
                    except Exception:
                        call_date = date_str
                    try:
                        time_obj = datetime.strptime(time_str, '%H-%M-%S')
                        call_time = time_obj.strftime('%H:%M')
                    except Exception:
                        call_time = time_str
                
                break
        
        station_name = self.STATION_CODES.get(station_number, f"Неизвестная станция ({station_number})")
        
        return phone_number, call_date, call_time, station_number, station_name, call_type

    def convert_legacy_to_new_category(self, legacy_category, call_type):
        """Преобразование старой категории в новую с учетом типа звонка"""
        if call_type == "Входящий":
            # Для входящих звонков используем прямое соответствие
            return self.LEGACY_TO_NEW_MAPPING.get(legacy_category, "1")
        elif call_type == "Исходящий":
            # Для исходящих звонков добавляем 13 к номеру категории
            incoming_category = self.LEGACY_TO_NEW_MAPPING.get(legacy_category, "1")
            incoming_num = int(incoming_category)
            outgoing_num = incoming_num + 13
            return str(outgoing_num)
        else:
            # Для неопределенных типов используем входящие
            return self.LEGACY_TO_NEW_MAPPING.get(legacy_category, "1")

    def get_category_name(self, category_num):
        """Получение названия категории по номеру"""
        return self.NEW_CATEGORIES.get(category_num, "НЕИЗВЕСТНАЯ КАТЕГОРИЯ")

    def get_category_group(self, category_name):
        """Получение группы категории"""
        return self.CATEGORY_GROUPS.get(category_name, "Неизвестно")

    def build_call_history_context(self, phone_number, current_filename, processed_calls):
        """Построение контекста предыдущих звонков клиента"""
        if not phone_number or phone_number == "Не распознан":
            return "КОНТЕКСТ: Первый звонок клиента или номер не распознан."

        client_calls = []
        for call in processed_calls:
            if call['Номер телефона'] == phone_number and call['Файл'] != current_filename:
                client_calls.append(call)

        if not client_calls:
            return f"КОНТЕКСТ: Первый звонок клиента {phone_number}."

        client_calls.sort(key=lambda x: (x['Дата'], x['Время']))

        history_text = f"КОНТЕКСТ: История звонков клиента {phone_number} ({len(client_calls)} предыдущих звонков):\n"

        for i, call in enumerate(client_calls[-5:], 1):
            history_text += f"{i}. {call['Дата']} {call['Время']} - {call['Категория']} ({call['Результат']})\n"
            history_text += f"   Обоснование: {call['Обоснование'][:100]}...\n"

        categories = [call['Результат'] for call in client_calls]

        # Проверяем записи (новые символьные коды)
        record_categories = ['IN.BOOK', 'OUT.BOOK', 'IN.FU.BOOK', 'OUT.FU.BOOK', '2', '15']
        has_records = any(cat in record_categories for cat in categories)
        if has_records:
            # Ищем последнюю запись
            for call in reversed(client_calls):
                if call['Результат'] in record_categories:
                    history_text += f"\n⚠️ КЛИЕНТ УЖЕ ЗАПИСАН: {call['Дата']} {call['Время']} в {call['Станция']}\n"
                    break

        # Проверяем запланированные перезвоны (новые символьные коды)
        callback_categories = ['IN.CONS.CB', 'OUT.CONS.CB', '10', '23']
        if any(cat in callback_categories for cat in categories):
            history_text += "ИСТОРИЯ: Были запланированные перезвоны\n"

        # Проверяем многократные отказы/размышления (новые символьные коды)
        think_categories = ['IN.CONS.THINK', 'OUT.CONS.THINK', '4', '17']
        think_count = sum(1 for cat in categories if cat in think_categories)
        if think_count > 1:
            history_text += "ИСТОРИЯ: Многократные отказы/размышления клиента\n"

        return history_text

    def build_training_examples_context(self, transcription):
        """Построение контекста с обучающими примерами"""
        similar_examples = self.training_manager.get_similar_examples(transcription, limit=3)
        
        if not similar_examples:
            return "ОБУЧАЮЩИЕ ПРИМЕРЫ: Нет похожих примеров в базе."
        
        examples_text = "ОБУЧАЮЩИЕ ПРИМЕРЫ (изучи паттерны):\n"
        for i, example in enumerate(similar_examples, 1):
            examples_text += f"{i}. Категория {example['category']}: {self.get_category_description(example['category'])}\n"
            examples_text += f"   Транскрипция: {example['transcription'][:150]}...\n"
            examples_text += f"   Обоснование: {example['reasoning'][:150]}...\n\n"
        
        return examples_text

    def get_category_description(self, category_num):
        """Получение описания категории"""
        return self.NEW_CATEGORIES.get(category_num, "НЕИЗВЕСТНАЯ КАТЕГОРИЯ")

    def _resolve_chat_completions_url(self):
        url = str(self.base_url or "").strip()
        if not url:
            return "https://api.deepseek.com/v1/chat/completions"
        lower_url = url.lower().rstrip("/")
        if lower_url.endswith("/chat/completions"):
            return url
        if lower_url.endswith("/v1"):
            return f"{url.rstrip('/')}/chat/completions"
        return url

    def _append_debug_log(self, event, payload):
        try:
            entry = {
                "ts": datetime.now().isoformat(timespec="seconds"),
                "event": str(event),
                "payload": payload or {},
            }
            self.debug_log_path.parent.mkdir(parents=True, exist_ok=True)
            with self.debug_log_path.open("a", encoding="utf-8") as f:
                f.write(json.dumps(entry, ensure_ascii=False) + "\n")
        except Exception:
            pass

    def _request_llm_with_retries(self, payload, max_retries=3, delay=5):
        url = self._resolve_chat_completions_url()
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json",
        }
        last_error = ""
        for attempt in range(max_retries):
            try:
                resp = requests.post(url, headers=headers, json=payload, timeout=90)
                if resp.status_code == 200:
                    self._append_debug_log(
                        "llm_http_ok",
                        {"url": url, "attempt": attempt + 1, "status_code": resp.status_code},
                    )
                    return resp
                last_error = f"HTTP {resp.status_code}: {resp.text[:500]}"
                self._append_debug_log(
                    "llm_http_error",
                    {
                        "url": url,
                        "attempt": attempt + 1,
                        "status_code": resp.status_code,
                        "body_preview": (resp.text or "")[:1200],
                    },
                )
            except Exception as exc:
                last_error = str(exc)
                self._append_debug_log(
                    "llm_request_exception",
                    {"url": url, "attempt": attempt + 1, "error": str(exc)},
                )
            if attempt < max_retries - 1:
                time.sleep(delay)
        raise RuntimeError(last_error or "LLM request failed")

    def _clean_llm_text(self, text):
        cleaned = str(text or "").strip()
        if cleaned.startswith("```"):
            cleaned = re.sub(r"^```[a-zA-Z0-9_-]*\s*", "", cleaned)
            cleaned = re.sub(r"\s*```$", "", cleaned).strip()
        return cleaned

    def _normalize_reasoning_text(self, text):
        """
        Приводит сырое обоснование LLM к короткому читабельному виду для Excel:
        убирает теги/служебные подписи/markdown и оставляет сам смысл.
        """
        cleaned = self._clean_llm_text(text)
        if not cleaned:
            return ""

        cleaned = cleaned.replace("\r\n", "\n").replace("\r", "\n")

        # Приоритет: если модель вернула явный тег [ОБОСНОВАНИЕ: ...] или [РЕЗУЛЬТАТ: ...],
        # берем только его содержимое.
        tag_match = re.search(r"\[\s*ОБОСНОВАНИЕ\s*:\s*(.*?)\s*\]", cleaned, re.IGNORECASE | re.DOTALL)
        if not tag_match:
            tag_match = re.search(r"\[\s*РЕЗУЛЬТАТ\s*:\s*(.*?)\s*\]", cleaned, re.IGNORECASE | re.DOTALL)
        if tag_match:
            cleaned = tag_match.group(1).strip()
        else:
            label_patterns = [
                r"(?:обоснование|reasoning|explanation|explain|объяснение|пояснение|comment|комментарий|why|here['’]s why)\s*[:\-]\s*(.+)$",
                r"(?:итог|summary|вывод)\s*[:\-]\s*(.+)$",
            ]
            for pattern in label_patterns:
                match = re.search(pattern, cleaned, re.IGNORECASE | re.DOTALL)
                if match:
                    cleaned = match.group(1).strip()
                    break

        # Удаляем типовые служебные блоки, которые модель иногда пишет перед самим объяснением.
        cleaned = re.sub(
            r"(?is)\b(?:классификация звонка|classification(?: of the call)?|тип звонка|call type|"
            r"целевой звонок\??|target call\??|категория|category)\b\s*[:\-]\s*.*?(?=(?:\b(?:обоснование|"
            r"reasoning|explanation|explain|объяснение|пояснение|comment|комментарий|why|here['’]s why)\b\s*[:\-])|$)",
            " ",
            cleaned,
        )

        # Очищаем markdown и лишние маркеры списков.
        cleaned = cleaned.replace("**", " ").replace("__", " ").replace("`", " ")
        cleaned = re.sub(r"(?m)^\s*[-*•]\s*", "", cleaned)
        cleaned = re.sub(r"(?i)^\s*(?:in|out)(?:\.[a-z0-9_]+)+\s+", "", cleaned)
        cleaned = re.sub(
            r"(?i)\b(?:in|out)(?:\.[a-z0-9_]+)+\b\s*(?=(?:обоснование|reasoning|explanation|объяснение|why)\s*:)",
            " ",
            cleaned,
        )
        cleaned = re.sub(r"(?is)\b(?:this conversation|based on the provided|the transcript shows|the call is)\b.*", " ", cleaned)
        cleaned = re.sub(r"\s+", " ", cleaned).strip()
        cleaned = re.sub(r"\s+([,.;:!?])", r"\1", cleaned)

        return cleaned

    def _extract_category_from_text(self, text):
        if not text:
            return None
        # Приоритет: коды новой схемы, если встретились в любом месте ответа.
        keys = sorted(self.NEW_CATEGORIES.keys(), key=len, reverse=True)
        for code in keys:
            if re.search(rf"(?<![A-Z0-9_.]){re.escape(code)}(?![A-Z0-9_.])", text):
                return code
        return None

    def _parse_llm_result(self, raw_text, call_type):
        text = self._clean_llm_text(raw_text)
        if not text:
            return None, None

        # Формат 0: тегированный ответ [КАТЕГОРИЯ:...] + [ОБОСНОВАНИЕ:...]
        tag_cat_match = re.search(r"\[\s*КАТЕГОРИЯ\s*:\s*([^\]]+)\]", text, re.IGNORECASE)
        if tag_cat_match:
            cat = self._extract_category_from_text(tag_cat_match.group(1))
            if cat in self.NEW_CATEGORIES:
                normalized = self._normalize_reasoning_text(text)
                return cat, normalized or self._clean_llm_text(text)

        # Формат 1: CODE|reasoning
        if "|" in text:
            left, right = text.split("|", 1)
            cat = self._extract_category_from_text(left.strip()) or left.strip()
            if cat in self.NEW_CATEGORIES:
                reasoning = self._normalize_reasoning_text(right.strip() or text)
                return cat, reasoning or self._clean_llm_text(text)

        # Формат 2: JSON
        try:
            parsed = json.loads(text)
            if isinstance(parsed, dict):
                cat_raw = (
                    parsed.get("category")
                    or parsed.get("result")
                    or parsed.get("code")
                    or parsed.get("category_num")
                    or ""
                )
                reasoning = (
                    parsed.get("reasoning")
                    or parsed.get("analysis")
                    or parsed.get("explanation")
                    or parsed.get("comment")
                    or ""
                )
                cat = self._extract_category_from_text(str(cat_raw))
                if cat in self.NEW_CATEGORIES:
                    normalized = self._normalize_reasoning_text(str(reasoning or text).strip())
                    return cat, normalized or self._clean_llm_text(text)
        except Exception:
            pass

        # Формат 3: "Категория: CODE" и "Обоснование: ..."
        cat_match = re.search(r"(?:категори[яи]|category|result|код)\s*[:\-]\s*([A-Z][A-Z0-9_.]+)", text, re.IGNORECASE)
        if cat_match:
            cat = self._extract_category_from_text(cat_match.group(1))
            if cat in self.NEW_CATEGORIES:
                normalized = self._normalize_reasoning_text(text)
                return cat, normalized or self._clean_llm_text(text)

        # Формат 4: код где-то в тексте
        cat = self._extract_category_from_text(text)
        if cat in self.NEW_CATEGORIES:
            normalized = self._normalize_reasoning_text(text)
            return cat, normalized or self._clean_llm_text(text)

        # Фолбэк: не роняем классификацию на "формате", выбираем безопасную категорию.
        fallback_cat = "IN.CONS.OTHER" if str(call_type).startswith("Вход") else "OUT.CONS.OTHER"
        normalized = self._normalize_reasoning_text(text)
        suffix = "[auto_fallback: формат ответа не распознан]"
        return fallback_cat, f"{normalized or self._clean_llm_text(text)} {suffix}".strip()

    def classify_call_with_reasoning(self, transcription, call_history_context="", training_examples_context="", call_type="Не определен"):
        """Классификация звонка с получением обоснования, учетом контекста и обучающих примеров"""
        try:
            # Генерируем системный промпт из актуальных правил
            system_prompt = self.rules_manager.generate_system_prompt(
                call_history_context, 
                training_examples_context,
                call_type
            )

            payload = {
                "model": self.model,
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": transcription},
                ],
                "temperature": 0.1,
                "max_tokens": 800,
            }
            transcript_hash = hashlib.md5((transcription or "").encode("utf-8")).hexdigest()
            self._append_debug_log(
                "llm_request",
                {
                    "url": self._resolve_chat_completions_url(),
                    "model": self.model,
                    "call_type": call_type,
                    "transcript_hash": transcript_hash,
                    "transcript_preview": (transcription or "")[:800],
                    "system_prompt_preview": (system_prompt or "")[:1000],
                },
            )
            response = self._request_llm_with_retries(payload, max_retries=3, delay=10)
            data = response.json()
            result = (
                data.get("choices", [{}])[0]
                .get("message", {})
                .get("content", "")
            ).strip()
            self._append_debug_log(
                "llm_response",
                {
                    "transcript_hash": transcript_hash,
                    "response_preview": (result or "")[:2000],
                },
            )
            if not result:
                return "Ошибка", "Пустой ответ от LLM"

            category_num, reasoning = self._parse_llm_result(result, call_type)
            self._append_debug_log(
                "llm_parsed",
                {
                    "transcript_hash": transcript_hash,
                    "parsed_category": category_num,
                    "reasoning_preview": (reasoning or "")[:800],
                },
            )
            if category_num in self.NEW_CATEGORIES:
                return category_num, reasoning
            return "Ошибка", "Некорректный формат ответа"

        except Exception as e:
            self._append_debug_log("llm_classification_exception", {"error": str(e)})
            return "Ошибка", f"API ошибка: {str(e)}"


    def validate_classification_with_context(self, category_num, phone_number, current_filename, processed_calls, transcription, reasoning=""):
        """Расширенная валидация классификации с учетом контекста и обучающих примеров"""

        # ПРИОРИТЕТНОЕ ПРАВИЛО 0: ЛЮБОЙ повторный звонок ОБЯЗАТЕЛЬНО должен быть последующим контактом
        # Проверяем наличие истории звонков от этого клиента
        client_previous_calls = [call for call in processed_calls
                                if call['Номер телефона'] == phone_number and call['Файл'] != current_filename]
        
        if client_previous_calls and phone_number and phone_number != "Не распознан":
            # Если есть история - проверяем, нужно ли переклассифицировать
            # НЕ переклассифицируем, если это уже правильные категории последующего контакта или не целевой
            if category_num not in ['IN.FU.BOOK', 'OUT.FU.BOOK', 'IN.INFO.FU.NOBOOK', 'OUT.INFO.FU.NOBOOK', 'IN.NE', 'OUT.NE']:
                # Определяем направление из текущей категории
                direction_prefix = 'IN.' if category_num.startswith('IN.') or (category_num.isdigit() and int(category_num) <= 13) else 'OUT.'
                
                # Проверяем, была ли ЗАПИСЬ в истории (клиент уже записан)
                has_previous_record = any(
                    call.get('Результат') in ['IN.BOOK', 'OUT.BOOK', 'IN.FU.BOOK', 'OUT.FU.BOOK', '2', '15'] 
                    for call in client_previous_calls
                )
                
                # Если AI классифицировал как новую запись (BOOK)
                if category_num in ['IN.BOOK', 'OUT.BOOK']:
                    if has_previous_record:
                        # Клиент УЖЕ ЗАПИСАН - это не новая запись, а последующий контакт БЕЗ записи
                        correct_category = f'{direction_prefix}INFO.FU.NOBOOK'
                        return correct_category, f'ПОВТОРНЫЙ ЗВОНОК: клиент уже записан - переклассифицировано из {category_num}'
                    else:
                        # Клиент не записан, но звонит не первый раз - последующий контакт С ЗАПИСЬЮ
                        correct_category = f'{direction_prefix}FU.BOOK'
                        return correct_category, f'ПОВТОРНЫЙ ЗВОНОК: новая запись - переклассифицировано из {category_num}'
                
                # Если AI классифицировал как консультацию или другое
                else:
                    # Любая консультация/уточнение при наличии истории - это последующий контакт БЕЗ записи
                    correct_category = f'{direction_prefix}INFO.FU.NOBOOK'
                    return correct_category, f'ПОВТОРНЫЙ ЗВОНОК: консультация - переклассифицировано из {category_num}'

        # Правило 1: Первый звонок не может быть "Последующим контактом"
        if category_num in ['IN.FU.BOOK', 'OUT.FU.BOOK', 'IN.INFO.FU.NOBOOK', 'OUT.INFO.FU.NOBOOK']:
            if not client_previous_calls:
                return 'IN.CONS.OTHER' if category_num.startswith('IN.') else 'OUT.CONS.OTHER', 'ПЕРВЫЙ ЗВОНОК КЛИЕНТА - переклассифицировано в "Консультация"'

        # Правило 1.1: Если клиент УЖЕ ЗАПИСАН - не может быть "Последующий контакт С ЗАПИСЬЮ"
        # Должен быть "Последующий контакт БЕЗ записи"
        if category_num in ['IN.FU.BOOK', 'OUT.FU.BOOK']:
            client_previous_calls = [call for call in processed_calls
                                    if call['Номер телефона'] == phone_number and call['Файл'] != current_filename]

            # Проверяем, есть ли уже запись у клиента
            has_existing_record = any(call.get('Результат') in ['IN.BOOK', 'OUT.BOOK', 'IN.FU.BOOK', 'OUT.FU.BOOK'] for call in client_previous_calls)
            
            if has_existing_record:
                return 'IN.INFO.FU.NOBOOK' if category_num.startswith('IN.') else 'OUT.INFO.FU.NOBOOK', 'КЛИЕНТ УЖЕ ЗАПИСАН - переклассифицировано в "Последующий контакт БЕЗ записи"'

        # Правило 2: Проверка ложной "Записи на сервис"
        if category_num in ['IN.BOOK', 'OUT.BOOK', 'IN.FU.BOOK', 'OUT.FU.BOOK']:
            reasoning_lower = reasoning.lower() if reasoning else ""
            
            not_recorded_phrases = [
                "подумаю", "найду решение", "перезвоню", "если надумает",
                "уже нашел", "спасибо", "уточнить информацию",
                "договоренность о дальнейшем контакте", "конкретной записи не было",
                "не происходит", "отложено", "время на раздумье"
            ]
            
            if any(phrase in reasoning_lower for phrase in not_recorded_phrases):
                return 'IN.CONS.OTHER' if category_num.startswith('IN.') else 'OUT.CONS.OTHER', 'НЕТ КОНКРЕТНОЙ ЗАПИСИ - переклассифицировано в "Консультация"'
            
            subsequent_contact_phrases = [
                "уже записан", "подтверждает запись", "существующая запись",
                "переносит запись", "уточняет детали записи", "завтра ждем",
                "повторный контакт по уже существующей записи", "подтверждение записи",
                "звонили вот сейчас на сегодня записанного", "вас ждем", "машина в сервисе",
                "машина на ремонте", "автомобиль уже у нас", "клиент записан"
            ]
            
            if any(phrase in reasoning_lower for phrase in subsequent_contact_phrases):
                return 'IN.INFO.FU.NOBOOK' if category_num.startswith('IN.') else 'OUT.INFO.FU.NOBOOK', 'ПОДТВЕРЖДЕНИЕ СУЩЕСТВУЮЩЕЙ ЗАПИСИ - переклассифицировано в "Последующий контакт БЕЗ записи"'

        return category_num, None

    def _normalize_station_for_report(self, station_number, station_display):
        """
        Подпись станции для сводных таблиц и колонки «Станция» в Excel.
        Берётся из настроек ЛК («Название для отчётов» и порядок карточек);
        при отсутствии настроек — встроенные краткие имена из STATION_CODES.
        """
        c = str(station_number or "").strip()
        d = str(station_display or "").strip()

        if c in getattr(self, "_report_label_by_code", {}):
            return self._report_label_by_code[c]

        d_clean = d.strip()
        if d_clean in getattr(self, "_report_label_by_display", {}):
            return self._report_label_by_display[d_clean]

        if c and c not in ("Не распознан", "не распознан") and c in self.STATION_CODES:
            disp = str(self.STATION_CODES[c]).strip()
            if disp in getattr(self, "_report_label_by_display", {}):
                return self._report_label_by_display[disp]

        if d_clean in self.ALL_STATIONS:
            return d_clean

        if c and c not in ("Не распознан", "не распознан") and c in self.STATION_CODES:
            if not self._has_user_station_list:
                return self.STATION_CODES[c]
            return "Прочие"

        return "Прочие"

    def _call_type_from_result(self, category_num, fallback_call_type):
        """
        Колонка «Тип звонка» в отчёте должна совпадать с направлением в «Результат» (IN.* / OUT.*).
        Имя файла нередко однозначно не различает направление, а классификация — различает.
        """
        c = str(category_num or "").strip()
        if c.startswith("IN."):
            return "Входящий"
        if c.startswith("OUT."):
            return "Исходящий"
        return fallback_call_type

    def safe_save_excel(self, results, output_file):
        """Безопасное сохранение Excel файла с обработкой случая, когда файл открыт"""
        df = pd.DataFrame(results)

        if not df.empty and 'Обоснование' in df.columns:
            df['Обоснование'] = df['Обоснование'].apply(self._normalize_reasoning_text)
        
        # Краткие названия для отчёта (по коду станции из файла — не зависят от полных имён в ЛК)
        if not df.empty and 'Станция' in df.columns:
            num_col = 'Номер станции' if 'Номер станции' in df.columns else None
            if num_col:
                df['Станция'] = df.apply(
                    lambda r: self._normalize_station_for_report(r.get(num_col), r.get('Станция')),
                    axis=1,
                )
            else:
                df['Станция'] = df['Станция'].apply(
                    lambda d: self._normalize_station_for_report('', d)
                )
        
        # Убираем исключённые станции (Арз / Хлз / Рпб) уже по кратким именам
        if not df.empty and 'Станция' in df.columns:
            df = df[~df['Станция'].isin(self.EXCLUDED_STATIONS)]
        
        if not df.empty and 'Станция' in df.columns:
            stations_order = list(self.ALL_STATIONS)
            if (df['Станция'] == 'Прочие').any() and 'Прочие' not in stations_order:
                stations_order = list(stations_order) + ['Прочие']
        else:
            stations_order = list(self.ALL_STATIONS)
        
        # Создаем сводные таблицы
        target_summary_df = self.create_summary_table(df, stations_order)
        reference_summary_df = self.create_reference_summary_table(df, stations_order)
        
        # Создаем временный файл
        temp_file = None
        try:
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_file_path = temp_file.name
            
            # Сохраняем во временный файл
            with pd.ExcelWriter(temp_file_path, engine='openpyxl') as writer:
                # 1. Сводная таблица (Целевые)
                target_summary_df.to_excel(writer, sheet_name='Сводная таблица (Целевые)', index=False)
                
                # 2. Детальные данные
                df.to_excel(writer, sheet_name='Детальные данные', index=False)
                
                # 3. Сводная таблица (Справочные)
                reference_summary_df.to_excel(writer, sheet_name='Сводная таблица (Справочные)', index=False)
                
                # Применяем форматирование
                self.apply_excel_formatting(writer, df, target_summary_df, reference_summary_df)
            
            # Пытаемся заменить оригинальный файл
            max_attempts = 3
            for attempt in range(max_attempts):
                try:
                    # Если файл существует, удаляем его
                    if os.path.exists(output_file):
                        os.remove(output_file)
                    
                    # Перемещаем временный файл на место оригинального
                    shutil.move(temp_file_path, output_file)
                    print(f"✅ Файл успешно сохранен: {output_file}")
                    return True
                    
                except PermissionError as e:
                    if attempt < max_attempts - 1:
                        print(f"⚠️ Файл {output_file} заблокирован (возможно, открыт в Excel). Попытка {attempt + 1}/{max_attempts}...")
                        time.sleep(2)  # Ждем 2 секунды перед следующей попыткой
                    else:
                        print(f"❌ Не удалось сохранить файл {output_file} после {max_attempts} попыток. Файл заблокирован.")
                        # Сохраняем с другим именем
                        backup_name = output_file.replace('.xlsx', f'_backup_{int(time.time())}.xlsx')
                        shutil.move(temp_file_path, backup_name)
                        print(f"💾 Файл сохранен как резервная копия: {backup_name}")
                        return False
                        
                except Exception as e:
                    print(f"❌ Ошибка при сохранении файла: {e}")
                    return False
            
        except Exception as e:
            print(f"❌ Ошибка при создании временного файла: {e}")
            return False
        finally:
            # Удаляем временный файл если он остался
            if temp_file and os.path.exists(temp_file_path):
                try:
                    os.remove(temp_file_path)
                except:
                    pass

    def get_main_group(self, category_num):
        """Определение главной группы по новой символьной схеме"""
        # Не целевые
        if category_num in ['IN.NE', 'OUT.NE']:
            return "Не целевые"
        # Справочные
        elif category_num in ['IN.INFO.FU.NOBOOK', 'OUT.INFO.FU.NOBOOK']:
            return "Справочные"
        # Целевые (все остальные IN.* и OUT.* кроме NE и INFO)
        elif category_num.startswith('IN.') or category_num.startswith('OUT.'):
            if not category_num.endswith('.NE') and 'INFO' not in category_num:
                return "Целевые"
        
        # Для старых числовых кодов (обратная совместимость)
        if category_num == "1":
            return "Не целевые"
        elif category_num in ["12", "13"]:
            return "Справочные"
        elif category_num in ["2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "14"]:
            return "Целевые"

        return "Не определен"

    def get_target_status(self, category_num):
        """Определение статуса Целевой/Не целевой/Справочные по новой схеме"""
        # Не целевые
        if category_num in ['IN.NE', 'OUT.NE']:
            return "Не целевые"
        # Справочные
        elif category_num in ['IN.INFO.FU.NOBOOK', 'OUT.INFO.FU.NOBOOK']:
            return "Справочные"
        # Целевые (все категории консультаций, записей и обзвонов)
        elif category_num in ['IN.BOOK', 'OUT.BOOK', 'IN.FU.BOOK', 'OUT.FU.BOOK',
                              'OUT.OBZ.BOOK', 'OUT.OBZ.NOBOOK',
                              'IN.CONS.MSG', 'IN.CONS.REDIR', 'IN.CONS.OWN', 'IN.CONS.THINK',
                              'IN.CONS.BUSY', 'IN.CONS.COST', 'IN.CONS.NODO', 'IN.CONS.CB', 'IN.CONS.OTHER',
                              'OUT.CONS.MSG', 'OUT.CONS.REDIR', 'OUT.CONS.OWN', 'OUT.CONS.THINK',
                              'OUT.CONS.BUSY', 'OUT.CONS.COST', 'OUT.CONS.NODO', 'OUT.CONS.CB', 'OUT.CONS.OTHER']:
            return "Целевые"
        
        # Для старых числовых кодов (обратная совместимость)
        category_name = self.get_category_name(category_num)
        return self.get_category_group(category_name)

    def get_recorded_status(self, category_num):
        """Определение статуса Записан/Не записан по новой схеме"""
        # Записан - только категории с записью (BOOK, но не INFO.FU.NOBOOK)
        if category_num in ['IN.BOOK', 'OUT.BOOK', 'IN.FU.BOOK', 'OUT.FU.BOOK', 'OUT.OBZ.BOOK']:
            return "Записан"
        
        # Для старых числовых кодов (обратная совместимость)
        if category_num in ["11", "24"]:
            return "Записан"

        return "Не записан"

    def process_folder(self, input_folder, output_file, context_days=7, progress_callback=None):
        """Обработка папки с файлами транскрипций"""
        results = []
        total_calls = 0

        # Загружаем историю звонков за указанный период только из папки uploads для контекста
        external_history = []
        if context_days > 0:  # Если context_days = 0, не загружаем историю
            try:
                uploads_dir = self.uploads_dir
                frames = []
                if uploads_dir.exists():
                    for f in sorted(uploads_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True):
                        try:
                            try:
                                one_df = pd.read_excel(f, sheet_name='Детальные данные')
                            except Exception:
                                one_df = pd.read_excel(f)
                            if one_df is not None and not one_df.empty:
                                frames.append(one_df)
                        except Exception:
                            continue
                all_df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
                
                if not all_df.empty:
                    # Фильтруем по периоду (не включая текущий день)
                    # Используем импортированный сверху datetime/timedelta, а не локальный импорт
                    today = datetime.now().date()
                    cutoff_date = today - timedelta(days=context_days)
                    
                    # Преобразуем даты для сравнения, обрабатывая ошибки
                    # Используем errors='coerce' чтобы невалидные даты стали NaT (Not a Time)
                    all_df['Дата_parsed'] = pd.to_datetime(all_df['Дата'], format='%d.%m.%Y', errors='coerce').dt.date
                    
                    # Фильтруем только строки с валидными датами
                    valid_dates_mask = all_df['Дата_parsed'].notna()
                    all_df_valid = all_df[valid_dates_mask].copy()
                    
                    if not all_df_valid.empty:
                        filtered_df = all_df_valid[all_df_valid['Дата_parsed'] < today]  # Исключаем текущий день
                        
                        if context_days < 90:  # Если не "все данные" (90 дней = все данные)
                            filtered_df = filtered_df[filtered_df['Дата_parsed'] >= cutoff_date]
                    else:
                        filtered_df = pd.DataFrame()  # Пустой DataFrame, если нет валидных дат
                    
                    # Преобразуем в нужный формат для контекста (НЕ включаем в отчет)
                    if not filtered_df.empty:
                        for _, row in filtered_df.iterrows():
                            call_record = {
                                'Файл': str(row.get('Файл', 'external')),
                                'Номер телефона': str(row.get('Номер телефона', '')),
                                'Дата': str(row.get('Дата', '')),
                                'Время': str(row.get('Время', '')),
                                'Номер станции': str(row.get('Номер станции', '')),
                                'Станция': str(row.get('Станция', '')),
                                'Результат': str(row.get('Результат', '')),
                                'Категория': str(row.get('Категория', '')),
                                'Целевой/Не целевой': str(row.get('Целевой/Не целевой', '')),
                                'Записан/Не записан': str(row.get('Записан/Не записан', '')),
                                'Обоснование': str(row.get('Обоснование', ''))
                            }
                            external_history.append(call_record)
                    
                    print(f"Загружено {len(external_history)} звонков из истории за {context_days} дней для контекста")
                    
            except Exception as e:
                print(f"Ошибка загрузки истории: {e}")

        # Получаем все текстовые файлы в папке
        if not os.path.exists(input_folder):
            raise FileNotFoundError(f"Папка {input_folder} не найдена")
        
        text_files = [f for f in os.listdir(input_folder) if f.endswith('.txt')]
        
        if not text_files:
            raise ValueError(f"В папке {input_folder} не найдено .txt файлов")

        # Сортируем файлы по дате и времени звонка, извлечённым из имени файла
        def _sort_key_by_datetime(filename):
            try:
                _, call_date, call_time, _, _, _ = self.extract_file_info(filename)
                dt = datetime.strptime(f"{call_date} {call_time}", "%d.%m.%Y %H:%M")
                return (dt, filename)
            except Exception:
                # Если не удалось распознать дату/время — отправляем в начало, сортируем по имени
                return (datetime.min, filename)

        text_files.sort(key=_sort_key_by_datetime)

        print(f"Найдено {len(text_files)} файлов для обработки...")

        for i, filename in enumerate(text_files, 1):
            if progress_callback:
                progress_callback(i, len(text_files), f"Обработка {filename}")
            
            print(f"Обработка файла {i}/{len(text_files)}: {filename}")
            
            # Извлекаем информацию из имени файла
            phone_number, call_date, call_time, station_number, station_name, call_type = self.extract_file_info(filename)
            
            # Читаем транскрипцию
            try:
                with open(os.path.join(input_folder, filename), 'r', encoding='utf-8') as f:
                    transcription = f.read().strip()
            except UnicodeDecodeError:
                try:
                    with open(os.path.join(input_folder, filename), 'r', encoding='cp1251') as f:
                        transcription = f.read().strip()
                except:
                    transcription = "Ошибка чтения файла"
            
            if not transcription or transcription == "Ошибка чтения файла":
                results.append({
                    'Файл': filename,
                    'Номер телефона': phone_number,
                    'Дата': call_date,
                    'Время': call_time,
                    'Номер станции': station_number,
                    'Станция': station_name,
                    'Тип звонка': call_type,
                    'Результат': 'Пропущен',
                    'Категория': 'Пустой файл',
                    'Целевой/Не целевой': 'Не определен',
                    'Записан/Не записан': 'Не определен',
                    'Обоснование': 'Файл пустой или не читается'
                })
                continue
            
            total_calls += 1
            
            # Объединяем текущие результаты с внешней историей для контекста
            combined_history = external_history + results

            # Строим контекст предыдущих звонков клиента
            call_history_context = self.build_call_history_context(phone_number, filename, combined_history)
            
            # Строим контекст с обучающими примерами
            training_examples_context = self.build_training_examples_context(transcription)

            # Классифицируем звонок с учетом контекста и обучающих примеров
            category_num, reasoning = self.classify_call_with_reasoning(
                transcription, call_history_context, training_examples_context, call_type
            )

            # Дополнительная валидация с учетом всей истории (внешняя история + текущая сессия).
            # Здесь применяется логика, которая уже умеет:
            # - отличать последующий контакт С записью от БЕЗ записи
            # - учитывать, что машина уже в сервисе / клиент уже записан
            # - не допускать "последующий контакт" для первого звонка клиента
            validated_category, correction_reason = self.validate_classification_with_context(
                category_num,
                phone_number,
                filename,
                combined_history,
                transcription,
                reasoning,
            )
            if validated_category != category_num:
                self._append_debug_log(
                    "context_validation_correction",
                    {
                        "filename": filename,
                        "phone_number": phone_number,
                        "old_category": category_num,
                        "new_category": validated_category,
                        "reason": correction_reason,
                    },
                )
                category_num = validated_category

            # Получаем описание категории
            category_desc = self.get_category_description(category_num)

            # Определяем статусы
            target_status = self.get_target_status(category_num)
            recorded_status = self.get_recorded_status(category_num)

            report_call_type = self._call_type_from_result(category_num, call_type)

            # Добавляем задержку между запросами
            time.sleep(1)

            results.append({
                'Файл': filename,
                'Номер телефона': phone_number,
                'Дата': call_date,
                'Время': call_time,
                'Номер станции': station_number,
                'Станция': station_name,
                'Тип звонка': report_call_type,
                'Результат': category_num,
                'Категория': category_desc,
                'Целевой/Не целевой': target_status,
                'Записан/Не записан': recorded_status,
                'Обоснование': reasoning
            })
        
        # Обновляем метрики в базе данных
        today = datetime.now().strftime('%Y-%m-%d')
        self.training_manager.update_daily_metrics(today, total_calls, total_calls, 0)
        
        # Сохраняем результаты в Excel
        self.save_results_to_excel(results, output_file)
        
        return results, 0, total_calls

    def save_results_to_excel(self, results, output_file):
        """Сохранение результатов в Excel с форматированием"""
        # Используем безопасное сохранение
        success = self.safe_save_excel(results, output_file)
        if not success:
            print(f"⚠️ Внимание: Файл {output_file} не удалось перезаписать. Проверьте, не открыт ли он в Excel.")

    def create_summary_table(self, results_df, stations_order=None):
        """Создание сводной таблицы по станциям и причинам согласно новой схеме.
        
        stations_order: список названий станций в нужном порядке. 
        Если не передан, используется self.ALL_STATIONS или уникальные станции из данных.
        """
        summary_data = []
        
        # Определяем порядок станций
        if stations_order is not None:
            stations = stations_order
        elif not results_df.empty and 'Станция' in results_df.columns:
            stations = sorted(results_df['Станция'].unique())
        else:
            stations = self.ALL_STATIONS
        
        # ЦЕЛЕВЫЕ ЗВОНКИ (объединенные входящие и исходящие)
        summary_data.append(['Целевые', ''] + [''] * len(stations))
        
        # Записанные (входящие + исходящие)
        recorded_calls = results_df[results_df['Записан/Не записан'] == 'Записан']
        recorded_counts = recorded_calls['Станция'].value_counts().reindex(stations, fill_value=0)
        summary_data.append(['  - Записанные', recorded_counts.sum()] + recorded_counts.tolist())
        
        # Не записанные - сумма всех консультаций
        consultation_categories = {
            'Консультация - ПЕРЕШЛИ В МЕССЕНДЖЕР': ['IN.CONS.MSG', 'OUT.CONS.MSG'],
            'Консультация - ПЕРЕАДРЕСАЦИЯ': ['IN.CONS.REDIR', 'OUT.CONS.REDIR'],
            'Консультация - СВОИ ЗАПЧАСТИ': ['IN.CONS.OWN', 'OUT.CONS.OWN'],
            'Консультация - ПОДУМАЕТ/ОТКАЗ': ['IN.CONS.THINK', 'OUT.CONS.THINK'],
            'Консультация - НЕТ ВРЕМЕНИ/ЗАНЯТО': ['IN.CONS.BUSY', 'OUT.CONS.BUSY'],
            'Консультация - ВЫСОКАЯ СТОИМОСТЬ': ['IN.CONS.COST', 'OUT.CONS.COST'],
            'Консультация - НЕ ВЫПОЛНЯЕМ РАБОТЫ': ['IN.CONS.NODO', 'OUT.CONS.NODO'],
            'Консультация - ЗАПЛАНИРОВАН ПЕРЕЗВОН': ['IN.CONS.CB', 'OUT.CONS.CB'],
            'Консультация - Общая': ['IN.CONS.OTHER', 'OUT.CONS.OTHER']
        }
        
        # Собираем все коды консультаций
        all_consultation_codes = []
        for cat_codes in consultation_categories.values():
            all_consultation_codes.extend(cat_codes)
        
        # Не записанные = все консультации
        not_recorded_calls = results_df[results_df['Результат'].isin(all_consultation_codes)]
        not_recorded_counts = not_recorded_calls['Станция'].value_counts().reindex(stations, fill_value=0)
        summary_data.append(['  - Не записанные', not_recorded_counts.sum()] + not_recorded_counts.tolist())
        
        # Детализация по категориям консультаций
        for cat_name, cat_codes in consultation_categories.items():
            cat_calls = results_df[results_df['Результат'].isin(cat_codes)]
            cat_counts = cat_calls['Станция'].value_counts().reindex(stations, fill_value=0)
            summary_data.append([f'    {cat_name}', cat_counts.sum()] + cat_counts.tolist())
        
        columns = ['Категория', 'Кол-во'] + stations
        summary_df = pd.DataFrame(summary_data, columns=columns)
        
        return summary_df

    def create_reference_summary_table(self, results_df, stations_order=None):
        """Создание сводной таблицы для справочных звонков согласно новой схеме.
        
        stations_order: список названий станций в нужном порядке. 
        Если не передан, используется self.ALL_STATIONS или уникальные станции из данных.
        """
        # Разделяем справочные звонки по типам
        incoming_reference = results_df[(results_df['Целевой/Не целевой'] == 'Справочные') & (results_df['Тип звонка'] == 'Входящий')]
        outgoing_reference = results_df[(results_df['Целевой/Не целевой'] == 'Справочные') & (results_df['Тип звонка'] == 'Исходящий')]
        
        summary_data = []
        
        # Определяем порядок станций
        if stations_order is not None:
            stations = stations_order
        elif not results_df.empty and 'Станция' in results_df.columns:
            stations = sorted(results_df['Станция'].unique())
        else:
            stations = self.ALL_STATIONS
        
        # ВХОДЯЩИЕ СПРАВОЧНЫЕ
        summary_data.append(['=== ВХОДЯЩИЕ СПРАВОЧНЫЕ ===', ''] + [''] * len(stations))
        
        total_incoming_ref = incoming_reference['Станция'].value_counts().reindex(stations, fill_value=0)
        summary_data.append(['Всего входящих справочных', total_incoming_ref.sum()] + total_incoming_ref.tolist())
        
        # Входящие справочные категории с новыми символьными кодами
        incoming_ref_categories = {
            "IN.INFO.FU.NOBOOK": "Последующий контакт без записи"
        }
        
        for cat_id, cat_name in incoming_ref_categories.items():
            cat_calls = incoming_reference[incoming_reference['Результат'] == cat_id]
            cat_counts = cat_calls['Станция'].value_counts().reindex(stations, fill_value=0)
            summary_data.append([f'  {cat_name}', cat_counts.sum()] + cat_counts.tolist())
        
        # ИСХОДЯЩИЕ СПРАВОЧНЫЕ
        summary_data.append(['=== ИСХОДЯЩИЕ СПРАВОЧНЫЕ ===', ''] + [''] * len(stations))
        
        total_outgoing_ref = outgoing_reference['Станция'].value_counts().reindex(stations, fill_value=0)
        summary_data.append(['Всего исходящих справочных', total_outgoing_ref.sum()] + total_outgoing_ref.tolist())
        
        # Исходящие справочные категории с новыми символьными кодами
        outgoing_ref_categories = {
            "OUT.INFO.FU.NOBOOK": "Последующий контакт без записи",
            "OUT.OBZ.BOOK": "Обзвон - С записью",
            "OUT.OBZ.NOBOOK": "Обзвон - Без записи"
        }
        
        for cat_id, cat_name in outgoing_ref_categories.items():
            cat_calls = outgoing_reference[outgoing_reference['Результат'] == cat_id]
            cat_counts = cat_calls['Станция'].value_counts().reindex(stations, fill_value=0)
            summary_data.append([f'  {cat_name}', cat_counts.sum()] + cat_counts.tolist())
        
        # ИТОГО СПРАВОЧНЫХ
        summary_data.append(['=== ИТОГО СПРАВОЧНЫХ ===', ''] + [''] * len(stations))
        
        all_reference = results_df[results_df['Целевой/Не целевой'] == 'Справочные']
        total_all_ref = all_reference['Станция'].value_counts().reindex(stations, fill_value=0)
        summary_data.append(['Всего справочных звонков', total_all_ref.sum()] + total_all_ref.tolist())
        
        columns = ['Категория', 'Кол-во'] + stations
        reference_summary_df = pd.DataFrame(summary_data, columns=columns)
        
        return reference_summary_df

    def apply_excel_formatting(self, writer, df, target_summary_df, reference_summary_df):
        """Применение форматирования к Excel файлу"""
        # Цвета для групп
        colors = {
            'Не целевой': PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid'),
            'Целевой': PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid'),
            'Справочные': PatternFill(start_color='CCCCFF', end_color='CCCCFF', fill_type='solid')
        }
        
        # Форматирование детальных данных
        worksheet = writer.sheets['Детальные данные']
        
        # Заголовки
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
        
        # Цветовое выделение строк по группам
        group_column_idx = df.columns.get_loc('Целевой/Не целевой') + 1
        
        for row_idx in range(2, len(df) + 2):
            group_value = df.iloc[row_idx - 2]['Целевой/Не целевой']
            
            if group_value in colors:
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = colors[group_value]
        
        # Автоматическое изменение ширины столбцов для детальных данных
        self._auto_adjust_column_width(worksheet, df)
        
        # Форматирование сводной таблицы (Целевые)
        self._format_target_summary_table(writer, target_summary_df)
        
        # Форматирование сводной таблицы (Справочные)
        self._format_reference_summary_table(writer, reference_summary_df)
    
    def _format_target_summary_table(self, writer, target_summary_df):
        """Форматирование сводной таблицы целевых звонков"""
        worksheet = writer.sheets['Сводная таблица (Целевые)']
        
        # Цвета согласно изображению
        light_green = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')  # Светло-зеленый
        light_red = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')    # Светло-красный  
        light_orange = PatternFill(start_color='FFE4CC', end_color='FFE4CC', fill_type='solid') # Светло-оранжевый
        
        # Заголовки
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        for col_idx in range(1, len(target_summary_df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
        
        # Применяем цвета к строкам согласно изображению
        for row_idx in range(2, len(target_summary_df) + 2):
            category = target_summary_df.iloc[row_idx - 2]['Категория']
            
            # Определяем цвет строки
            fill_color = None
            if 'Записанные' in category:
                fill_color = light_green
            elif 'Не записанные' in category:
                fill_color = light_red
            elif 'ПЕРЕШЛИ В МЕССЕНДЖЕР' in category:
                fill_color = light_green
            elif 'ПЕРЕАДРЕСАЦИЯ' in category:
                fill_color = light_green
            elif 'ЗАПЛАНИРОВАН ПЕРЕЗВОН' in category:
                fill_color = light_green
            elif 'Общая' in category:
                fill_color = light_red
            elif any(keyword in category for keyword in ['СВОИ ЗАПЧАСТИ', 'ПОДУМАЕТ/ОТКАЗ', 'НЕТ ВРЕМЕНИ/ЗАНЯТО', 'ВЫСОКАЯ СТОИМОСТЬ', 'НЕ ВЫПОЛНЯЕМ РАБОТЫ']):
                fill_color = light_orange
            
            # Применяем цвет ко всей строке
            if fill_color:
                for col_idx in range(1, len(target_summary_df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = fill_color
        
        # Применяем общую табличную раскладку (границы, выравнивание, ширина)
        self._apply_summary_table_layout(worksheet, target_summary_df)
    
    def _format_reference_summary_table(self, writer, reference_summary_df):
        """Форматирование сводной таблицы справочных звонков"""
        worksheet = writer.sheets['Сводная таблица (Справочные)']
        
        # Заголовки
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        for col_idx in range(1, len(reference_summary_df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
        
        # Применяем общую табличную раскладку
        self._apply_summary_table_layout(worksheet, reference_summary_df)

    def _apply_summary_table_layout(self, worksheet, dataframe):
        """Единый стиль для сводных таблиц: границы, центровка и узкие столбцы"""
        thin_side = Side(style='thin', color='000000')
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        
        total_columns = len(dataframe.columns)
        column_widths = [40, 10] + [6] * max(0, total_columns - 2)
        
        for idx in range(total_columns):
            column_letter = get_column_letter(idx + 1)
            width = column_widths[idx] if idx < len(column_widths) else 6
            worksheet.column_dimensions[column_letter].width = width
        
        max_row = len(dataframe) + 1
        for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=total_columns):
            for cell in row:
                cell.border = thin_border
                if cell.column == 1:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
                else:
                    cell.alignment = center_alignment
    
    def _auto_adjust_column_width(self, worksheet, df):
        """Автоматическое изменение ширины столбцов для лучшей читаемости"""
        for col_idx, column in enumerate(df.columns, 1):
            # Находим максимальную длину содержимого в столбце
            max_length = 0
            
            # Проверяем заголовок
            max_length = max(max_length, len(str(column)))
            
            # Проверяем все значения в столбце
            for value in df[column]:
                max_length = max(max_length, len(str(value)))
            
            # Устанавливаем ширину столбца с небольшим запасом
            # Минимальная ширина 10, максимальная 50 для читаемости
            adjusted_width = min(max(max_length + 2, 10), 50)
            
            # Получаем буквенное обозначение столбца (A, B, C, ...)
            column_letter = worksheet.cell(row=1, column=col_idx).column_letter
            
            # Устанавливаем ширину столбца
            worksheet.column_dimensions[column_letter].width = adjusted_width
